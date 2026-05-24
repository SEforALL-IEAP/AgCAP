import dash
from dash import dcc, html, dash_table, Input, Output, State, callback_context, clientside_callback, MATCH, ALL
import dash.dash_table.FormatTemplate as FormatTemplate
from dash.dash_table.Format import Format, Scheme
import pandas as pd
import geopandas as gpd
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
import base64
import io
import uuid
import json
import os
import tempfile
from datetime import datetime
import requests
import random
from shapely import wkt
import dash_bootstrap_components as dbc

# Add scripts directory to path for translation_loader import
import sys
from pathlib import Path
scripts_dir = Path(__file__).parent
if str(scripts_dir) not in sys.path:
    sys.path.insert(0, str(scripts_dir))

from translation_loader import t, get_column_translation, get_spider_label_patterns, COLUMN_TRANSLATIONS, load_column_translations

# =============================================================================
# PARAMETERS — edit this section to adapt to a different country or dataset
# =============================================================================

COUNTRY_CODE  = 'MOZ'
FIGURE_TITLE  = 'AgCAP Mozambique'
DEFAULT_COLUMN = 'Ag Cooling Demand ALL Markets'

# Folder that create_app() / standalone run searches for the input dataset.
# Notebooks ignore this — they load data themselves and pass it to agcap_explorer().
# create_app() picks the most recently modified .gpkg in this folder automatically.
DATA_FOLDER = 'data/processed/input_analyzed'

# Column-tooltip dictionary file inside docs/ (xlsx with EN + PT sheets, or omit to disable tooltips).
# Change this filename when adapting for a different country.
DICTIONARY_FILE = 'settlements_data_dictionary_MOZ.xlsx'

# Map defaults
DEFAULT_MAP_CENTER = {'lat': -18.66, 'lon': 35.52}
DEFAULT_MAP_ZOOM   = 5

# Administrative hierarchy for population sunburst (top → bottom).
# Must match the exact column names in the dataset.
ADMIN_LEVELS = ['Province', 'District', 'Posto']

# Column used for population totals in the sunburst chart
POPULATION_COLUMN = 'Population 2025'

# Column used for electrification sunburst
ELECTRIFICATION_COLUMN = 'Electrification status (IEP study)'

# Analysis chart column names (English — translations applied at display time).
# All names must exist as columns in the dataset.
AG_METRICS = [
    'Ag Cooling Demand Export Market', 'Ag Cooling Demand National Market',
    'Ag Cooling Demand Fresh Markets', 'Ag Cooling Demand ALL Markets'
]
FISH_METRICS = [
    'Fish Cooling Demand Export Market', 'Fish Cooling Demand National Market',
    'Fish Cooling Demand Fresh Markets', 'Fish Cooling Demand ALL Markets'
]
PROD_METRICS = [
    'Banana production', 'Cassava production', 'Citrus production',
    'Cashew Nut production', 'Other Roots & Tubers production',
    'Tropical Fruit production', 'Vegetables production',
    'Potato production', 'Sweet Potato production',
    'Temperate Fruit production', 'Tomato production'
]

# =============================================================================

# Load country-specific column translations
load_column_translations(COUNTRY_CODE)

# Load column tooltips once at module level (keyed by lang code)
_COLUMN_TOOLTIPS = {}
try:
    import openpyxl as _openpyxl
    _wb = _openpyxl.load_workbook(scripts_dir.parent / 'docs' / DICTIONARY_FILE)
    _en_names = [r[1] for r in _wb['EN'].iter_rows(min_row=2, values_only=True) if r[1]]
    for _sheet, _lcode in [('EN', 'en'), ('PT', 'pt')]:
        _descs = [r[2] for r in _wb[_sheet].iter_rows(min_row=2, values_only=True) if r[1]]
        _COLUMN_TOOLTIPS[_lcode] = {n: d for n, d in zip(_en_names, _descs) if d}
    del _wb, _en_names, _sheet, _lcode, _descs
except Exception:
    pass

MAPBOX_TOKEN = os.environ.get("MAPBOX_TOKEN", "")
external_scripts = [
    {'src': 'https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js'}
]
# External stylesheet for Google Fonts (Mulish & Barlow Condensed) and Icons
external_stylesheets = [
    {
        'href': 'https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=Mulish:wght@300;400;700&display=swap',
        'rel': 'stylesheet'
    },
    dbc.themes.BOOTSTRAP,
    {
        'href': 'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css',
        'rel': 'stylesheet'
    }
]

# --- CSS Gradient Mappings for Legend ---
SCALE_MAPPINGS = {
    'Turbo': 'linear-gradient(to right, #30123b, #46f7f6, #a4fc3c, #e34f18, #7a0403)',
    'Jet': 'linear-gradient(to right, #00008f, #0000ff, #00ffff, #00ff00, #ffff00, #ff0000, #800000)',
    'Electric': 'linear-gradient(to right, #000000, #472265, #265c79, #1e9274, #7cba3d, #f8f74e)',
    'Rainbow': 'linear-gradient(to right, #96005a, #0000c8, #0084a9, #007a00, #b2b200, #bd7b00, #c80000)',
    'Portland': 'linear-gradient(to right, #0c3383, #0a88ba, #f2d338, #f28f38, #d91e18)',
    'Viridis': 'linear-gradient(to right, #440154, #31688e, #35b779, #fde725)',
    'Plasma': 'linear-gradient(to right, #0d0887, #cc4778, #f0f921)',
    'Hot': 'linear-gradient(to right, #0b0000, #440000, #7d0000, #b60000, #ef0000, #ff4900, #ff9000, #ffce00, #ffff00, #ffffff)',
    'Blackbody': 'linear-gradient(to right, #000000, #e60000, #e6d200, #ffffff, #a0c8ff)',
    'RdBu': 'linear-gradient(to right, #053061, #2166ac, #4393c3, #92c5de, #d1e5f0, #f7f7f7, #fddbc7, #f4a582, #d6604d, #b2182b, #67001f)',
    'Earth': 'linear-gradient(to right, #000030, #1c0c66, #2d3635, #7d603e, #dcc67e, #f1f2e8)'
}

# --- Global Data Manager Class (ROBUST SESSION VERSION) ---
class DataManager:
    def __init__(self):
        # Structure: { 'user_uuid_string': { 'layers': {}, 'metadata': [] } }
        self.sessions = {}
        self.session_timestamps = {}  # {session_id: last_access_datetime}

        # Store the startup data here, to be copied to new users later
        self.default_context = None
        self.color_cycle = ['#583CA5', '#e74c3c', '#2ecc71', '#3498db', '#9b59b6', '#f1c40f', '#e67e22', '#1abc9c']

    def get_user_data(self, session_id):
        """Helper to get the specific dictionary for a user."""
        if session_id not in self.sessions:
            # Initialize empty container for this user
            self.sessions[session_id] = {'layers': {}, 'metadata': [], 'geojson': {}}
        else:
            # Backfill geojson key for sessions created before polygon support
            if 'geojson' not in self.sessions[session_id]:
                self.sessions[session_id]['geojson'] = {}
        self.session_timestamps[session_id] = datetime.now()
        return self.sessions[session_id]

    def set_default_context(self, df, name, default_column):
        """Saves the initial dataframe as a template for new users."""
        self.default_context = {
            'df': df,
            'name': name,
            'default_column': default_column
        }

    def initialize_user_session(self, session_id):
        """Hydrates a new user session with the default context."""
        if not session_id: return
        user_data = self.get_user_data(session_id)
        
        # Only initialize if empty (prevent resetting on every callback)
        if not user_data['layers'] and self.default_context:
            self.add_layer(
                session_id, 
                self.default_context['df'], 
                self.default_context['name'], 
                is_primary=True,
                initial_column=self.default_context['default_column']
            )

    def _generate_column_mapping(self, columns):
        """Generates a map of {LongName: cX} for serialization efficiency."""
        # Columns that must NOT be renamed so backend logic continues to find them
        exclude = ['ID', 'row_uuid', 'orig_uuid', 'Notes', 'Score']
        col_map = {}
        counter = 0
        for col in columns:
            if col in exclude:
                col_map[col] = col
            else:
                col_map[col] = f"c{counter}"
                counter += 1
        return col_map

    def add_layer(self, session_id, df, name, filename="", is_primary=False, initial_column=None):
        if not session_id: return None
        user_data = self.get_user_data(session_id) # <--- Access User Specific Data
        
        layer_id = str(uuid.uuid4())
        
        # --- ROBUSTNESS: Generate Internal IDs (optimized - uses index instead of UUIDs) ---
        df = df.copy(deep=False)  # Shallow copy shares column data between users
        # Deep-copy editable columns to prevent cross-session contamination
        if 'Notes' in df.columns:
            df['Notes'] = df['Notes'].copy()
        if 'Score' in df.columns:
            df['Score'] = df['Score'].copy()
        df['row_uuid'] = df.index.astype(str)
        df.set_index('row_uuid', inplace=True)
        
        existing_id_col = next((c for c in df.columns if c.lower() in ['id', 'fid', 'name']), None)
        if not existing_id_col:
            df.insert(0, 'ID', range(1, len(df) + 1))
        
        numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        
        if name == "Settlements":
            def_size, def_opacity, def_color = 4, 0.8, '#FBB800'
        else:
            def_size, def_opacity, def_color = 15, 1.0, random.choice(self.color_cycle)
        
        # Save to user_data
        user_data['layers'][layer_id] = df
        
        # --- NEW: Generate Column Map ---
        col_mapping = self._generate_column_mapping(df.columns)

        meta = {
            'id': layer_id,
            'name': name if name else f"Layer {len(user_data['metadata'])+1}",
            'filename': filename,
            'visible': True,
            'is_primary': is_primary,
            'size': def_size,
            'opacity': def_opacity,
            'color_mode': 'single', 
            'color_column': 'Single Color',
            'color_scale': 'Portland',
            'single_color_hex': def_color, 
            'columns': df.columns.tolist(),
            'numeric_columns': numeric_cols,
            'column_map': col_mapping  # <--- Stored here
        }

        # Apply specific color settings if provided (e.g. for the primary layer)
        if initial_column:
            meta['color_mode'] = 'column'
            meta['color_column'] = initial_column

        user_data['metadata'].insert(0, meta)
        return layer_id

    def add_polygon_layer(self, session_id, gdf, name, filename=""):
        """Add a polygon/multipolygon layer. Stores GeoJSON geometry separately from attribute DataFrame."""
        if not session_id: return None
        user_data = self.get_user_data(session_id)

        layer_id = str(uuid.uuid4())

        # Build attribute DataFrame (no geometry) with row_uuid injected
        attr_df = pd.DataFrame(gdf.drop(columns='geometry'))
        # Store centroids for hover trace
        attr_df['centroid_lat'] = gdf.geometry.centroid.y.values
        attr_df['centroid_lon'] = gdf.geometry.centroid.x.values
        attr_df = attr_df.copy(deep=False)
        attr_df['row_uuid'] = attr_df.index.astype(str)
        attr_df.set_index('row_uuid', inplace=True)

        existing_id_col = next((c for c in attr_df.columns if c.lower() in ['id', 'fid', 'name']), None)
        if not existing_id_col:
            attr_df.insert(0, 'ID', range(1, len(attr_df) + 1))

        # Type conversion
        for col in attr_df.columns:
            if pd.api.types.is_numeric_dtype(attr_df[col]):
                attr_df[col] = pd.to_numeric(attr_df[col], errors='coerce').astype('float64')
            else:
                attr_df[col] = attr_df[col].fillna('').apply(str)

        numeric_cols = [c for c in attr_df.columns if pd.api.types.is_numeric_dtype(attr_df[c])
                        and c not in ('centroid_lat', 'centroid_lon')]
        all_cols = [c for c in attr_df.columns if c not in ('centroid_lat', 'centroid_lon')]
        cat_cols = [c for c in all_cols if c not in numeric_cols and c not in ('ID', 'row_uuid')]

        # Build GeoJSON with row_uuid injected into feature properties for JS join
        gdf_export = gdf.copy()
        gdf_export['row_uuid'] = gdf_export.index.astype(str)
        geojson_dict = json.loads(gdf_export.to_json())

        user_data['layers'][layer_id] = attr_df
        user_data['geojson'][layer_id] = geojson_dict

        col_mapping = self._generate_column_mapping(attr_df.columns)
        def_color = random.choice(self.color_cycle)

        meta = {
            'id': layer_id,
            'name': name if name else f"Layer {len(user_data['metadata'])+1}",
            'filename': filename,
            'visible': True,
            'is_primary': False,
            'layer_type': 'polygon',
            'opacity': 0.7,
            'fill_color': '#3498db',
            'fill_opacity': 0.4,
            'outline_color': def_color,
            'outline_weight': 2,
            'outline_opacity': 1.0,
            'color_mode': 'single',
            'color_column': 'Single Color',
            'color_scale': 'Portland',
            'single_color_hex': def_color,
            'columns': all_cols,
            'numeric_columns': numeric_cols,
            'cat_columns': cat_cols,
            'column_map': col_mapping,
            # point layer fields set to neutral defaults so callbacks don't crash
            'size': 8,
        }
        user_data['metadata'].insert(0, meta)
        return layer_id

    def move_layer(self, session_id, layer_id, direction):
        if not session_id: return
        user_data = self.get_user_data(session_id)
        idx = next((i for i, item in enumerate(user_data['metadata']) if item["id"] == layer_id), -1)
        if idx == -1: return
        
        if direction == 'up' and idx > 0:
            user_data['metadata'][idx], user_data['metadata'][idx-1] = user_data['metadata'][idx-1], user_data['metadata'][idx]
        elif direction == 'down' and idx < len(user_data['metadata']) - 1:
            user_data['metadata'][idx], user_data['metadata'][idx+1] = user_data['metadata'][idx+1], user_data['metadata'][idx]

    def delete_layer(self, session_id, layer_id):
        if not session_id: return
        user_data = self.get_user_data(session_id)
        user_data['metadata'] = [l for l in user_data['metadata'] if l['id'] != layer_id]
        if layer_id in user_data['layers']:
            del user_data['layers'][layer_id]
        user_data.get('geojson', {}).pop(layer_id, None)

    def toggle_visibility(self, session_id, layer_id):
        if not session_id: return
        user_data = self.get_user_data(session_id)
        for layer in user_data['metadata']:
            if layer['id'] == layer_id:
                layer['visible'] = not layer['visible']
                break

    def update_setting(self, session_id, layer_id, key, value):
        if not session_id: return
        user_data = self.get_user_data(session_id)
        for layer in user_data['metadata']:
            if layer['id'] == layer_id:
                layer[key] = value
                break

    def get_df(self, session_id, layer_id):
        if not session_id: return None
        return self.get_user_data(session_id)['layers'].get(layer_id)

    def get_metadata(self, session_id):
        """Safe getter for metadata"""
        if not session_id: return []
        return self.get_user_data(session_id)['metadata']

    # --- Session Management ---
    def get_session_state(self, session_id):
        if not session_id: return {}
        user_data = self.get_user_data(session_id)
        
        serialized_layers = {}
        for lid, df in user_data['layers'].items():
            serialized_layers[lid] = df.to_json(orient='split', date_format='iso')
            
        # Include polygon GeoJSON (stored as dicts, serialize to JSON strings)
        serialized_geojson = {}
        for lid, gj in user_data.get('geojson', {}).items():
            serialized_geojson[lid] = json.dumps(gj)

        return {
            'metadata': user_data['metadata'],
            'layers': serialized_layers,
            'geojson': serialized_geojson,
            'version': '1.1'
        }

    def load_session_state(self, session_id, session_data):
        if not session_id: return False, "No session ID"
        try:
            user_data = self.get_user_data(session_id)
            user_data['metadata'] = session_data.get('metadata', [])
            user_data['layers'] = {}
            for lid, df_json in session_data.get('layers', {}).items():
                df = pd.read_json(io.StringIO(df_json), orient='split')
                df.index = df.index.astype(str)
                user_data['layers'][lid] = df
            # Restore polygon GeoJSON
            user_data['geojson'] = {}
            for lid, gj_str in session_data.get('geojson', {}).items():
                user_data['geojson'][lid] = json.loads(gj_str)
            return True, "Session loaded successfully."
        except Exception as e:
            return False, f"Error loading session: {str(e)}"

    def create_composite_index(self, session_id, layer_id, column_name, selected_columns, weights, method='minmax', negative_effects=None):
        """Create a composite index column based on weighted combination of numeric columns."""
        if not session_id or not layer_id:
            return False

        user_data = self.get_user_data(session_id)
        if layer_id not in user_data['layers']:
            return False

        df = user_data['layers'][layer_id]
        if negative_effects is None:
            negative_effects = {}

        if not selected_columns or not weights:
            return False

        total_weight = sum(weights.values())
        if total_weight == 0:
            return False
        normalized_weights = {col: w / total_weight for col, w in weights.items()}

        composite = pd.Series(0.0, index=df.index)

        for col in selected_columns:
            if col not in df.columns:
                continue

            data = df[col].copy()
            data = data.fillna(data.mean() if not data.isna().all() else 0)

            if method == 'minmax':
                col_min = data.min()
                col_max = data.max()
                if col_max > col_min:
                    normalized = (data - col_min) / (col_max - col_min)
                else:
                    normalized = pd.Series(0.5, index=data.index)
            else:
                col_mean = data.mean()
                col_std = data.std()
                if col_std > 0:
                    normalized = (data - col_mean) / col_std
                    normalized = normalized.clip(-3, 3)
                    normalized = (normalized + 3) / 6
                else:
                    normalized = pd.Series(0.5, index=data.index)

            if negative_effects.get(col, False):
                normalized = 1.0 - normalized

            composite += normalized * normalized_weights[col]

        df[column_name] = (composite * 100).round(2)

        meta = next((m for m in user_data['metadata'] if m['id'] == layer_id), None)
        if meta:
            if column_name not in meta.get('columns', []):
                meta['columns'].append(column_name)
            if column_name not in meta.get('numeric_columns', []):
                meta['numeric_columns'].append(column_name)
            meta['column_map'] = self._generate_column_mapping(df.columns)

        return True

    def get_composite_indices(self, session_id):
        """Get all composite indices for a session."""
        if not session_id:
            return {}
        user_data = self.get_user_data(session_id)
        return user_data.get('composite_indices', {})

    def delete_composite_index(self, session_id, index_name):
        """Delete a composite index column and its metadata."""
        if not session_id:
            return False
        user_data = self.get_user_data(session_id)
        indices = user_data.get('composite_indices', {})
        if index_name not in indices:
            return False

        idx_info = indices[index_name]
        layer_id = idx_info.get('layer_id')
        if layer_id and layer_id in user_data['layers']:
            df = user_data['layers'][layer_id]
            if index_name in df.columns:
                df.drop(columns=[index_name], inplace=True)
            meta = next((m for m in user_data['metadata'] if m['id'] == layer_id), None)
            if meta:
                if index_name in meta.get('columns', []):
                    meta['columns'].remove(index_name)
                if index_name in meta.get('numeric_columns', []):
                    meta['numeric_columns'].remove(index_name)
                meta['column_map'] = self._generate_column_mapping(df.columns)

        del indices[index_name]
        return True

    def cleanup_session(self, session_id):
        """Remove a session and free its memory."""
        if session_id and session_id in self.sessions:
            del self.sessions[session_id]
            self.session_timestamps.pop(session_id, None)
            return True
        return False

# Initialize Global Manager
dm = DataManager()

# --- Periodic Session Cleanup Thread ---
import threading
import gc
import time as _time

def _session_cleanup_loop(dm_instance, max_age_hours=1, interval=600):
    """Background thread that removes stale sessions every `interval` seconds."""
    while True:
        _time.sleep(interval)
        try:
            now = datetime.now()
            stale = [sid for sid, ts in list(dm_instance.session_timestamps.items())
                     if (now - ts).total_seconds() / 3600 > max_age_hours]
            for sid in stale:
                dm_instance.cleanup_session(sid)
                print(f"[Session Cleanup MOZ] Removed stale session: {sid[:8]}...")
            if stale:
                gc.collect()

                # Force OS memory release (Linux-specific, with robust fallback)
                try:
                    import ctypes
                    import platform
                    if platform.system() == 'Linux':
                        ctypes.CDLL('libc.so.6').malloc_trim(0)
                        print(f"[Session Cleanup MOZ] Forced OS memory release (malloc_trim)")
                except:
                    pass  # Silently fail on Windows/Mac or if library unavailable

                print(f"[Session Cleanup MOZ] Cleaned {len(stale)} sessions.")
        except Exception as e:
            print(f"[Session Cleanup MOZ] Error: {e}")

_cleanup_thread = threading.Thread(target=_session_cleanup_loop, args=(dm,), daemon=True)
_cleanup_thread.start()

def agcap_explorer(settles_gdf, default_column, figure_title, lang='en', url_base_pathname='/'):
    """
    Create AgCAP Explorer app with language support

    Args:
        settles_gdf: GeoDataFrame with settlement data
        default_column: Default column to display
        figure_title: Title for the application
        lang: Language code ('en' or 'pt')
    """
    # --- Guard: Only process data on FIRST call (prevent redundant processing for en/pt) ---
    if dm.default_context is None:
        # --- Initial Data Processing ---
        if settles_gdf.crs is not None and settles_gdf.crs.to_string() != "EPSG:4326":
            try:
                settles_gdf = settles_gdf.to_crs(epsg=4326)
            except Exception as e:
                print(f"Warning: Could not convert CRS. Error: {e}")

        df_main = settles_gdf.copy()
        centroids = df_main.geometry.centroid
        df_main['lat'] = centroids.y
        df_main['lon'] = centroids.x
        df_main = df_main.drop(columns='geometry')

        # Initialize Custom Columns
        if 'Notes' not in df_main.columns:
            df_main['Notes'] = ""
        if 'Score' not in df_main.columns:
            df_main['Score'] = 0

        # Detect Types for Filter Panel
        numeric_columns = []
        categorical_columns = []
        for col in df_main.columns:
            if pd.api.types.is_numeric_dtype(df_main[col]):
                df_main[col] = pd.to_numeric(df_main[col], errors='coerce').astype('float64')
                numeric_columns.append(col)
            else:
                df_main[col] = df_main[col].astype(str)
                categorical_columns.append(col)

        all_filterable_columns = numeric_columns + categorical_columns
        if default_column not in numeric_columns and numeric_columns:
            default_column = numeric_columns[0]

        # CHANGED: Do NOT add the layer to a specific user yet.
        # Instead, save this setup as the "Default Context" in the DataManager.
        # A callback will trigger later to copy this into the user's session.
        dm.set_default_context(df_main, "Settlements", default_column)
        print(f"[AgCAP Explorer MOZ] Initial data processing completed (lang={lang})")

    # Get the already-processed data (runs on both calls - en/pt)
    df_main = dm.default_context['df']
    numeric_columns = [c for c in df_main.columns if pd.api.types.is_numeric_dtype(df_main[c])]
    categorical_columns = [c for c in df_main.columns if c not in numeric_columns]
    all_filterable_columns = numeric_columns + categorical_columns

    # Available Color Scales
    color_scales = [
        'Turbo', 'Jet', 'Electric', 'Rainbow', 'Portland', 'Viridis', 
        'Plasma', 'Hot', 'Blackbody', 'RdBu', 'Earth'
    ]
    
    # Basemap Options
    basemap_options = [
        {'label': 'OpenStreetMap', 'value': 'open-street-map'},
        {'label': 'Carto Positron', 'value': 'carto-positron'},
        {'label': 'Carto Dark Matter', 'value': 'carto-darkmatter'},
        {'label': 'Carto Voyager', 'value': 'carto-voyager'},
        {'label': 'Satellite', 'value': 'satellite'},
        {'label': 'Satellite Streets', 'value': 'satellite-streets'},
        {'label': 'Light', 'value': 'light'},
        {'label': 'Dark', 'value': 'dark'},
    ]
    
    # 1. DEFINE THE CORRECT ASSETS PATH
    # Get the folder where this script lives (/scripts)
    current_script_folder = os.path.dirname(os.path.abspath(__file__))
    # Go up one level to the project root
    project_root = os.path.dirname(current_script_folder)
    # Point to the 'assets' folder in the root
    assets_path = os.path.join(project_root, 'assets')

    app = dash.Dash(__name__, external_stylesheets=external_stylesheets, external_scripts=external_scripts, assets_folder=assets_path, requests_pathname_prefix=url_base_pathname)
    app.title = figure_title

    # Store language in app config for use in callbacks
    app.lang = lang

    # Derive language-switch hrefs from url_base_pathname.
    # For deployment (e.g. '/moz/en/'): _lang_root='/moz' → '/moz/en/', '/moz/pt/'
    # For standalone/notebook ('/'): _lang_root='' → '/', '/'
    _base = url_base_pathname.rstrip('/')
    _lang_root = _base.rsplit('/', 1)[0] if '/' in _base else ''
    _en_href = f'{_lang_root}/en/' if _lang_root else '/'
    _pt_href = f'{_lang_root}/pt/' if _lang_root else '/'

    # --- Flask cleanup endpoint for session management ---
    from flask import request as flask_request

    @app.server.route('/api/cleanup-session', methods=['POST'])
    def cleanup_session_endpoint():
        data = flask_request.get_json(silent=True) or {}
        session_id = data.get('session_id', '')
        if session_id:
            dm.cleanup_session(session_id)
            gc.collect()
        return {'status': 'ok'}, 200

    # ---------------- GLOBAL CSS -------------------
    app.index_string = """
    <!DOCTYPE html>
    <html>
        <head>
            {%metas%}
            <title>{%title%}</title>
            {%favicon%}
            {%css%}
            <style>
                * { box-sizing: border-box; }
                :root {
                    --brand-gold: #FBB800;
                    --brand-dark-blue: #44546A;
                    --brand-grey: #555555;
                    --brand-purple: #583CA5;
                    --brand-light: #ecf0f1;
                    --accent-contrast: #3498db;
                }
                html, body, #_dash-app-content {
                    margin: 0 !important; padding: 0 !important;
                    height: 100% !important; overflow: hidden !important;
                    font-family: 'Mulish', sans-serif;
                }
                h1, h2, h3, h4, h5, h6, .barlow-title {
                    font-family: 'Barlow Condensed', sans-serif !important;
                    text-transform: uppercase;
                }
                .js-plotly-plot, .plot-container, .modebar { margin: 0 !important; padding: 0 !important; }
                
                /* Scrollbars */
                ::-webkit-scrollbar { width: 8px; height: 8px;}
                ::-webkit-scrollbar-track { background: var(--brand-dark-blue); }
                ::-webkit-scrollbar-thumb { background: #8899aa; border-radius: 4px; }
                ::-webkit-scrollbar-thumb:hover { background: var(--brand-gold); }
                
                #map-graph { width: 100%; height: 100vh; }
                
                /* --- COMPACT DROPDOWN STYLING --- */
                .Select-control {
                    background-color: var(--brand-dark-blue) !important;
                    border: 1px solid #777 !important;
                    color: white !important;
                    height: 26px !important; min-height: 26px !important;
                }
                .Select-value { line-height: 26px !important; }
                .Select-input { height: 26px !important; }
                .Select-placeholder { line-height: 26px !important; color: #aaa !important; }
                .Select-value-label { color: white !important; line-height: 26px !important; }
                .Select-menu-outer {
                    background-color: var(--brand-dark-blue) !important;
                    border: 1px solid #777 !important;
                    color: white !important;
                    margin-top: 0px !important;
                }
                .Select-option {
                    background-color: var(--brand-dark-blue) !important;
                    color: white !important;
                    padding: 4px 8px !important;
                }
                .Select-option:hover, .Select-option.is-focused {
                    background-color: var(--brand-purple) !important;
                    color: white !important;
                }
                .Select-arrow-zone { color: white !important; }
                
                input[type="text"], input[type="number"], input[type="password"] {
                    background-color: var(--brand-grey) !important;
                    border: 1px solid #777 !important;
                    color: white !important;
                    height: 26px !important;
                    font-size: 11px !important;
                }
                input[type="color"] {
                    height: 26px !important; padding: 0 !important; border: none !important;
                }

                .rc-slider-track { background-color: var(--brand-purple) !important; }
                .rc-slider-rail { background-color: #333 !important; }
                .rc-slider-handle {
                    border: 2px solid var(--brand-gold) !important;
                    background-color: var(--brand-gold) !important;
                    opacity: 1 !important;
                    width: 10px !important; height: 10px !important;
                    margin-top: -3px !important;
                }
                .rc-slider-handle:hover { border-color: white !important; }
                
                /* Table Styles */
                .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner input:not([type=radio]):not([type=checkbox]) {
                    color: white !important;
                    font-family: 'Mulish', sans-serif !important;
                    background-color: #3a4d63 !important;
                    border: 1px solid #8899aa !important;
                }
                .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner input::placeholder {
                    color: #aabbcc !important;
                }
                .dash-table-container .previous-next-container .page-number,
                .dash-table-container .previous-next-container .last-page,
                .dash-table-container .previous-next-container .first-page,
                .dash-table-container .previous-next-container .previous-page,
                .dash-table-container .previous-next-container .next-page {
                    color: #ffffff !important;
                    font-size: 11px !important;
                    font-weight: 600 !important;
                    background-color: #583CA5 !important;
                    border: 1px solid #7B61C4 !important;
                    border-radius: 4px !important;
                    padding: 4px 10px !important;
                    margin: 0 2px !important;
                    cursor: pointer !important;
                }
                .dash-table-container .previous-next-container .page-number:hover,
                .dash-table-container .previous-next-container .last-page:hover,
                .dash-table-container .previous-next-container .first-page:hover,
                .dash-table-container .previous-next-container .previous-page:hover,
                .dash-table-container .previous-next-container .next-page:hover {
                    background-color: #6B4FBB !important;
                }
                .dash-table-container .previous-next-container .page-number.current-page {
                    color: #ffffff !important;
                    background-color: #FBB800 !important;
                    border-color: #FBB800 !important;
                }
                .dash-table-container .previous-next-container {
                    padding: 6px 8px !important;
                    background-color: #3a4d63 !important;
                    display: flex !important;
                    align-items: center !important;
                    gap: 2px !important;
                }
                
                .sidebar-btn {
                    display: flex; align-items: center; justify-content: center;
                    width: 34px; height: 34px; margin: 15px auto;
                    border: none; border-radius: 50%;
                    background-color: transparent; color: #bdc3c7;
                    cursor: pointer; font-size: 16px; transition: all 0.2s ease;
                }
                .sidebar-btn:hover {
                    background-color: rgba(255, 255, 255, 0.1);
                    color: var(--brand-gold); 
                    box-shadow: 0 0 8px rgba(251, 184, 0, 0.4);
                }
                
                .export {
                    background-color: var(--brand-purple) !important;
                    color: white !important;
                    border: none !important;
                    border-radius: 4px !important;
                    padding: 4px 10px !important;
                    margin-bottom: 5px !important;
                    font-family: 'Barlow Condensed', sans-serif !important;
                    font-weight: bold;
                    letter-spacing: 0.5px;
                    cursor: pointer !important;
                    font-size: 12px;
                }
                .export:hover {
                    background-color: var(--brand-gold) !important;
                    color: var(--brand-dark-blue) !important;
                }

                button { transition: all 0.3s ease; }
                
                #resize-handle {
                    width: 100%; height: 8px;
                    background-color: var(--brand-dark-blue);
                    border-top: 1px solid var(--brand-grey);
                    border-bottom: 1px solid var(--brand-grey);
                    cursor: ns-resize;
                    display: flex; justify-content: center; align-items: center;
                }
                #resize-handle::after {
                    content: "......"; color: #7f8c8d; font-size: 14px;
                    line-height: 0px; letter-spacing: 2px; margin-top: -8px; 
                }
                #resize-handle:hover { background-color: var(--brand-purple); }
                #resize-handle:hover::after { color: white; }

                /* Custom Map Toolbar */
                .toolbar-btn {
                    width: 30px; height: 30px; border: none; border-radius: 4px;
                    background-color: transparent; color: #bdc3c7; cursor: pointer;
                    font-size: 13px; display: flex; align-items: center; justify-content: center;
                    transition: all 0.15s ease; padding: 0;
                }
                .toolbar-btn:hover { background-color: rgba(255, 255, 255, 0.15); color: #FBB800; }
                .toolbar-btn.active-tool { background-color: #FBB800 !important; color: #1e1e1e !important; }
                .lasso-icon {
                    width: 18px; height: 18px;
                    -webkit-mask: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cellipse cx='11' cy='10' rx='8' ry='6.5' stroke='%23000' stroke-width='2.2' fill='none'/%3E%3Cpath d='M7 14.5C5 17 6.5 20 9 18.5C10.5 17.5 9 15.5 7 14.5Z' stroke='%23000' stroke-width='1.8' fill='none'/%3E%3Cpolygon points='16,13 23,23 19,23 16.5,18 14,23 14,13' fill='%23000' stroke='%23000' stroke-width='0.5' stroke-linejoin='round'/%3E%3C/svg%3E") center/contain no-repeat;
                    mask: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cellipse cx='11' cy='10' rx='8' ry='6.5' stroke='%23000' stroke-width='2.2' fill='none'/%3E%3Cpath d='M7 14.5C5 17 6.5 20 9 18.5C10.5 17.5 9 15.5 7 14.5Z' stroke='%23000' stroke-width='1.8' fill='none'/%3E%3Cpolygon points='16,13 23,23 19,23 16.5,18 14,23 14,13' fill='%23000' stroke='%23000' stroke-width='0.5' stroke-linejoin='round'/%3E%3C/svg%3E") center/contain no-repeat;
                    background-color: currentColor;
                }

                /* Custom cursor follower for lasso/box modes */
                #cursor-follower {
                    position: fixed;
                    pointer-events: none;
                    z-index: 9999;
                    display: none;
                    background: rgba(40,30,70,0.82);
                    color: #FBB800;
                    font-size: 11px;
                    font-family: Mulish, sans-serif;
                    font-weight: 700;
                    padding: 2px 7px;
                    border-radius: 4px;
                    white-space: nowrap;
                    letter-spacing: 0.03em;
                }
                #cursor-follower.active { display: block; }

                /* Scale bar */
                #map-scalebar {
                    position: absolute;
                    bottom: 12px;
                    left: 50%;
                    transform: translateX(-50%);
                    z-index: 5;
                    pointer-events: none;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    gap: 2px;
                }
                #map-scalebar-bar {
                    width: 120px;
                    height: 4px;
                    background: #fff;
                    border: 1.5px solid #333;
                    border-top: none;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.4);
                }
                #map-scalebar-label {
                    font-family: Mulish, sans-serif;
                    font-size: 11px;
                    font-weight: 700;
                    color: #fff;
                    text-shadow: 0 0 3px #000, 0 0 3px #000;
                    letter-spacing: 0.04em;
                }

                /* Global loading spinner */
                @keyframes agcap-spin { from { transform: rotate(0deg); } to { transform: rotate(360deg); } }
                #global-loading-spinner {
                    position: fixed;
                    left: calc(59vw - 30px);
                    top: calc(50vh - 30px);
                    z-index: 9998;
                    width: 60px; height: 60px;
                    border: 6px solid rgba(251, 184, 0, 0.2);
                    border-top-color: #FBB800;
                    border-radius: 50%;
                    animation: agcap-spin 0.8s linear infinite;
                    pointer-events: none;
                    display: none;
                }
                #global-loading-spinner.active { display: block; }

                #right-pane-resize-left:hover { background-color: rgba(251, 184, 0, 0.3); }
                #right-pane-resize-bottom:hover { background-color: rgba(251, 184, 0, 0.3); }
                #right-pane-resize-corner:hover { background-color: rgba(251, 184, 0, 0.5); }

                #legend-container {
                    background: rgba(44, 62, 80, 0.85);
                    border-radius: 5px; padding: 10px;
                    max-height: 300px; overflow-y: auto;
                    color: white; font-family: 'Mulish', sans-serif;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.3);
                    backdrop-filter: blur(4px); border: 1px solid #555;
                }
                .legend-item { display: flex; align-items: center; margin-bottom: 8px; font-size: 11px; }
                .legend-marker { border-radius: 50%; margin-right: 8px; border: 1px solid white; flex-shrink: 0; }
                .legend-gradient-bar { height: 8px; width: 60px; border-radius: 2px; margin: 0 5px; border: 1px solid #ccc; flex-shrink: 0; }
                .legend-label { color: #ecf0f1; font-weight: 300; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 120px; }

                /* Details/Summary Styling for Analysis */
                details { margin-bottom: 10px; border: 1px solid #555; border-radius: 4px; overflow: hidden; }
                summary { padding: 8px 10px; background-color: #34495e; cursor: pointer; color: #ecf0f1; font-size: 12px; font-weight: bold; font-family: 'Barlow Condensed'; letter-spacing: 0.5px; list-style: none; display: flex; align-items: center; justify-content: space-between; }
                summary:hover { background-color: #44546A; color: #FBB800; }
                summary::-webkit-details-marker { display: none; }
                summary::after { content: '+'; font-size: 14px; font-weight: bold; }
                details[open] summary::after { content: '-'; }
                details[open] summary { border-bottom: 1px solid #555; }
                .analysis-content-inner { padding: 10px; background-color: rgba(0,0,0,0.2); }

                /* Radio Items Styling */
                .radio-group label { margin-right: 15px; color: #bdc3c7; font-size: 11px; cursor: pointer; }
                .radio-group input { margin-right: 5px; }
                
                /* Loading Screen Fade Out Animation */
                .fade-out {
                    opacity: 0;
                    visibility: hidden;
                    transition: opacity 1s ease-out, visibility 1s ease-out;
                }
            </style>
        </head>
        <body>
            <div id="global-loading-spinner"></div>
            <div id="cursor-follower"></div>
            {%app_entry%}
            <footer>
                {%config%}
                {%scripts%}
                {%renderer%}
                <script>
                    document.addEventListener('DOMContentLoaded', function() {
                        // Global state for pane/table sizing
                        if (!window.__agcap_state) {
                            window.__agcap_state = {
                                tableHeight: '25vh',
                                rightPaneWidth: '22vw',
                                tableIsOpen: false
                            };
                        }

                        // Helper: update left/right pane bottoms to sit above the table
                        function updatePaneBottoms() {
                            var tablePanel = document.getElementById('table-panel');
                            var leftPane = document.getElementById('left-pane-container');
                            var rightPane = document.getElementById('right-pane-container');
                            var tableH = (window.__agcap_state.tableIsOpen && tablePanel) ? tablePanel.offsetHeight : 0;
                            if (leftPane && leftPane.style.display !== 'none') {
                                leftPane.style.bottom = tableH + 'px';
                            }
                            if (rightPane && rightPane.style.display !== 'none') {
                                rightPane.style.bottom = tableH + 'px';
                            }
                        }
                        window.__agcap_updatePaneBottoms = updatePaneBottoms;

                        setTimeout(function() {
                            // Table resize handle
                            const handle = document.getElementById('resize-handle');
                            const panel = document.getElementById('table-panel');
                            if (handle && panel) {
                                let isDragging = false;
                                let startY, startHeight;

                                handle.addEventListener('mousedown', function(e) {
                                    isDragging = true;
                                    startY = e.clientY;
                                    startHeight = parseInt(window.getComputedStyle(panel).height, 10);
                                    document.body.style.cursor = 'ns-resize';
                                    e.preventDefault();
                                });

                                document.addEventListener('mousemove', function(e) {
                                    if (!isDragging) return;
                                    const dy = startY - e.clientY;
                                    let newHeight = startHeight + dy;
                                    if (newHeight < 50) newHeight = 50;
                                    if (newHeight > window.innerHeight - 100) newHeight = window.innerHeight - 100;
                                    panel.style.height = newHeight + 'px';
                                    updatePaneBottoms();
                                });

                                document.addEventListener('mouseup', function() {
                                    if (isDragging) {
                                        isDragging = false;
                                        document.body.style.cursor = 'default';
                                        window.__agcap_state.tableHeight = panel.offsetHeight + 'px';
                                        updatePaneBottoms();
                                    }
                                });
                            }

                            // Right pane resize functionality
                            const rPane = document.getElementById('right-pane-container');
                            const rLeft = document.getElementById('right-pane-resize-left');
                            const rBottom = document.getElementById('right-pane-resize-bottom');
                            const rCorner = document.getElementById('right-pane-resize-corner');

                            function initRightResize(handle, mode) {
                                if (!handle || !rPane) return;
                                handle.addEventListener('mousedown', function(e) {
                                    e.preventDefault();
                                    const startX = e.clientX, startY = e.clientY;
                                    const startW = rPane.offsetWidth;

                                    function onMove(ev) {
                                        if (mode === 'ew' || mode === 'corner') {
                                            const newW = Math.max(200, startW - (ev.clientX - startX));
                                            rPane.style.width = newW + 'px';
                                        }
                                    }
                                    function onUp() {
                                        document.removeEventListener('mousemove', onMove);
                                        document.removeEventListener('mouseup', onUp);
                                        document.body.style.cursor = 'default';
                                        window.__agcap_state.rightPaneWidth = rPane.offsetWidth + 'px';
                                    }
                                    document.body.style.cursor = (mode === 'ew') ? 'ew-resize' : 'nesw-resize';
                                    document.addEventListener('mousemove', onMove);
                                    document.addEventListener('mouseup', onUp);
                                });
                            }

                            initRightResize(rLeft, 'ew');
                            initRightResize(rCorner, 'corner');
                        }, 2000);
                    });

                    // Session cleanup handled by periodic server thread (4h) and logout route

                    // --- Middle-click pan + S-key lasso + custom cursor follower ---
                    (function() {
                        var follower = document.getElementById('cursor-follower');

                        function getMapEl() {
                            var el = document.getElementById('map');
                            if (!el) return null;
                            return el._fullLayout ? el : (el.querySelector('.js-plotly-plot') || el);
                        }
                        function setDragMode(mode) {
                            var gd = getMapEl();
                            if (gd) Plotly.relayout(gd, {'dragmode': mode});
                        }
                        function getCurrentMode() {
                            var gd = getMapEl();
                            if (!gd || !gd._fullLayout) return 'pan';
                            return gd._fullLayout.dragmode || 'pan';
                        }

                        // Middle-click pan: while middle button is held, force pan mode
                        document.addEventListener('mousedown', function(e) {
                            if (e.button !== 1) return;
                            var mapEl = document.getElementById('map');
                            if (!mapEl || !mapEl.contains(e.target)) return;
                            e.preventDefault();
                            var prev = getCurrentMode();
                            setDragMode('pan');
                            updateCursor('pan');
                            function onUp(ev) {
                                if (ev.button !== 1) return;
                                setDragMode(prev);
                                updateCursor(prev);
                                document.removeEventListener('mouseup', onUp);
                            }
                            document.addEventListener('mouseup', onUp);
                        });

                        // Update cursor follower text based on current drag mode
                        function updateCursor(mode) {
                            if (!follower) return;
                            if (mode === 'lasso') {
                                var btn = document.getElementById('tool-lasso');
                                follower.textContent = btn ? btn.title : 'Lasso';
                                follower.classList.add('active');
                            } else if (mode === 'select') {
                                var btn = document.getElementById('tool-box');
                                follower.textContent = btn ? btn.title : 'Box';
                                follower.classList.add('active');
                            } else {
                                follower.classList.remove('active');
                            }
                        }

                        // Temporarily highlight a toolbar button (key feedback)
                        function setToolHighlight(toolId, on) {
                            var btn = document.getElementById(toolId);
                            if (!btn) return;
                            var modeMap = {'tool-lasso':'lasso','tool-box':'select','tool-pan':'pan'};
                            if (on) {
                                btn.classList.add('active-tool');
                            } else {
                                // Only remove if not already the real active tool
                                if (window.__agcap_active_tool !== modeMap[toolId]) btn.classList.remove('active-tool');
                            }
                        }

                        // Track toolbar button clicks to know the "real" active tool
                        window.__agcap_active_tool = 'pan';
                        ['tool-lasso','tool-box','tool-pan'].forEach(function(id) {
                            document.addEventListener('click', function(e) {
                                if (e.target && (e.target.id === id || e.target.closest('#' + id))) {
                                    var modeMap = {'tool-lasso':'lasso','tool-box':'select','tool-pan':'pan'};
                                    window.__agcap_active_tool = modeMap[id];
                                    updateCursor(window.__agcap_active_tool);
                                }
                            });
                        });

                        // Mouse move: keep follower 14px to bottom-right of pointer
                        document.addEventListener('mousemove', function(e) {
                            if (follower && follower.classList.contains('active')) {
                                follower.style.left = (e.clientX + 14) + 'px';
                                follower.style.top  = (e.clientY + 14) + 'px';
                            }
                        });

                        // Temporary tool keys: S=lasso, R=box select, P=pan
                        var tempKeyPrev = null; // {mode, toolId, key}
                        document.addEventListener('keydown', function(e) {
                            if (e.repeat || tempKeyPrev !== null) return;
                            var tag = document.activeElement && document.activeElement.tagName;
                            if (tag === 'INPUT' || tag === 'TEXTAREA') return;
                            var k = e.key;
                            var target = null;
                            if (k === 's' || k === 'S') target = {mode:'lasso', toolId:'tool-lasso'};
                            else if (k === 'r' || k === 'R') target = {mode:'select', toolId:'tool-box'};
                            else if (k === 'p' || k === 'P') {
                                var cur = getCurrentMode();
                                if (cur === 'lasso' || cur === 'select') target = {mode:'pan', toolId:'tool-pan'};
                            }
                            if (!target) return;
                            tempKeyPrev = {mode: getCurrentMode(), key: k, toolId: target.toolId};
                            setDragMode(target.mode);
                            setToolHighlight(target.toolId, true);
                            updateCursor(target.mode);
                        });
                        document.addEventListener('keyup', function(e) {
                            if (!tempKeyPrev) return;
                            var k = e.key;
                            if (k.toLowerCase() !== tempKeyPrev.key.toLowerCase()) return;
                            setDragMode(tempKeyPrev.mode);
                            setToolHighlight(tempKeyPrev.toolId, false);
                            updateCursor(tempKeyPrev.mode);
                            tempKeyPrev = null;
                        });
                    })();

                    // --- Global loading spinner: intercept fetch to detect server callbacks ---
                    (function() {
                        var pending = 0;
                        var spinner = document.getElementById('global-loading-spinner');
                        var origFetch = window.fetch;
                        window.fetch = function() {
                            var url = (arguments[0] && arguments[0].url) || arguments[0] || '';
                            var isCallback = (typeof url === 'string' && url.indexOf('_dash-update-component') !== -1);
                            if (isCallback) {
                                pending++;
                                if (spinner) spinner.classList.add('active');
                            }
                            return origFetch.apply(this, arguments).finally(function() {
                                if (isCallback) {
                                    pending--;
                                    if (pending <= 0) {
                                        pending = 0;
                                        if (spinner) spinner.classList.remove('active');
                                    }
                                }
                            });
                        };
                    })();
                </script>
            </footer>
        </body>
    </html>
    """

    # --- COMPONENT SETUP ---
    sidebar_items = []
    
 # --- MODIFIED: LOGO WITH LINK ---
    sidebar_items.append(
        html.A(
            html.Img(
                src=app.get_asset_url('images/logo.png'),  # Use absolute path from web root
                style={'width': '35px', 'display': 'block', 'margin': '0 auto'}
            ),
            href="https://www.seforall.org/",
            target="_blank",  # Opens in new tab
            style={'display': 'block', 'margin': '15px auto 5px auto', 'textDecoration': 'none'}
        )
    )
    # --------------------------------
    
    sidebar_items.extend([
        html.A(html.I(className="fas fa-arrow-left"), href="/", title=t('sidebar.back', lang), className="sidebar-btn", style={'textDecoration': 'none', 'display': 'flex'}),
        html.Button(html.I(className="fas fa-layer-group"), id="layers-btn", n_clicks=0, title=t('sidebar.layers', lang), className="sidebar-btn"),
        html.Button(html.I(className="fas fa-sliders-h"), id="filters-btn", n_clicks=0, title=t('sidebar.filters', lang), className="sidebar-btn"),
        html.Button(html.I(className="fas fa-calculator"), id="composite-btn", n_clicks=0, title=t('composite.title', lang), className="sidebar-btn"),
        html.Button(html.I(className="fas fa-info-circle"), id="info-btn", n_clicks=0, title=t('sidebar.documentation', lang), className="sidebar-btn"),
        html.Button(html.I(className="fas fa-globe"), id="language-btn", n_clicks=0, title=t('sidebar.language', lang), className="sidebar-btn"),
    ])

    sidebar_items.extend([
        html.Div(style={'flexGrow': 1}),
        html.Button(html.I(className="fas fa-save"), id="save-session-btn", n_clicks=0, title=t('sidebar.save_session', lang), className="sidebar-btn", style={'color': '#FBB800'}),
        dcc.Upload(
            id='load-session-upload',
            children=html.Div(html.I(className="fas fa-folder-open"), className="sidebar-btn", title=t('sidebar.load_session', lang), style={'color': '#FBB800'}),
            accept='.agcap,.json',
            multiple=False,
            style={'cursor': 'pointer'}
        )
    ])

    map_overlays = [
        dcc.Graph(
                id='map', 
                # Initialize with hidden axes so it looks clean even if the loader fades fast
                figure={'layout': {'xaxis': {'visible': False}, 'yaxis': {'visible': False}, 'paper_bgcolor': '#44546A', 'plot_bgcolor': '#44546A'}},
                style={'width': '100%', 'height': '100vh'}, 
                config={'scrollZoom': True, 'displayModeBar': False, 'responsive': True}
            ),

        # Custom Map Toolbar (Left side of map)
        html.Div(id='map-toolbar', children=[
            html.Button(html.Div(className='lasso-icon'), id='tool-lasso', n_clicks=0,
                        title=t('toolbar.lasso_select', lang), className='toolbar-btn'),
            html.Button(html.I(className="fas fa-vector-square"), id='tool-box', n_clicks=0,
                        title=t('toolbar.box_select', lang), className='toolbar-btn'),
            html.Button(html.I(className="fas fa-hand-paper"), id='tool-pan', n_clicks=0,
                        title=t('toolbar.pan', lang), className='toolbar-btn active-tool'),
            html.Div(style={'width': '1px', 'height': '20px', 'backgroundColor': '#666', 'margin': '0 3px'}),
            html.Button(html.I(className="fas fa-search-plus"), id='tool-zoom-in', n_clicks=0,
                        title=t('toolbar.zoom_in', lang), className='toolbar-btn'),
            html.Button(html.I(className="fas fa-search-minus"), id='tool-zoom-out', n_clicks=0,
                        title=t('toolbar.zoom_out', lang), className='toolbar-btn'),
            html.Div(style={'width': '1px', 'height': '20px', 'backgroundColor': '#666', 'margin': '0 3px'}),
            html.Button(html.I(className="fas fa-camera"), id='tool-screenshot', n_clicks=0,
                        title=t('toolbar.screenshot', lang), className='toolbar-btn'),
            html.Button(html.I(className="fas fa-crosshairs"), id='tool-reset-view', n_clicks=0,
                        title=t('toolbar.reset_view', lang), className='toolbar-btn'),
            html.Button(html.I(className="fas fa-redo-alt"), id='tool-reset-platform', n_clicks=0,
                        title=t('toolbar.reset_platform', lang), className='toolbar-btn'),
        ], style={
            'position': 'absolute', 'top': '15px', 'left': '35px', 'zIndex': 10,
            'display': 'flex', 'alignItems': 'center', 'gap': '2px',
            'backgroundColor': 'rgba(30, 30, 30, 0.85)', 'borderRadius': '6px',
            'padding': '4px 6px', 'boxShadow': '0 2px 8px rgba(0,0,0,0.4)',
            'backdropFilter': 'blur(4px)',
        }),

        # Hidden dummy output for side-effect-only callbacks
        html.Div(id='toolbar-dummy-output', style={'display': 'none'}),

        # Geocoder Bar (Final Integrated Search Lens - Small)
        html.Div([
            dcc.Input(
                id='geocoder-input',
                type='text',
                placeholder=t('geocoder.placeholder', lang), 
                style={
                    'width': '100%', 
                    'height': '26px',         # Smaller Height
                    'padding': '0 30px 0 10px', # Reduced Padding
                    'borderRadius': '13px',   # Adjusted Radius
                    'border': 'none', 
                    'outline': 'none', 
                    'backgroundColor': '#555555',
                    'color': 'white',
                    'fontFamily': 'Mulish',
                    'fontSize': '11px',       # Smaller Font
                    'lineHeight': '26px',
                    'boxShadow': '0 2px 5px rgba(0,0,0,0.3)'
                }
            ),
            html.Button(
                html.I(className="fas fa-search", style={'fontSize': '11px'}), # Smaller Icon
                id='geocoder-btn', 
                n_clicks=0, 
                style={
                    'position': 'absolute',
                    'right': '0',
                    'top': '0',
                    'height': '26px',         # Matches Input Height
                    'width': '30px',          # Smaller Width
                    'border': 'none', 
                    'backgroundColor': 'transparent',
                    'color': '#ecf0f1', 
                    'cursor': 'pointer',
                    'display': 'flex',
                    'alignItems': 'center',
                    'justifyContent': 'center',
                    'borderRadius': '0 13px 13px 0', # Adjusted Radius
                    'zIndex': 1
                }
            )
        ], style={
            'position': 'absolute', 
            'top': '15px', 
            'left': '50%', 
            'transform': 'translateX(-50%)', 
            'zIndex': 10, 
            'width': '180px', # Narrower Container
            'display': 'flex',
            'alignItems': 'center',
            'height': '26px',
        }),
        
        # Legend Container (Left side of map)
        html.Div(id='legend-container', style={'position': 'absolute', 'top': '60px', 'left': '15px', 'zIndex': 5, 'width': '220px', 'display': 'none'}),

        # Scale bar (bottom centre of map)
        html.Div(id='map-scalebar', children=[
            html.Div(id='map-scalebar-label', children=''),
            html.Div(id='map-scalebar-bar'),
        ]),

        # Bottom right logos (all files in assets/logos/MOZ/, sorted alphabetically)
        html.Div(
            id='map-logos',
            children=[
                html.Img(
                    src=app.get_asset_url(f'logos/MOZ/{fname}'),
                    style={'height': '8vh', 'width': 'auto', 'display': 'block'}
                )
                for fname in sorted([
                    f for f in (
                        os.listdir(scripts_dir.parent / 'assets' / 'logos' / 'MOZ')
                        if (scripts_dir.parent / 'assets' / 'logos' / 'MOZ').exists() else []
                    )
                    if f.lower().endswith(('.png', '.svg', '.jpg', '.jpeg', '.webp'))
                ])
            ],
            style={
                'position': 'absolute',
                'bottom': '35px',
                'right': '20px',
                'display': 'flex',
                'flexDirection': 'row',
                'alignItems': 'center',
                'gap': '10px',
                'zIndex': 5,
                'opacity': 0.9,
                'pointerEvents': 'none',
            }
        ),

    ]

    # --- SPLASH SCREEN COMPONENT ---
    loading_screen = html.Div(
        id='initial-loader',
        children=[
            html.Img(
                src=app.get_asset_url('images/logo.png'), 
                style={'width': '150px', 'marginBottom': '20px'}
            ),
            html.Div(
                dbc.Spinner(color="white", type="grow"), 
                style={'display': 'flex', 'justifyContent': 'center'}
            )
        ],
        style={
            'position': 'fixed',
            'top': 0,
            'left': 0,
            'width': '100%',
            'height': '100%',
            'backgroundColor': '#583CA5', # Your requested color
            'zIndex': 9999,
            'display': 'flex',
            'flexDirection': 'column',
            'alignItems': 'center',
            'justifyContent': 'center',
        }
    )

    # ---------------------- LAYOUT ----------------------

    # ---------------------- LAYOUT ----------------------
    # Define the main layout
    # Create column translation dictionary for clientside callbacks
    column_translations = {}
    for col in df_main.columns:
        column_translations[col] = get_column_translation(col, lang, COUNTRY_CODE)

    # Look up column tooltips from module-level cache (parsed once at import time)
    column_tooltips = _COLUMN_TOOLTIPS.get(lang, _COLUMN_TOOLTIPS.get('en', {}))

    app.layout = html.Div([
        loading_screen,
        dcc.Store(id='client-layer-data'),
        dcc.Store(id='column-translations-store', data=column_translations),  # Column translations for clientside callbacks
        dcc.Store(id='column-tooltips-store', data=column_tooltips),  # Column tooltips from data dictionary
        dcc.Store(id='current-selection-store', data=[]), # <--- NEW: Persist selections here
        dcc.Store(id='table-state-store', data=False),
        dcc.Store(id='layer-manager-trigger', data=0),
        dcc.Store(id='layer-data-content-trigger', data=0), # Only fires on Add/Delete
        dcc.Store(id='visible-layers-list', data=[]),       # Keeps track of what is visible
        dcc.Store(id='map-view-store', data={'center': None, 'zoom': None}), 
        dcc.Store(id='save-status-store'),
        dcc.Store(id='user-session-id', storage_type='local'),
        dcc.Store(id='active-tool-store', data='pan'),
        dcc.Store(id='bubble-clicked-point'),
        dcc.Store(id='bubble-clear-trigger', data=0),
        dcc.Store(id='econ-clicked-point'),
        dcc.Store(id='econ-clear-trigger', data=0),
        dcc.Store(id='geo-clicked-point'),
        dcc.Store(id='geo-clear-trigger', data=0),
        dcc.Store(id='default-map-view', data={'center': DEFAULT_MAP_CENTER, 'zoom': DEFAULT_MAP_ZOOM}),
        dcc.Store(id='pending-filter-values'),
        dcc.Store(id='filter-log-scale-store', data={}),
        dcc.Store(id='filter-histogram-data', data={}),
        dcc.Download(id="download-session"),
        dcc.Download(id="download-table-csv"),

        # Language Switcher Modal
        html.Div(id='language-modal', style={'display': 'none'}, children=[
            html.Div(style={
                'position': 'fixed',
                'top': 0,
                'left': 0,
                'width': '100%',
                'height': '100%',
                'backgroundColor': 'rgba(0, 0, 0, 0.7)',
                'zIndex': 9998,
            }),
            html.Div([
                html.H3('Select Language / Escolha o idioma', style={'color': '#FBB800', 'marginBottom': '20px', 'fontSize': '18px'}),
                html.Div([
                    html.A(
                        html.Div([
                            html.I(className="fas fa-globe", style={'marginRight': '10px', 'fontSize': '18px'}),
                            html.Span('English', style={'fontSize': '16px'})
                        ], style={
                            'padding': '15px 30px',
                            'backgroundColor': '#583CA5' if lang == 'en' else '#333',
                            'borderRadius': '5px',
                            'cursor': 'pointer',
                            'marginBottom': '10px',
                            'border': '2px solid #FBB800' if lang == 'en' else '2px solid #555',
                            'display': 'flex',
                            'alignItems': 'center',
                            'transition': 'all 0.3s'
                        }),
                        href=_en_href,
                        style={'textDecoration': 'none', 'color': '#ecf0f1', 'display': 'block'}
                    ),
                    html.A(
                        html.Div([
                            html.I(className="fas fa-globe", style={'marginRight': '10px', 'fontSize': '18px'}),
                            html.Span('Português', style={'fontSize': '16px'})
                        ], style={
                            'padding': '15px 30px',
                            'backgroundColor': '#583CA5' if lang == 'pt' else '#333',
                            'borderRadius': '5px',
                            'cursor': 'pointer',
                            'marginBottom': '10px',
                            'border': '2px solid #FBB800' if lang == 'pt' else '2px solid #555',
                            'display': 'flex',
                            'alignItems': 'center',
                            'transition': 'all 0.3s'
                        }),
                        href=_pt_href,
                        style={'textDecoration': 'none', 'color': '#ecf0f1', 'display': 'block'}
                    ),
                ]),
                html.Button('Close / Fechar', id='close-language-modal', n_clicks=0, style={
                    'marginTop': '20px',
                    'padding': '10px 20px',
                    'backgroundColor': '#555',
                    'color': '#ecf0f1',
                    'border': 'none',
                    'borderRadius': '5px',
                    'cursor': 'pointer',
                    'fontSize': '14px'
                })
            ], style={
                'position': 'fixed',
                'top': '50%',
                'left': '50%',
                'transform': 'translate(-50%, -50%)',
                'backgroundColor': '#1e1e1e',
                'padding': '30px',
                'borderRadius': '10px',
                'zIndex': 9999,
                'minWidth': '300px',
                'border': '2px solid #FBB800'
            })
        ]),

        # Info Modal
        html.Div(id='info-modal', style={'display': 'none'}, children=[
            html.Div(style={
                'position': 'fixed', 'top': 0, 'left': 0, 'width': '100%', 'height': '100%',
                'backgroundColor': 'rgba(0, 0, 0, 0.7)', 'zIndex': 9998,
            }),
            html.Div([
                html.H3(t('info.title', lang), style={'color': '#FBB800', 'marginBottom': '15px', 'fontSize': '18px'}),
                html.P(t('info.description', lang), style={'color': '#ecf0f1', 'fontSize': '13px', 'lineHeight': '1.6', 'marginBottom': '20px'}),
                html.A(
                    html.Div([
                        html.I(className="fas fa-code-branch", style={'marginRight': '10px'}),
                        html.Span(t('info.model_source', lang))
                    ], style={'padding': '10px 20px', 'backgroundColor': '#583CA5', 'borderRadius': '5px', 'display': 'inline-flex', 'alignItems': 'center'}),
                    href='https://github.com/SEforALL-IEAP/AgCAP', target='_blank',
                    style={'textDecoration': 'none', 'color': '#ecf0f1', 'display': 'block', 'marginBottom': '20px'}
                ),
                html.P([html.Span('Sustainable Energy for All', style={'fontWeight': 'bold'}), ' — ', html.A('seforall.org', href='https://www.seforall.org/', target='_blank', style={'color': '#FBB800'})], style={'color': '#bbb', 'fontSize': '11px', 'marginBottom': '15px'}),
                html.Button(t('info.close', lang), id='close-info-modal', n_clicks=0, style={
                    'padding': '8px 20px', 'backgroundColor': '#555', 'color': '#ecf0f1',
                    'border': 'none', 'borderRadius': '5px', 'cursor': 'pointer', 'fontSize': '13px'
                })
            ], style={
                'position': 'fixed', 'top': '50%', 'left': '50%', 'transform': 'translate(-50%, -50%)',
                'backgroundColor': '#1e1e1e', 'padding': '30px', 'borderRadius': '10px',
                'zIndex': 9999, 'minWidth': '350px', 'maxWidth': '500px', 'border': '2px solid #FBB800'
            })
        ]),

        # 1. Left Menu
        html.Div(sidebar_items, style={'position':'fixed','top':0,'left':0,'width':'45px','height':'100vh','backgroundColor':'#1e1e1e','zIndex':30, 'borderRight': '1px solid #333', 'display': 'flex', 'flexDirection': 'column'}),

        # 2. Left Pane
        html.Div(id="left-pane-container", children=[
            html.Div(id='left-pane', style={'width':'18vw','height':'100%','backgroundColor':'#44546A','color':'#ecf0f1','padding':'10px','boxSizing':'border-box','boxShadow':'inset -5px 0 10px -5px rgba(0,0,0,0.5)','overflowY':'auto'}, children=[

                html.Div([
                    html.H2(figure_title, style={'color': '#FBB800', 'marginBottom': '0px', 'fontWeight':'700', 'letterSpacing':'1px', 'fontSize': '22px'}),
                    html.H5(t('branding.subtitle', lang), style={'color': '#bdc3c7', 'marginTop': '5px', 'marginBottom': '0px', 'fontWeight': '300', 'fontSize':'12px', 'fontStyle':'italic', 'fontFamily': 'Mulish'})
                ], style={'borderBottom': '1px solid #555555', 'paddingBottom': '15px', 'marginBottom': '15px', 'textAlign':'left'}),

                # --- LAYERS TAB ---
                html.Div(id='layers-content', style={'display': 'none'}, children=[
                    html.H3(t('tabs.layers', lang), style={'borderBottom':'1px solid #555555', 'paddingBottom':'5px', 'marginBottom': '10px', 'fontSize': '16px'}),

                    # Basemap Settings
                    html.Div([
                        html.H4(t('basemap.title', lang), style={'color': '#ecf0f1', 'fontSize': '12px', 'marginBottom': '5px'}),
                        dcc.Dropdown(
                            id='basemap-dropdown',
                            options=basemap_options,
                            value='open-street-map',
                            clearable=False,
                            style={'width': '100%', 'marginBottom': '10px', 'color': '#333', 'fontSize': '11px'}
                        ),
                        dcc.Input(
                            id='mapbox-token-input',
                            type='password',
                            value=MAPBOX_TOKEN,
                            placeholder=t('basemap.token_placeholder', lang),
                            style={'width': '100%', 'padding': '5px', 'borderRadius': '4px', 'border': '1px solid #555555', 'backgroundColor': '#555555', 'color': 'white', 'marginBottom': '10px', 'fontSize': '11px'}
                        )
                    ], style={'marginBottom': '10px', 'borderBottom': '1px dashed #555555', 'paddingBottom': '10px'}),

                    dcc.Upload(
                        id='upload-data',
                        children=html.Button(t('layers.add_layer', lang), style={'width':'100%','padding':'8px','border':'2px dashed #7f8c8d','background':'transparent','color':'#FBB800','borderRadius':'4px','cursor':'pointer', 'fontWeight':'bold', 'marginBottom':'10px', 'fontSize': '12px'}),
                        multiple=False
                    ),
                    html.Div(id='upload-status', style={'marginBottom':'10px', 'color':'#e74c3c', 'fontSize':'11px'}),

                    html.Div(id='layers-list-container')
                ]),

                # --- FILTERS TAB ---
                html.Div(id='filters-content', style={'display': 'block'}, children=[
                    html.H3(t('tabs.filters_primary', lang), style={'borderBottom':'1px solid #555555', 'paddingBottom':'5px', 'marginBottom':'15px', 'fontSize': '16px'}),
                    html.Label(t('filters.filter_by', lang), style={'fontWeight':'bold', 'color':'#FBB800', 'fontSize': '12px'}),
                    dcc.Dropdown(id='column-dropdown', options=[{'label': get_column_translation(col, lang, COUNTRY_CODE), 'value': col} for col in all_filterable_columns], value=[], multi=True, style={'width': '100%', 'marginBottom': '15px', 'borderRadius': '4px', 'color': '#333', 'fontSize': '11px'}),
                    html.Div(id='histograms-container')
                ]),

                # --- COMPOSITE INDEX TAB ---
                html.Div(id='composite-content', style={'display': 'none'}, children=[
                    html.H3([html.I(className="fas fa-calculator", style={'marginRight':'8px'}), t('composite.title', lang)],
                        style={'borderBottom':'1px solid #555555', 'paddingBottom':'5px', 'marginBottom':'15px', 'fontSize':'16px'}),
                    html.Div([
                        html.H4([t('composite.saved_indices', lang), " ", html.I(className="fas fa-info-circle", id="manager-tooltip", style={'fontSize':'12px', 'color':'#bbb', 'cursor':'help'})],
                            style={'fontSize':'12px', 'color':'#FBB800', 'marginBottom':'8px'}),
                        dbc.Tooltip(t('composite.help_manager', lang), target="manager-tooltip", placement="right"),
                        html.Div(id='composite-manager-container', style={'marginBottom':'20px'})
                    ]),
                    html.Hr(style={'borderColor':'#555', 'margin':'10px 0'}),
                    html.H4([t('composite.create_new', lang), " ", html.I(className="fas fa-info-circle", id="create-tooltip", style={'fontSize':'12px', 'color':'#bbb', 'cursor':'help'})],
                        style={'fontSize':'12px', 'color':'#FBB800', 'marginBottom':'8px'}),
                    dbc.Tooltip(t('composite.help_create', lang), target="create-tooltip", placement="right"),
                    html.Div([
                        html.Label([t('composite.column_name', lang), " ", html.I(className="fas fa-question-circle", id="name-tooltip", style={'fontSize':'10px', 'color':'#bbb', 'cursor':'help'})],
                            style={'fontSize':'10px', 'color':'#ecf0f1', 'marginBottom':'3px', 'fontWeight':'bold'}),
                        dbc.Tooltip(t('composite.help_name', lang), target="name-tooltip", placement="right"),
                        dcc.Input(id='composite-name-input', type='text', placeholder=t('composite.placeholder_name', lang), value='',
                                  style={'width':'100%', 'fontSize':'11px', 'padding':'4px', 'borderRadius':'4px', 'border':'1px solid #ccc'})
                    ], style={'marginBottom':'12px'}),
                    html.Div([
                        html.Label([t('composite.target_layer', lang), " ", html.I(className="fas fa-question-circle", id="layer-tooltip", style={'fontSize':'10px', 'color':'#bbb', 'cursor':'help'})],
                            style={'fontSize':'10px', 'color':'#ecf0f1', 'marginBottom':'3px', 'fontWeight':'bold'}),
                        dbc.Tooltip(t('composite.help_layer', lang), target="layer-tooltip", placement="right"),
                        dcc.Dropdown(id='composite-layer-dropdown', clearable=False, style={'fontSize':'11px', 'color':'#333'})
                    ], style={'marginBottom':'12px'}),
                    html.Div([
                        html.Label([t('composite.attributes', lang), " ", html.I(className="fas fa-question-circle", id="columns-tooltip", style={'fontSize':'10px', 'color':'#bbb', 'cursor':'help'})],
                            style={'fontSize':'10px', 'color':'#ecf0f1', 'marginBottom':'3px', 'fontWeight':'bold'}),
                        dbc.Tooltip(t('composite.help_columns', lang), target="columns-tooltip", placement="right"),
                        dcc.Dropdown(id='composite-columns-dropdown', multi=True, placeholder=t('composite.placeholder_attrs', lang), style={'fontSize':'11px', 'color':'#333'})
                    ], style={'marginBottom':'12px'}),
                    html.Div([
                        html.Label([t('composite.method', lang), " ", html.I(className="fas fa-question-circle", id="method-tooltip", style={'fontSize':'10px', 'color':'#bbb', 'cursor':'help'})],
                            style={'fontSize':'10px', 'color':'#ecf0f1', 'marginBottom':'5px', 'fontWeight':'bold'}),
                        dbc.Tooltip(t('composite.help_method', lang), target="method-tooltip", placement="right"),
                        dcc.RadioItems(id='composite-normalization-method',
                            options=[{'label': t('composite.minmax', lang), 'value': 'minmax'}, {'label': t('composite.zscore', lang), 'value': 'zscore'}],
                            value='minmax', inline=False, style={'fontSize':'10px', 'color':'#ecf0f1'})
                    ], style={'marginBottom':'12px'}),
                    html.Div(id='composite-weight-controls-container', style={'marginBottom':'8px'}),
                    html.Div([
                        html.Span(t('composite.sum_label', lang), style={'fontSize':'10px', 'color':'#bbb', 'marginRight':'5px'}),
                        html.Span(id='composite-sum-display', children="0%", style={'fontSize':'11px', 'fontWeight':'bold', 'color':'#e74c3c'})
                    ], style={'marginBottom':'8px', 'textAlign':'center'}),
                    html.Div([
                        html.Button([html.I(className="fas fa-undo", style={'marginRight':'5px'}), t('composite.reset', lang)],
                            id='composite-reset-weights-btn', n_clicks=0,
                            style={'flex':'1', 'padding':'6px', 'backgroundColor':'#555', 'color':'#ecf0f1', 'border':'none', 'borderRadius':'4px', 'cursor':'pointer', 'fontSize':'10px', 'marginRight':'5px', 'fontFamily':'Barlow Condensed'}),
                        html.Button([html.I(className="fas fa-calculator", style={'marginRight':'5px'}), t('composite.calculate', lang)],
                            id='composite-calculate-btn', n_clicks=0,
                            style={'flex':'2', 'padding':'8px', 'backgroundColor':'#555', 'color':'#888', 'border':'none', 'borderRadius':'4px', 'cursor':'not-allowed', 'fontWeight':'bold', 'fontSize':'11px', 'fontFamily':'Barlow Condensed', 'letterSpacing':'1px'})
                    ], style={'display':'flex', 'marginBottom':'10px'}),
                    html.Div(id='composite-status-message', style={'fontSize':'10px', 'padding':'8px', 'borderRadius':'4px', 'display':'none'})
                ]),

            ])
        ], style={'position':'fixed','top':0,'left':'45px','bottom':0,'zIndex':20}),

        # 3. Collapse Button
        html.Button("‹", id="collapse-btn", n_clicks=0, style={'position':'fixed','top':'20px','left':'calc(45px + 18vw)','zIndex':40,'width':'20px','height':'30px','backgroundColor':'#555555','color':'white','border':'none','borderTopRightRadius':'4px','borderBottomRightRadius':'4px','cursor':'pointer','fontSize':'18px','padding':'0','lineHeight':'28px'}),

        # 4. Map & Overlays
        html.Div(map_overlays, id='map-container', style={'position':'fixed', 'top':0, 'left':'calc(45px + 18vw)', 'right':0, 'height':'100vh', 'zIndex':1, 'transition': 'left 0.3s ease'}),

        # 5a. Analysis Toggle Button (fixed position, above pane)
        html.Button([
            html.I(className="fas fa-chart-pie", style={'marginRight': '6px'}),
            t('sidebar.analysis', lang)
        ], id='analysis-toggle-btn', n_clicks=0, style={
            'position':'fixed', 'top':'8px', 'right':'20px', 'zIndex':25,
            'padding':'8px 16px', 'borderRadius':'20px', 'border':'none',
            'backgroundColor':'#44546A', 'color':'white', 'fontSize':'14px',
            'cursor':'pointer', 'fontFamily':'Barlow Condensed', 'letterSpacing':'1px',
            'boxShadow':'0 4px 6px rgba(0,0,0,0.3)', 'transition':'all 0.2s ease'
        }),

        # 5b. Right Analysis Pane (hidden by default, toggled via button)
        html.Div(id='right-pane-container', children=[
            # Resize handles
            html.Div(id='right-pane-resize-left', style={
                'position':'absolute', 'left':0, 'top':0, 'width':'6px',
                'height':'100%', 'cursor':'ew-resize', 'zIndex':2
            }),
            html.Div(id='right-pane-resize-bottom', style={
                'position':'absolute', 'left':0, 'bottom':0, 'width':'100%',
                'height':'6px', 'cursor':'ns-resize', 'zIndex':2
            }),
            html.Div(id='right-pane-resize-corner', style={
                'position':'absolute', 'left':0, 'bottom':0, 'width':'12px',
                'height':'12px', 'cursor':'nesw-resize', 'zIndex':3
            }),
            # Content wrapper (scrollable)
            html.Div(id='right-pane-content', style={
                'width':'100%', 'height':'100%', 'overflowY':'auto',
                'padding':'10px', 'paddingTop':'45px', 'boxSizing':'border-box'
            }, children=[
                html.H3(t('tabs.analysis', lang), style={'borderBottom':'1px solid #555555', 'paddingBottom':'5px', 'fontSize': '16px', 'marginBottom': '15px'}),

                # Data Subset Selector
                html.Div([
                    html.Label(t('analysis.analyze_label', lang), style={'color': '#FBB800', 'fontWeight': 'bold', 'fontSize': '12px', 'marginRight': '10px'}),
                    dcc.RadioItems(
                        id='analysis-subset-selector',
                        options=[
                            {'label': t('analysis.subset_selected', lang), 'value': 'selected'},
                            {'label': t('analysis.subset_filtered', lang), 'value': 'filtered'},
                            {'label': t('analysis.subset_all', lang), 'value': 'all'},
                        ],
                        value='selected',
                        inline=True,
                        className='radio-group',
                        style={'display': 'inline-block'}
                    )
                ], style={'marginBottom': '15px', 'paddingBottom': '10px', 'borderBottom': '1px solid #555'}),

                # Thematic Sections
                html.Div([
                    html.Details([
                        html.Summary(t('analysis.section.population', lang)),
                        html.Div([
                            html.Div(id='settlement-count-text', style={
                                'textAlign': 'center', 'color': '#FBB800', 'fontWeight': 'bold',
                                'fontSize': '12px', 'marginBottom': '4px', 'fontFamily': 'Barlow Condensed'
                            }),
                            html.Div(id='population-total-text', style={
                                'textAlign': 'center', 'color': '#FBB800', 'fontWeight': 'bold',
                                'fontSize': '12px', 'marginBottom': '4px', 'fontFamily': 'Barlow Condensed'
                            }),
                            dcc.Graph(id='sunburst-population', config={'displayModeBar': False}, style={'height': '300px'})
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.electrification', lang)),
                        html.Div([
                            dcc.Graph(id='sunburst-electrification', config={'displayModeBar': False}, style={'height': '300px'})
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.economic', lang)),
                        html.Div([
                            html.Div(id='economic-kpi-row', style={'display':'flex', 'flexWrap':'wrap', 'gap':'4px', 'marginBottom':'8px'}),
                            dcc.Graph(id='chart-livelihoods', config={'displayModeBar': False}, style={'height': '180px'}),
                            html.Div([
                                html.Button('×', id='econ-clear-btn', title=t('analysis.bubble.clear_highlight', lang),
                                    style={'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                                           'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                                           'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                                           'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer',
                                           'display':'none', 'padding':'0'}),
                                dcc.Graph(id='chart-wealth-food', config={'displayModeBar': False, 'scrollZoom': False}, style={'height': '220px'}),
                            ], style={'position':'relative'}),
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.geography', lang)),
                        html.Div([
                            html.Div(id='geography-kpi-row', style={'display':'flex', 'flexWrap':'wrap', 'gap':'4px', 'marginBottom':'8px'}),
                            dcc.Graph(id='chart-climate-class', config={'displayModeBar': False}, style={'height': '180px'}),
                            html.Div([
                                html.Button('×', id='geo-clear-btn', title=t('analysis.bubble.clear_highlight', lang),
                                    style={'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                                           'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                                           'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                                           'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer',
                                           'display':'none', 'padding':'0'}),
                                dcc.Graph(id='chart-cooling-solar', config={'displayModeBar': False, 'scrollZoom': False}, style={'height': '220px'}),
                            ], style={'position':'relative'}),
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.market', lang)),
                        html.Div([
                            html.Div(id='market-kpi-row', style={'display':'flex', 'flexWrap':'wrap', 'gap':'4px', 'marginBottom':'8px'}),
                            dcc.Graph(id='chart-market-indices', config={'displayModeBar': False}, style={'height': '160px'}),
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.ag_demand', lang)),
                        html.Div([
                            dcc.Graph(id='spider-chart-ag', config={'displayModeBar': False}, style={'height': '200px'})
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.ag_production', lang)),
                        html.Div([
                            html.Div(id='prod-total-text', style={
                                'textAlign': 'center', 'color': '#FBB800', 'fontWeight': 'bold',
                                'fontSize': '12px', 'marginBottom': '4px', 'fontFamily': 'Barlow Condensed'
                            }),
                            dcc.Graph(id='spider-chart-prod', config={'displayModeBar': False}, style={'height': '200px'})
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.fish_demand', lang)),
                        html.Div([
                            dcc.Graph(id='spider-chart-fish', config={'displayModeBar': False}, style={'height': '200px'})
                        ], className="analysis-content-inner")
                    ], open=True),
                    html.Details([
                        html.Summary(t('analysis.section.bubble_chart', lang)),
                        html.Div([
                            html.Div([
                                html.Div([
                                    html.Label(t('analysis.bubble.x_axis', lang), style={'fontSize':'10px', 'color':'#bbb', 'marginBottom':'2px'}),
                                    dcc.Dropdown(id='bubble-x-col', options=[{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c} for c in numeric_columns], style={'fontSize':'11px', 'color':'#333'})
                                ], style={'flex':'1', 'marginRight':'4px'}),
                                html.Div([
                                    html.Label(t('analysis.bubble.y_axis', lang), style={'fontSize':'10px', 'color':'#bbb', 'marginBottom':'2px'}),
                                    dcc.Dropdown(id='bubble-y-col', options=[{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c} for c in numeric_columns], style={'fontSize':'11px', 'color':'#333'})
                                ], style={'flex':'1', 'marginLeft':'4px'}),
                            ], style={'display':'flex', 'marginBottom':'5px'}),
                            html.Div([
                                html.Div([
                                    html.Label(t('analysis.bubble.size', lang), style={'fontSize':'10px', 'color':'#bbb', 'marginBottom':'2px'}),
                                    dcc.Dropdown(id='bubble-size-col', options=[{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c} for c in numeric_columns], placeholder=t('analysis.bubble.optional', lang), style={'fontSize':'11px', 'color':'#333'})
                                ], style={'flex':'1', 'marginRight':'4px'}),
                                html.Div([
                                    html.Label(t('analysis.bubble.color', lang), style={'fontSize':'10px', 'color':'#bbb', 'marginBottom':'2px'}),
                                    dcc.Dropdown(id='bubble-color-col', options=[{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c} for c in all_filterable_columns], placeholder=t('analysis.bubble.optional', lang), style={'fontSize':'11px', 'color':'#333'})
                                ], style={'flex':'1', 'marginLeft':'4px'}),
                            ], style={'display':'flex', 'marginBottom':'5px'}),
                            html.Div([
                                html.Button('×', id='bubble-clear-btn', title=t('analysis.bubble.clear_highlight', lang),
                                    style={'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                                           'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                                           'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                                           'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer',
                                           'display':'none', 'padding':'0'}),
                                dcc.Graph(id='bubble-chart', config={'displayModeBar': True, 'scrollZoom': True}, style={'height': '350px'})
                            ], style={'position':'relative'})
                        ], className="analysis-content-inner")
                    ], open=True),

                ])
            ])
        ], style={
            'position':'fixed', 'top':0, 'right':0, 'width':'22vw', 'bottom':0,
            'backgroundColor':'#44546A', 'color':'#ecf0f1',
            'boxShadow':'inset 5px 0 10px -5px rgba(0,0,0,0.5)',
            'zIndex':20, 'display':'none'
        }),

        # 6. Table Interface (full width: from icon bar to right edge)
        html.Div(id='table-ui-wrapper', style={'position':'fixed', 'bottom':0, 'left':'45px', 'right':0, 'zIndex': 50, 'display': 'flex', 'flexDirection': 'column', 'justifyContent': 'flex-end', 'alignItems': 'center', 'pointerEvents': 'none'}, children=[
            # Toggle Button
            html.Button(children=[html.I(className="fas fa-table"), f" {t('table.data_button', lang)}"], id="table-toggle-btn", n_clicks=0, style={'pointerEvents': 'auto', 'marginBottom': '15px', 'padding': '8px 16px', 'borderRadius': '30px', 'border': 'none', 'backgroundColor': '#44546A', 'color': 'white', 'fontSize': '14px', 'boxShadow': '0 4px 6px rgba(0,0,0,0.3)', 'cursor': 'pointer', 'transition': 'transform 0.2s', 'fontFamily': 'Barlow Condensed', 'letterSpacing': '1px'}),

            # Table Panel
            html.Div(id='table-panel', style={'pointerEvents': 'auto', 'width': '100%', 'height': '0vh', 'backgroundColor': '#44546A', 'overflow': 'hidden', 'transition': 'height 0.3s ease-in-out', 'display': 'flex', 'flexDirection': 'column', 'boxShadow': '0 -4px 10px rgba(0,0,0,0.3)'}, children=[

                # Resizer Handle
                html.Div(id="resize-handle"),

                # Header Row (Single compact line with all controls)
                html.Div([
                    html.Button([html.I(className="fas fa-download", style={'marginRight': '5px'}), t('table.export', lang)], id="export-table-btn", n_clicks=0, style={'backgroundColor': '#583CA5', 'color': 'white', 'border': 'none', 'padding': '3px 10px', 'borderRadius': '4px', 'cursor': 'pointer', 'fontSize': '10px', 'fontWeight': 'bold', 'fontFamily': 'Barlow Condensed', 'letterSpacing': '0.5px', 'marginRight': '10px'}),
                    html.Div([
                        html.Label(t('table.layer_label', lang), style={'color': '#FBB800', 'marginRight': '6px', 'fontSize': '10px', 'fontWeight': 'bold', 'fontFamily': 'Barlow Condensed'}),
                        dcc.Dropdown(id='table-layer-dropdown', clearable=False, style={'width': '130px', 'color': '#333', 'fontSize': '10px', 'height': '24px'})
                    ], style={'display': 'flex', 'alignItems': 'center', 'marginRight': '10px', 'borderRight': '1px solid #555', 'paddingRight': '10px'}),
                    html.Div(id='header-stats-container', style={'display': 'flex', 'alignItems': 'center', 'gap': '5px', 'marginRight': '10px'}),
                    html.Button(t('table.select_all', lang), id="select-all-btn", n_clicks=0, style={'marginRight': '4px', 'backgroundColor': '#583CA5', 'color': 'white', 'border': 'none', 'padding': '3px 8px', 'borderRadius': '12px', 'cursor': 'pointer', 'fontSize': '9px', 'fontWeight': 'bold', 'textTransform': 'uppercase'}),
                    html.Button(t('table.deselect', lang), id="deselect-all-btn", n_clicks=0, style={'marginRight': '4px', 'backgroundColor': '#555555', 'color': 'white', 'border': 'none', 'padding': '3px 8px', 'borderRadius': '12px', 'cursor': 'pointer', 'fontSize': '9px', 'fontWeight': 'bold', 'textTransform': 'uppercase'}),
                    html.Div(style={'flexGrow': 1}),
                    html.Button(t('table.close', lang), id="close-table-btn", n_clicks=0, style={'border':'none','background':'transparent','fontSize':'18px','cursor':'pointer', 'color':'#ecf0f1', 'lineHeight': '1'})
                ], style={'padding':'4px 10px','display':'flex','alignItems':'center','borderBottom':'1px solid #555555', 'backgroundColor':'#34495e', 'height': '32px', 'flexShrink': 0}),

                # Table Content (scrollable)
                html.Div(id='table-content', style={'flex':'1', 'overflowY':'auto', 'overflowX':'auto', 'padding':'0'}, children=[
                    html.Div(id='table-placeholder', children=[
                        html.I(className="fas fa-hand-pointer", style={'fontSize': '30px', 'marginBottom': '15px'}),
                        html.H3(t('table.no_points', lang), style={'marginBottom': '5px', 'color': '#ecf0f1'}),
                        html.P(t('table.select_instruction', lang), style={'color': '#bbb', 'fontSize': '12px'})
                    ], style={'display': 'flex', 'flexDirection': 'column', 'alignItems': 'center', 'justifyContent': 'center', 'height': '100%', 'color': '#7f8c8d', 'textAlign': 'center'}),
                    dash_table.DataTable(
                        id={'type': 'table-export', 'index': 0},
                        columns=[{'name': ' ', 'id': '_init'}],
                        data=[],
                        tooltip_header={},
                        tooltip_delay=0,
                        tooltip_duration=None,
                        page_action='custom',
                        page_current=0,
                        page_size=50,
                        page_count=1,
                        sort_action='custom',
                        sort_mode='single',
                        sort_by=[],
                        editable=True,
                        style_table={'height':'100%', 'minWidth':'100%', 'overflowX':'auto', 'display': 'none'},
                        style_cell={'textAlign':'left','padding':'4px 6px','whiteSpace':'normal','height':'auto','fontFamily':'Mulish, sans-serif','fontSize':'10px','backgroundColor':'#44546A','color':'#ecf0f1','border':'1px solid #555555','minWidth':'90px','width':'90px','maxWidth':'180px'},
                        style_header={'fontWeight':'600','backgroundColor':'#583CA5','color':'white','fontSize':'10px','border':'1px solid #555555', 'fontFamily': 'Barlow Condensed', 'letterSpacing': '0.3px'},
                        style_data_conditional=[
                            {'if': {'state': 'selected'}, 'backgroundColor': 'rgba(88, 60, 165, 0.35)', 'color': '#fff', 'border': '1px solid #583CA5'},
                            {'if': {'state': 'active'}, 'backgroundColor': 'rgba(255,200,0,0.25)', 'color': '#fff', 'border': '1px solid rgba(255,200,0,0.6)'},
                        ],
                        css=[{'selector': '.dash-tooltip', 'rule': 'max-width: 320px; white-space: normal; font-size: 11px;'}],
                    ),
                    dcc.Store(id='table-edits-store', data={})
                ])
            ])
        ])
    ])

# ---------------------- CALLBACKS ----------------------

# --- 1. SESSION INITIALIZATION ---
    @app.callback(
        Output('user-session-id', 'data'),
        Input('table-state-store', 'data'), # Trigger
        State('user-session-id', 'data')    # Check existing state
    )
    def create_session_id(_, current_id):
        # FIX: If a session ID already exists, do NOT regenerate it.
        # This prevents the session from resetting when the table is toggled.
        if current_id:
            return dash.no_update
            
        # Generate a unique ID for this browser tab only if one doesn't exist
        return str(uuid.uuid4())

    @app.callback(
        [Output('layer-manager-trigger', 'data', allow_duplicate=True),
         Output('layer-data-content-trigger', 'data', allow_duplicate=True)], # <--- ADDED OUTPUT
        Input('user-session-id', 'data'),
        prevent_initial_call='initial_duplicate'
    )
    def hydrate_new_session(session_id):
        # When session ID is created, load the default "Settlements" data into it
        if not session_id: return dash.no_update, dash.no_update

        dm.initialize_user_session(session_id)

        # Store session_id in Flask session for logout cleanup
        from flask import session as flask_session
        try:
            flask_session['dash_session_id'] = session_id
        except Exception:
            pass

        # Trigger BOTH the UI manager and the Data Content manager
        # This ensures populate_client_store runs immediately after startup
        trigger = np.random.randint(100000)
        return trigger, trigger

    # Store session_id in global JS variable for beforeunload handler
    app.clientside_callback(
        """
        function(sessionId) {
            if (sessionId) { window.__agcap_session_id = sessionId; }
            return window.dash_clientside.no_update;
        }
        """,
        Output('toolbar-dummy-output', 'children', allow_duplicate=True),
        Input('user-session-id', 'data'),
        prevent_initial_call=True
    )

    # Right analysis pane toggle
    app.clientside_callback(
        """
        function(n) {
            const defaultBtn = {'position':'fixed','top':'8px','right':'20px','zIndex':25,'padding':'8px 16px','borderRadius':'20px','border':'none','backgroundColor':'#44546A','color':'white','fontSize':'14px','cursor':'pointer','fontFamily':'Barlow Condensed','letterSpacing':'1px','boxShadow':'0 4px 6px rgba(0,0,0,0.3)','transition':'all 0.2s ease'};
            const activeBtn = Object.assign({}, defaultBtn, {'backgroundColor':'#FBB800','color':'#333'});

            if (!n) return [{'display':'none'}, defaultBtn];

            var st = window.__agcap_state || { tableHeight: '25vh', rightPaneWidth: '22vw', tableIsOpen: false };
            const isOpen = (n % 2 === 1);

            var paneBottom = '0px';
            if (isOpen && st.tableIsOpen) {
                var tp = document.getElementById('table-panel');
                if (tp) paneBottom = tp.offsetHeight + 'px';
            }

            const paneStyle = isOpen
                ? {'position':'fixed','top':0,'right':0,'width':st.rightPaneWidth,'bottom':paneBottom,'backgroundColor':'#44546A','color':'#ecf0f1','boxShadow':'inset 5px 0 10px -5px rgba(0,0,0,0.5)','zIndex':20,'display':'block'}
                : {'display':'none'};

            return [paneStyle, isOpen ? activeBtn : defaultBtn];
        }
        """,
        [Output("right-pane-container", "style"), Output("analysis-toggle-btn", "style")],
        Input("analysis-toggle-btn", "n_clicks")
    )

    # --- TOOLBAR CALLBACKS ---

    # Lasso / Box / Pan drag mode switching
    app.clientside_callback(
        """
        function(lasso_n, box_n, pan_n) {
            const ctx = dash_clientside.callback_context;
            if (!ctx.triggered.length) return window.dash_clientside.no_update;
            const btnId = ctx.triggered[0].prop_id.split('.')[0];
            let mode = 'lasso';
            if (btnId === 'tool-box') mode = 'select';
            else if (btnId === 'tool-pan') mode = 'pan';
            var el = document.getElementById('map');
            if (el) {
                var gd = el._fullLayout ? el : el.querySelector('.js-plotly-plot') || el;
                Plotly.relayout(gd, {'dragmode': mode});
            }
            return mode;
        }
        """,
        Output('active-tool-store', 'data'),
        [Input('tool-lasso', 'n_clicks'), Input('tool-box', 'n_clicks'), Input('tool-pan', 'n_clicks')],
        prevent_initial_call=True
    )

    # Highlight the active tool button (use className, not style, to override CSS)
    app.clientside_callback(
        """
        function(activeTool) {
            var ac = 'toolbar-btn active-tool';
            var ic = 'toolbar-btn';
            return [
                activeTool === 'lasso' ? ac : ic,
                activeTool === 'select' ? ac : ic,
                activeTool === 'pan' ? ac : ic
            ];
        }
        """,
        [Output('tool-lasso', 'className'), Output('tool-box', 'className'), Output('tool-pan', 'className')],
        Input('active-tool-store', 'data')
    )

    # Zoom In / Zoom Out
    app.clientside_callback(
        """
        function(zoomIn_n, zoomOut_n) {
            const ctx = dash_clientside.callback_context;
            if (!ctx.triggered.length) return window.dash_clientside.no_update;
            const btnId = ctx.triggered[0].prop_id.split('.')[0];
            var el = document.getElementById('map');
            if (!el) return window.dash_clientside.no_update;
            var gd = el._fullLayout ? el : el.querySelector('.js-plotly-plot') || el;
            var mb = (gd._fullLayout || {}).map || (gd.layout || {}).map;
            if (!mb) return window.dash_clientside.no_update;
            var zoom = mb.zoom || 5;
            if (btnId === 'tool-zoom-in') zoom = Math.min(zoom + 1, 20);
            else zoom = Math.max(zoom - 1, 1);
            Plotly.relayout(gd, {'map.zoom': zoom});
            return window.dash_clientside.no_update;
        }
        """,
        Output('toolbar-dummy-output', 'children'),
        [Input('tool-zoom-in', 'n_clicks'), Input('tool-zoom-out', 'n_clicks')],
        prevent_initial_call=True
    )

    # Screenshot: composite Plotly map (WebGL) + legend + logos overlays
    app.clientside_callback(
        """
        function(n_clicks) {
            if (!n_clicks) return window.dash_clientside.no_update;
            var mapEl = document.getElementById('map');
            if (!mapEl) return window.dash_clientside.no_update;
            var gd = mapEl._fullLayout ? mapEl : mapEl.querySelector('.js-plotly-plot') || mapEl;
            var W = mapEl.offsetWidth, H = mapEl.offsetHeight;

            // Step 1: get map tiles via Plotly (handles WebGL canvas)
            Plotly.toImage(gd, {format: 'png', width: W, height: H, scale: 2}).then(function(mapDataUrl) {
                var scale = 2;
                var canvas = document.createElement('canvas');
                canvas.width  = W * scale;
                canvas.height = H * scale;
                var ctx = canvas.getContext('2d');

                // Draw map base
                var mapImg = new Image();
                mapImg.onload = function() {
                    ctx.drawImage(mapImg, 0, 0);

                    // Capture overlays sequentially then composite
                    var overlayIds = ['legend-container', 'map-scalebar', 'map-logos'];
                    var promises = overlayIds.map(function(id) {
                        var el = document.getElementById(id);
                        if (!el || el.style.display === 'none') return Promise.resolve(null);
                        return html2canvas(el, {
                            backgroundColor: null,
                            scale: scale,
                            useCORS: true,
                            logging: false
                        }).then(function(ovCanvas) {
                            var rect = el.getBoundingClientRect();
                            var mapRect = mapEl.getBoundingClientRect();
                            return {
                                canvas: ovCanvas,
                                x: (rect.left - mapRect.left) * scale,
                                y: (rect.top  - mapRect.top)  * scale
                            };
                        });
                    });

                    Promise.all(promises).then(function(overlays) {
                        overlays.forEach(function(ov) {
                            if (ov) ctx.drawImage(ov.canvas, ov.x, ov.y);
                        });
                        canvas.toBlob(function(blob) {
                            var url = URL.createObjectURL(blob);
                            var a = document.createElement('a');
                            a.href = url;
                            a.download = 'AgCAP_map_screenshot.png';
                            document.body.appendChild(a);
                            a.click();
                            document.body.removeChild(a);
                            setTimeout(function() { URL.revokeObjectURL(url); }, 1000);
                        }, 'image/png');
                    });
                };
                mapImg.src = mapDataUrl;
            });
            return window.dash_clientside.no_update;
        }
        """,
        Output('toolbar-dummy-output', 'children', allow_duplicate=True),
        Input('tool-screenshot', 'n_clicks'),
        prevent_initial_call=True
    )

    # Scale bar: update on every pan/zoom via relayoutData
    app.clientside_callback(
        """
        function(relayoutData) {
            var BAR_PX = 120;
            // Read current zoom and center lat from the Plotly graph layout
            var gd = document.getElementById('map');
            if (gd) { gd = gd._fullLayout ? gd : gd.querySelector('.js-plotly-plot') || gd; }
            var mapbox = (gd && gd._fullLayout) ? gd._fullLayout.map : null;
            if (!mapbox) return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            var zoom = mapbox.zoom;
            var lat  = mapbox.center ? mapbox.center.lat : 0;

            // Web Mercator: meters per pixel at this zoom and latitude
            var mpp = 156543.034 * Math.cos(lat * Math.PI / 180) / Math.pow(2, zoom);
            var totalM = mpp * BAR_PX;

            // Round to a clean number
            var magnitude = Math.pow(10, Math.floor(Math.log10(totalM)));
            var nice = [1, 2, 5, 10];
            var raw = totalM / magnitude;
            var factor = nice.reduce(function(prev, cur) {
                return Math.abs(cur - raw) < Math.abs(prev - raw) ? cur : prev;
            });
            var niceM = factor * magnitude;

            var label = niceM >= 1000 ? (niceM / 1000) + ' km' : Math.round(niceM) + ' m';

            // Adjust bar width to exactly match the nice distance
            var barW = Math.round(niceM / mpp);
            var barEl = document.getElementById('map-scalebar-bar');
            if (barEl) barEl.style.width = barW + 'px';

            return [label, window.dash_clientside.no_update];
        }
        """,
        [Output('map-scalebar-label', 'children'),
         Output('map-scalebar-bar', 'style')],
        Input('map', 'relayoutData'),
        prevent_initial_call=False
    )

    # Reset View
    app.clientside_callback(
        """
        function(n_clicks, defaultView) {
            if (!n_clicks || !defaultView) return window.dash_clientside.no_update;
            var el = document.getElementById('map');
            if (!el) return window.dash_clientside.no_update;
            var gd = el._fullLayout ? el : el.querySelector('.js-plotly-plot') || el;
            Plotly.relayout(gd, {
                'map.center.lat': defaultView.center.lat,
                'map.center.lon': defaultView.center.lon,
                'map.zoom': defaultView.zoom
            });
            return window.dash_clientside.no_update;
        }
        """,
        Output('toolbar-dummy-output', 'children', allow_duplicate=True),
        Input('tool-reset-view', 'n_clicks'),
        State('default-map-view', 'data'),
        prevent_initial_call=True
    )

    # Reset Platform
    app.clientside_callback(
        """
        async function(n_clicks, sessionId) {
            if (!n_clicks) return window.dash_clientside.no_update;
            if (!confirm('Reset the platform to its default state? All changes will be lost.')) {
                return window.dash_clientside.no_update;
            }
            try {
                const prefix = window.location.pathname.replace(/\\/$/, '');
                await fetch(prefix + '/api/cleanup-session', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({session_id: sessionId})
                });
            } catch(e) { console.warn('Cleanup request failed:', e); }
            localStorage.clear();
            sessionStorage.clear();
            window.location.reload();
            return window.dash_clientside.no_update;
        }
        """,
        Output('toolbar-dummy-output', 'children', allow_duplicate=True),
        Input('tool-reset-platform', 'n_clicks'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )

    # Language modal toggle
    @app.callback(
        Output('language-modal', 'style'),
        [Input('language-btn', 'n_clicks'),
         Input('close-language-modal', 'n_clicks')],
        prevent_initial_call=True
    )
    def toggle_language_modal(lang_clicks, close_clicks):
        ctx = callback_context
        if not ctx.triggered:
            return {'display': 'none'}

        button_id = ctx.triggered[0]['prop_id'].split('.')[0]

        if button_id == 'language-btn':
            return {'display': 'block'}
        else:  # close button
            return {'display': 'none'}

    # Info modal toggle
    @app.callback(
        Output('info-modal', 'style'),
        [Input('info-btn', 'n_clicks'),
         Input('close-info-modal', 'n_clicks')],
        prevent_initial_call=True
    )
    def toggle_info_modal(info_clicks, close_clicks):
        ctx = callback_context
        if not ctx.triggered:
            return {'display': 'none'}
        button_id = ctx.triggered[0]['prop_id'].split('.')[0]
        if button_id == 'info-btn':
            return {'display': 'block'}
        return {'display': 'none'}

    # Export table to CSV (server-side: reads full dataset since table uses custom paging)
    @app.callback(
        Output('download-table-csv', 'data'),
        Input('export-table-btn', 'n_clicks'),
        State('current-selection-store', 'data'),
        State('table-layer-dropdown', 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'id'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def export_table_csv(n_clicks, selected_uuids, active_layer_id, slider_vals, slider_ids, cat_vals, cat_ids, session_id):
        if not n_clicks or not session_id or not active_layer_id:
            return dash.no_update
        df = dm.get_df(session_id, active_layer_id)
        if df is None:
            return dash.no_update
        # Apply filters
        if slider_ids and slider_vals:
            for value, sid in zip(slider_vals, slider_ids):
                col = sid['column']
                if col in df.columns and value:
                    df = df[(df[col] >= value[0]) & (df[col] <= value[1])]
        if cat_ids and cat_vals:
            for values, cid in zip(cat_vals, cat_ids):
                col = cid['column']
                if col in df.columns and values:
                    df = df[df[col].isin(values)]
        # Apply selection
        is_select_all = (selected_uuids and len(selected_uuids) == 1 and selected_uuids[0] == 'ALL')
        if not is_select_all:
            if selected_uuids:
                valid = [u for u in selected_uuids if u in df.index]
                if valid:
                    df = df.loc[valid]
                else:
                    return dash.no_update
            else:
                return dash.no_update
        # Translate and export
        lang = app.lang
        drop_cols = [c for c in ['row_uuid', 'orig_uuid', 'geometry'] if c in df.columns]
        export_df = df.drop(columns=drop_cols)
        translate_map = {col: get_column_translation(col, lang, COUNTRY_CODE) for col in export_df.columns}
        export_df = export_df.rename(columns=translate_map)
        buf = io.StringIO()
        export_df.to_csv(buf, index=False)
        return dict(content=buf.getvalue(), filename='agcap_export.csv')

    # Reset table column filters
    # --- NEW: SEND DATA TO CLIENT STORE (OPTIMIZED) ---
    # --- NEW: SEND ALL VISIBLE LAYERS TO CLIENT STORE ---
    # --- NEW: SEND ALL VISIBLE LAYERS TO CLIENT STORE ---
    @app.callback(
        Output('client-layer-data', 'data'),
        Input('layer-data-content-trigger', 'data'), 
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def populate_client_store(trigger, session_id):
        if not session_id: return None

        user_meta = dm.get_metadata(session_id)
        user_data = dm.get_user_data(session_id)

        layers_payload = {}
        for layer in user_meta:
            lid = layer['id']
            df = dm.get_df(session_id, lid)
            if df is not None:
                dff = df.reset_index()
                data_dict = dff.to_dict('split')
                data_dict['name'] = layer['name']
                data_dict['layer_type'] = layer.get('layer_type', 'point')
                # For polygon layers, include GeoJSON geometry
                if layer.get('layer_type') == 'polygon':
                    geojson = user_data.get('geojson', {}).get(lid)
                    if geojson:
                        data_dict['geojson'] = geojson
                layers_payload[lid] = data_dict

        return layers_payload


# --- CLIENTSIDE: HIDE LOADER WHEN MAP FIGURE IS POPULATED ---
    app.clientside_callback(
        """
        function(figure) {
            // 1. Check if figure exists
            // 2. Check if figure has 'data' (traces) or a specific layout indicating it is ready
            if (figure && figure.data && figure.data.length > 0) {
                
                // OPTIONAL: Add a small delay (e.g., 1000ms) to allow the map tiles 
                // to visually paint before lifting the curtain.
                // You can remove the setTimeout wrapper if you want it instant.
                
                setTimeout(function() {
                    const loader = document.getElementById('initial-loader');
                    if(loader) {
                        loader.style.opacity = '0';
                        loader.style.transition = 'opacity 1s ease-out';
                        
                        // Actually remove it from layout after fade
                        setTimeout(function() {
                            loader.style.display = 'none';
                        }, 1000);
                    }
                }, 1000); // <--- Adjust this 1000 (1 second) to your liking
                
                return window.dash_clientside.no_update;
            }
            return window.dash_clientside.no_update;
        }
        """,
        Output('initial-loader', 'children'), # Dummy output just to make the callback valid
        Input('map', 'figure') # <--- THE KEY: We wait for the map data
    )


    # --- 2. CLIENTSIDE UI ACTIONS (Instant Response) ---
    # Replaces 'switch_tab' (Layers / Filters / Composite — 3 tabs)
    app.clientside_callback(
        """
        function(btn1, btn2, btn3) {
            const ctx = dash_clientside.callback_context;
            const active = {'color': '#FBB800', 'boxShadow': '0 0 8px rgba(251, 184, 0, 0.4)'};
            const inactive = {'color': '#bdc3c7', 'boxShadow': 'none'};
            const show = {'display': 'block'};
            const hide = {'display': 'none'};

            if (!ctx.triggered.length) {
                return [hide, show, hide, inactive, active, inactive];
            }

            const bid = ctx.triggered[0].prop_id.split('.')[0];

            if (bid === 'layers-btn') {
                return [show, hide, hide, active, inactive, inactive];
            } else if (bid === 'composite-btn') {
                return [hide, hide, show, inactive, inactive, active];
            } else {
                return [hide, show, hide, inactive, active, inactive];
            }
        }
        """,
        [Output('layers-content', 'style'), Output('filters-content', 'style'), Output('composite-content', 'style'),
         Output('layers-btn', 'style'), Output('filters-btn', 'style'), Output('composite-btn', 'style')],
        [Input('layers-btn', 'n_clicks'), Input('filters-btn', 'n_clicks'), Input('composite-btn', 'n_clicks')]
    )

    # Replaces 'toggle_layout'
    app.clientside_callback(
        """
        function(n) {
            const collapsed = (n % 2 === 1);
            const left_pos = collapsed ? '45px' : 'calc(45px + 18vw)';

            var paneBottom = '0px';
            if (!collapsed && window.__agcap_state && window.__agcap_state.tableIsOpen) {
                var tp = document.getElementById('table-panel');
                if (tp) paneBottom = tp.offsetHeight + 'px';
            }

            const pane = collapsed
                ? {'display': 'none'}
                : {'position':'fixed','top':0,'left':'45px','bottom':paneBottom,'zIndex':20};

            return [
                pane,
                {'position':'fixed','top':'20px','left':left_pos,'zIndex':40,'width':'20px','height':'30px','backgroundColor':'#555555','color':'white','border':'none','borderTopRightRadius':'4px','borderBottomRightRadius':'4px','cursor':'pointer','fontSize':'18px','padding':'0','lineHeight':'28px'},
                {'position':'fixed', 'top':0, 'left':left_pos, 'right':0, 'height':'100vh', 'zIndex':1, 'transition': 'left 0.3s ease'},
                {'position':'fixed', 'bottom':0, 'left':'45px', 'right':0, 'zIndex': 50, 'display': 'flex', 'flexDirection': 'column', 'justifyContent': 'flex-end', 'alignItems': 'center', 'pointerEvents': 'none'},
                collapsed ? '›' : '‹'
            ];
        }
        """,
        [Output("left-pane-container", "style"), Output("collapse-btn", "style"), Output("map-container", "style"), Output("table-ui-wrapper", "style"), Output("collapse-btn", "children")],
        Input("collapse-btn", "n_clicks")
    )

    # Replaces 'toggle_table'
    # Replaces 'toggle_table' - FIXED (Animation Removed to prevent Crash)
    app.clientside_callback(
        """
        function(btn, close, is_open) {
            const ctx = dash_clientside.callback_context;
            
            // Define Styles (Transition REMOVED to fix RangeError)
            const style_closed = {
                'pointerEvents': 'auto', 
                'marginBottom': '15px', 
                'padding': '8px 16px', 
                'borderRadius': '30px', 
                'border': 'none', 
                'backgroundColor': '#44546A', 
                'color': 'white', 
                'fontSize': '14px', 
                'boxShadow': '0 4px 6px rgba(0,0,0,0.3)', 
                'cursor': 'pointer', 
                'fontFamily': 'Barlow Condensed', 
                'letterSpacing': '1px'
            };
            
            const style_open = {
                'pointerEvents': 'auto', 
                'marginBottom': '10px', 
                'padding': '8px 16px', 
                'borderRadius': '30px', 
                'border': 'none', 
                'backgroundColor': '#FBB800', 
                'color': '#333', 
                'fontSize': '14px', 
                'boxShadow': '0 4px 6px rgba(0,0,0,0.3)', 
                'cursor': 'pointer', 
                'fontFamily': 'Barlow Condensed', 
                'letterSpacing': '1px'
            };

            // Initialize state if needed
            if (!window.__agcap_state) {
                window.__agcap_state = { tableHeight: '25vh', rightPaneWidth: '22vw', tableIsOpen: false };
            }

            // 1. Initial Load: Return closed style immediately
            if (!ctx.triggered.length) {
                window.__agcap_state.tableIsOpen = false;
                return [{
                    'height': '0vh',
                    'width': '100%',
                    'backgroundColor': '#44546A',
                    'overflow': 'hidden',
                    'display': 'flex',
                    'flexDirection': 'column',
                    'boxShadow': '0 -4px 10px rgba(0,0,0,0.3)',
                    'pointerEvents': 'auto'
                }, style_closed, false];
            }

            const trig = ctx.triggered[0].prop_id.split('.')[0];

            // 2. Logic: Toggle if main button, Close if X button
            let new_state = is_open;
            if (trig === 'table-toggle-btn') {
                new_state = !is_open;
            } else if (trig === 'close-table-btn') {
                new_state = false;
            }

            const h = new_state ? window.__agcap_state.tableHeight : '0vh';
            const btn_style = new_state ? style_open : style_closed;
            window.__agcap_state.tableIsOpen = new_state;

            // Update pane bottoms after Dash renders
            setTimeout(function() {
                if (window.__agcap_updatePaneBottoms) window.__agcap_updatePaneBottoms();
            }, 50);

            return [
                {
                    'height': h,
                    'pointerEvents': 'auto',
                    'width': '100%',
                    'backgroundColor': '#44546A',
                    'overflow': 'hidden',
                    'display': 'flex',
                    'flexDirection': 'column',
                    'boxShadow': '0 -4px 10px rgba(0,0,0,0.3)'
                },
                btn_style,
                new_state
            ];
        }
        """,
        [Output("table-panel", "style"), Output("table-toggle-btn", "style"), Output("table-state-store", "data")],
        [Input("table-toggle-btn", "n_clicks"), Input("close-table-btn", "n_clicks")],
        State("table-state-store", "data")
    )

    # --- 3. HELPER FUNCTION FOR SPIDER CHART ---
    def build_spider_figure(df, metrics, title, color, agg='mean', manual_range=None, lang='en'):
        # Translate column names for display
        translated_metrics = [get_column_translation(m, lang, COUNTRY_CODE) for m in metrics]

        # Get language-specific patterns for label stripping
        patterns = get_spider_label_patterns(lang)

        # Strip text based on language patterns (use translated names for labels)
        labels = []
        for translated_m in translated_metrics:
            label = translated_m
            # Remove prefix patterns
            label = label.replace(patterns.get('ag', ''), '')
            label = label.replace(patterns.get('fish', ''), '')
            # Remove production suffix/prefix
            label = label.replace(patterns.get('prod_prefix', ''), '')
            label = label.replace(patterns.get('prod_suffix', ''), '')
            labels.append(label)

        # Compute values using ENGLISH column names (for data lookup in df)
        values = []
        if df.empty:
            values = [0] * len(metrics)
        else:
            for m in metrics:  # Use English column names for lookup
                if m in df.columns:
                    series = pd.to_numeric(df[m], errors='coerce')
                    if agg == 'sum': val = series.sum()
                    else: val = series.mean()
                    clean_val = val if not pd.isna(val) else 0
                    if manual_range == [0, 1]: values.append(min(max(clean_val, 0), 1))
                    else: values.append(clean_val if clean_val > 0 else 0)
                else: values.append(0)

        values.append(values[0])
        labels.append(labels[0])

        if manual_range:
            axis_config = dict(visible=True, range=manual_range, tickfont=dict(color='#bdc3c7', size=8), gridcolor='#555')
        else:
            top_val = max(values) if values else 10
            limit = top_val * 1.1 if top_val > 0 else 10
            axis_config = dict(visible=True, range=[0, limit], tickfont=dict(color='#bdc3c7', size=8), gridcolor='#555')

        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(
            r=values, theta=labels, fill='toself', name=title, line_color=color,
            fillcolor=f"rgba{tuple(int(color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) + (0.4,)}" 
        ))

        fig.update_layout(
            polar=dict(radialaxis=axis_config, angularaxis=dict(tickfont=dict(color='#ecf0f1', size=10), gridcolor='#555', rotation=90), bgcolor='rgba(0,0,0,0)'),
            showlegend=False, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=35, r=35, t=20, b=20), font=dict(family='Mulish', color='#ecf0f1')
        )
        return fig

    # --- 4. PYTHON CALLBACKS (WITH SESSION ID) ---

    @app.callback(
        [Output('spider-chart-ag', 'figure'),
         Output('spider-chart-fish', 'figure'),
         Output('spider-chart-prod', 'figure'),
         Output('prod-total-text', 'children'),
         Output('settlement-count-text', 'children'),
         Output('population-total-text', 'children'),
         Output('sunburst-population', 'figure'),
         Output('sunburst-electrification', 'figure'),
         Output('economic-kpi-row', 'children'),
         Output('chart-livelihoods', 'figure'),
         Output('chart-wealth-food', 'figure'),
         Output('geography-kpi-row', 'children'),
         Output('chart-climate-class', 'figure'),
         Output('chart-cooling-solar', 'figure'),
         Output('market-kpi-row', 'children'),
         Output('chart-market-indices', 'figure')],
        Input('analysis-subset-selector', 'value'),
        Input('layer-manager-trigger', 'data'),
        Input('table-layer-dropdown', 'value'),
        Input({'type': 'histogram-slider', 'column': ALL}, 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'id'),
        Input({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        Input('current-selection-store', 'data'),
        Input('econ-clear-trigger', 'data'),
        Input('geo-clear-trigger', 'data'),
        State('user-session-id', 'data'),
        State('econ-clicked-point', 'data'),
        State('geo-clicked-point', 'data'),
    )
    def update_analysis_charts(subset_mode, trigger, active_layer_id, slider_vals, slider_ids, cat_vals, cat_ids, selected_uuids, econ_clear, geo_clear, session_id, econ_clicked_point, geo_clicked_point):
        lang = app.lang

        empty_fig = go.Figure()
        empty_fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', xaxis={'visible':False}, yaxis={'visible':False})
        empty_text = t('spider_charts.total_production', lang, value='0')
        empty_settle = ''
        empty_pop = ''

        empty_9 = (empty_fig, empty_fig, empty_fig, empty_text, empty_settle, empty_pop, empty_fig, empty_fig,
                   [], empty_fig, empty_fig, [], empty_fig, empty_fig, [], empty_fig)

        def themed_fig():
            f = go.Figure()
            f.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0.15)',
                font=dict(family='Mulish', color='#ecf0f1', size=10),
                margin=dict(l=40, r=10, t=28, b=36),
                xaxis=dict(gridcolor='rgba(255,255,255,0.08)', zerolinecolor='rgba(255,255,255,0.15)'),
                yaxis=dict(gridcolor='rgba(255,255,255,0.08)', zerolinecolor='rgba(255,255,255,0.15)'),
            )
            return f

        def kpi_badge(label, value, unit=''):
            return html.Div([
                html.Div(f"{value}{unit}", style={
                    'fontSize': '15px', 'fontWeight': 'bold', 'color': '#FBB800',
                    'fontFamily': 'Barlow Condensed', 'lineHeight': '1'
                }),
                html.Div(label, style={
                    'fontSize': '9px', 'color': '#bbb', 'lineHeight': '1.2', 'marginTop': '2px'
                })
            ], style={
                'background': 'rgba(0,0,0,0.25)', 'borderRadius': '5px',
                'padding': '5px 7px', 'flex': '1', 'minWidth': '60px', 'textAlign': 'center'
            })

        def safe_mean(series):
            s = pd.to_numeric(series, errors='coerce')
            return s.mean() if not s.isna().all() else None

        def safe_sum(series):
            s = pd.to_numeric(series, errors='coerce')
            return s.sum() if not s.isna().all() else None

        def fmt(val, decimals=1):
            if val is None or (isinstance(val, float) and np.isnan(val)):
                return '—'
            return f"{val:,.{decimals}f}"

        if not session_id or not active_layer_id:
            return empty_9

        # USE SESSION MANAGER
        df = dm.get_df(session_id, active_layer_id)
        if df is None: return empty_9
        
        user_meta = dm.get_metadata(session_id)
        layer_meta = next((l for l in user_meta if l['id'] == active_layer_id), None)
        is_primary = layer_meta.get('is_primary', False) if layer_meta else False

        if subset_mode in ['filtered', 'selected']:
            if is_primary:
                if slider_ids and slider_vals:
                    for value, sid in zip(slider_vals, slider_ids):
                        col = sid['column']
                        if col in df.columns and value:
                            df = df[(df[col]>=value[0]) & (df[col]<=value[1])]
                if cat_ids:
                    for values, cid in zip(cat_vals, cat_ids):
                        col = cid['column']
                        if col in df.columns and values:
                            df = df[df[col].isin(values)]

        if subset_mode == 'selected':
            is_select_all = (selected_uuids and len(selected_uuids) == 1 and selected_uuids[0] == 'ALL')
            if is_select_all:
                pass  # Use all filtered data
            elif selected_uuids:
                df = df[df.index.isin(selected_uuids)]
            else:
                df = df.iloc[0:0]

        # --- SPIDER CHARTS ---
        total_prod = 0
        if not df.empty:
            valid_prod_cols = [c for c in PROD_METRICS if c in df.columns]
            total_prod = df[valid_prod_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum().sum()

        prod_text = t('spider_charts.total_production', lang, value=f"{total_prod:,.0f}")

        fig_ag = build_spider_figure(df, AG_METRICS, t('spider_charts.ag_demand', lang), "#FBB800", agg='mean', manual_range=[0,1], lang=lang)
        fig_fish = build_spider_figure(df, FISH_METRICS, t('spider_charts.fish_demand', lang), "#3498db", agg='mean', manual_range=[0,1], lang=lang)
        fig_prod = build_spider_figure(df, PROD_METRICS, t('spider_charts.production', lang), "#2ecc71", agg='sum', manual_range=None, lang=lang)

        # --- POPULATION & DEMOGRAPHICS ---
        settlement_count = len(df)
        settle_text = t('analysis.total_settlements', lang, value=f"{settlement_count:,}")

        total_pop = 0
        if not df.empty and POPULATION_COLUMN in df.columns:
            total_pop = pd.to_numeric(df[POPULATION_COLUMN], errors='coerce').fillna(0).sum()
        pop_text = t('analysis.total_population', lang, value=f"{total_pop:,.0f}")

        # Population Sunburst (uses ADMIN_LEVELS hierarchy)
        required_cols = ADMIN_LEVELS + [POPULATION_COLUMN]
        if not df.empty and all(c in df.columns for c in required_cols):
            sun_df = df[required_cols].copy()
            sun_df[POPULATION_COLUMN] = pd.to_numeric(sun_df[POPULATION_COLUMN], errors='coerce').fillna(0)
            sun_df = sun_df.groupby(ADMIN_LEVELS, as_index=False)[POPULATION_COLUMN].sum()
            sun_df = sun_df[sun_df[POPULATION_COLUMN] > 0]

            if not sun_df.empty:
                fig_pop_sunburst = px.sunburst(
                    sun_df, path=ADMIN_LEVELS, values=POPULATION_COLUMN,
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                pop_label = t('analysis.population_label', lang)
                fig_pop_sunburst.update_layout(
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=10, r=10, t=10, b=10),
                    font=dict(family='Mulish', color='#ecf0f1', size=10)
                )
                fig_pop_sunburst.update_traces(
                    textinfo='label+value',
                    hovertemplate=f'<b>%{{label}}</b><br>{pop_label}: %{{value:,.0f}}<extra></extra>',
                    insidetextorientation='radial'
                )
            else:
                fig_pop_sunburst = empty_fig
        else:
            fig_pop_sunburst = empty_fig

        # --- ELECTRIFICATION SUNBURST ---
        elec_col = ELECTRIFICATION_COLUMN
        if not df.empty and elec_col in df.columns and POPULATION_COLUMN in df.columns:
            elec_df = df[[elec_col, POPULATION_COLUMN]].copy()
            elec_df[POPULATION_COLUMN] = pd.to_numeric(elec_df[POPULATION_COLUMN], errors='coerce').fillna(0)
            elec_df[elec_col] = pd.to_numeric(elec_df[elec_col], errors='coerce')
            elec_df = elec_df.dropna(subset=[elec_col])
            # Map numeric codes to translated labels: 1 = electrified, 99 = non-electrified
            label_electrified = t('analysis.electrification.electrified', lang)
            label_non_electrified = t('analysis.electrification.non_electrified', lang)
            elec_df['elec_label'] = elec_df[elec_col].map({1.0: label_electrified, 99.0: label_non_electrified})
            elec_df['elec_label'] = elec_df['elec_label'].fillna(elec_df[elec_col].astype(int).astype(str))
            elec_df = elec_df.groupby('elec_label', as_index=False)[POPULATION_COLUMN].sum()
            elec_df = elec_df[elec_df[POPULATION_COLUMN] > 0]

            if not elec_df.empty:
                color_map = {
                    label_electrified: '#2ecc71',
                    label_non_electrified: '#e74c3c',
                }
                fig_elec_sunburst = px.sunburst(
                    elec_df, path=['elec_label'], values=POPULATION_COLUMN,
                    color='elec_label',
                    color_discrete_map=color_map
                )
                pop_label = t('analysis.population_label', lang)
                fig_elec_sunburst.update_layout(
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=10, r=10, t=10, b=10),
                    font=dict(family='Mulish', color='#ecf0f1', size=10),
                    showlegend=False
                )
                fig_elec_sunburst.update_traces(
                    textinfo='label+percent parent',
                    hovertemplate=f'<b>%{{label}}</b><br>{pop_label}: %{{value:,.0f}}<br>%{{percentRoot:.1%}}<extra></extra>'
                )
            else:
                fig_elec_sunburst = empty_fig
        else:
            fig_elec_sunburst = empty_fig

        # ── ECONOMIC, LIVELIHOODS & RISKS ────────────────────────────────────────
        wealth_mean   = safe_mean(df.get('Relative Wealth Index'))
        food_mean_raw = safe_mean(df.get('Percentage of food-insecure people'))
        food_mean     = food_mean_raw * 100 if food_mean_raw is not None else None
        conflict_sum  = safe_sum(df.get('Conflict fatalities nearby'))
        cyclone_mean  = safe_mean(df.get('Cyclone hazard (1 in 100 years wind speed)'))

        economic_kpis = [
            kpi_badge(t('analysis.economic.avg_wealth', lang),         fmt(wealth_mean, 2)),
            kpi_badge(t('analysis.economic.food_insecure', lang),       fmt(food_mean,  1), '%'),
            kpi_badge(t('analysis.economic.conflict_fatalities', lang), fmt(conflict_sum, 0) if conflict_sum is not None else '—'),
            kpi_badge(t('analysis.economic.cyclone_risk', lang),        fmt(cyclone_mean, 0), ' km/h'),
        ]

        # Livelihoods bar chart
        if not df.empty and 'Livelihoods' in df.columns:
            lv_counts = df['Livelihoods'].dropna().value_counts().sort_values(ascending=True)
            fig_livelihoods = themed_fig()
            fig_livelihoods.add_trace(go.Bar(
                y=lv_counts.index.tolist(),
                x=lv_counts.values.tolist(),
                orientation='h',
                marker_color='#FBB800',
                hovertemplate='<b>%{y}</b><br>Settlements: %{x:,}<extra></extra>'
            ))
            fig_livelihoods.update_layout(
                title=dict(text=t('analysis.economic.livelihoods_title', lang), font=dict(size=10), x=0, xanchor='left'),
                margin=dict(l=10, r=10, t=28, b=10),
                xaxis=dict(title='', gridcolor='rgba(255,255,255,0.08)'),
                yaxis=dict(title='', automargin=True),
            )
        else:
            fig_livelihoods = empty_fig

        # Wealth vs Population Density scatter (bubble = population, click-to-map)
        econ_scatter_cols = ['Relative Wealth Index', 'Population density']
        econ_avail_extra = [c for c in ['lat', 'lon', POPULATION_COLUMN] + list(ADMIN_LEVELS) if c in df.columns]
        if not df.empty and all(c in df.columns for c in econ_scatter_cols):
            es_df = df[econ_scatter_cols + econ_avail_extra].copy()
            es_df['_uuid'] = df.index
            for c in econ_scatter_cols:
                es_df[c] = pd.to_numeric(es_df[c], errors='coerce')
            if POPULATION_COLUMN in es_df.columns:
                es_df[POPULATION_COLUMN] = pd.to_numeric(es_df[POPULATION_COLUMN], errors='coerce')
            es_df = es_df.dropna(subset=econ_scatter_cols)
            if not es_df.empty:
                pop_s = es_df[POPULATION_COLUMN].fillna(0) if POPULATION_COLUMN in es_df.columns else pd.Series([6]*len(es_df))
                max_pop = pop_s.max() or 1
                sizes = (pop_s / max_pop * 18 + 4).tolist()
                lat_vals = es_df['lat'].tolist() if 'lat' in es_df.columns else [None]*len(es_df)
                lon_vals = es_df['lon'].tolist() if 'lon' in es_df.columns else [None]*len(es_df)
                econ_admin_cols = [ac for ac in ADMIN_LEVELS if ac in es_df.columns]
                full_customdata = []
                for i, (uuid_v, lat_v, lon_v, pop_v) in enumerate(zip(
                    es_df['_uuid'].tolist(), lat_vals, lon_vals, pop_s.tolist()
                )):
                    full_customdata.append([uuid_v, lat_v, lon_v, pop_v] + [es_df.iloc[i][ac] if ac in es_df.columns else '' for ac in econ_admin_cols])
                hover_parts = [
                    '<b>ID: %{customdata[0]}</b>',
                    'Wealth: %{x:.2f}',
                    'Pop. density: %{y:,.1f} pp/km²',
                    'Population: %{customdata[3]:,.0f}',
                ]
                for i, ac in enumerate(econ_admin_cols):
                    hover_parts.append(f'{ac}: %{{customdata[{4+i}]}}')
                hover_parts.append('<extra></extra>')
                fig_wealth_food = themed_fig()
                fig_wealth_food.add_trace(go.Scatter(
                    x=es_df['Relative Wealth Index'].tolist(),
                    y=es_df['Population density'].tolist(),
                    mode='markers',
                    marker=dict(size=sizes, color='#FBB800', opacity=0.65,
                                line=dict(width=0.5, color='rgba(255,255,255,0.3)')),
                    customdata=full_customdata,
                    hovertemplate='<br>'.join(hover_parts),
                    name='settlements',
                    selected=dict(marker=dict(color='#FBB800', opacity=1)),
                    unselected=dict(marker=dict(color='#FBB800', opacity=0.15)),
                ))
                if econ_clicked_point and econ_clicked_point.get('uuid'):
                    hl_row = es_df[es_df['_uuid'] == econ_clicked_point['uuid']]
                    if not hl_row.empty:
                        r = hl_row.iloc[0]
                        fig_wealth_food.add_trace(go.Scatter(
                            x=[r['Relative Wealth Index']], y=[r['Population density']],
                            mode='markers',
                            marker=dict(size=20, color='rgba(0,255,136,0.15)',
                                        line=dict(width=2, color='#00ff88')),
                            showlegend=False, hoverinfo='skip', name='highlight'
                        ))
                fig_wealth_food.update_layout(
                    title=dict(text=t('analysis.economic.wealth_pop_title', lang), font=dict(size=10), x=0, xanchor='left'),
                    showlegend=False,
                    xaxis=dict(title='Relative Wealth Index', gridcolor='rgba(255,255,255,0.08)', zerolinecolor='rgba(255,255,255,0.2)'),
                    yaxis=dict(title='Pop. density (pp/km²)', gridcolor='rgba(255,255,255,0.08)', zerolinecolor='rgba(255,255,255,0.2)'),
                )
            else:
                fig_wealth_food = empty_fig
        else:
            fig_wealth_food = empty_fig

        # ── PHYSICAL GEOGRAPHY & CLIMATE ─────────────────────────────────────────
        elev_mean   = safe_mean(df.get('Elevation'))
        temp_mean   = safe_mean(df.get('Mean temperature'))
        precip_mean = safe_mean(df.get('Precipitation'))
        solar_mean  = safe_mean(df.get('Solar PV Output (kWh/year per installed kW)'))

        geography_kpis = [
            kpi_badge(t('analysis.geography.avg_elevation', lang), fmt(elev_mean, 0), ' m'),
            kpi_badge(t('analysis.geography.avg_temp', lang),      fmt(temp_mean, 1), '°C'),
            kpi_badge(t('analysis.geography.avg_precip', lang),    fmt(precip_mean, 0), ' mm'),
            kpi_badge(t('analysis.geography.avg_solar', lang),     fmt(solar_mean, 0), ' kWh'),
        ]

        # Climate class distribution bar
        if not df.empty and 'Climate class' in df.columns:
            cc_counts = df['Climate class'].dropna().value_counts().sort_values(ascending=True)
            fig_climate = themed_fig()
            fig_climate.add_trace(go.Bar(
                y=cc_counts.index.tolist(),
                x=cc_counts.values.tolist(),
                orientation='h',
                marker_color='#3498db',
                hovertemplate='<b>%{y}</b><br>Settlements: %{x:,}<extra></extra>'
            ))
            fig_climate.update_layout(
                title=dict(text=t('analysis.geography.climate_title', lang), font=dict(size=10), x=0, xanchor='left'),
                margin=dict(l=10, r=10, t=28, b=10),
                xaxis=dict(title='', gridcolor='rgba(255,255,255,0.08)'),
                yaxis=dict(title='', automargin=True),
            )
        else:
            fig_climate = empty_fig

        # Cooling Degree Days vs Solar PV (bubble = population, click-to-map)
        geo_scatter_cols = ['Cooling Degree Days Cold Room at 4C', 'Solar PV Output (kWh/year per installed kW)']
        geo_avail_extra = [c for c in ['lat', 'lon', POPULATION_COLUMN] + list(ADMIN_LEVELS) if c in df.columns]
        if not df.empty and all(c in df.columns for c in geo_scatter_cols):
            gs_df = df[geo_scatter_cols + geo_avail_extra].copy()
            gs_df['_uuid'] = df.index
            for c in geo_scatter_cols:
                gs_df[c] = pd.to_numeric(gs_df[c], errors='coerce')
            if POPULATION_COLUMN in gs_df.columns:
                gs_df[POPULATION_COLUMN] = pd.to_numeric(gs_df[POPULATION_COLUMN], errors='coerce')
            gs_df = gs_df.dropna(subset=geo_scatter_cols)
            if not gs_df.empty:
                pop_s = gs_df[POPULATION_COLUMN].fillna(0) if POPULATION_COLUMN in gs_df.columns else pd.Series([6]*len(gs_df))
                max_pop = pop_s.max() or 1
                sizes = (pop_s / max_pop * 18 + 4).tolist()
                lat_vals = gs_df['lat'].tolist() if 'lat' in gs_df.columns else [None]*len(gs_df)
                lon_vals = gs_df['lon'].tolist() if 'lon' in gs_df.columns else [None]*len(gs_df)
                geo_admin_cols = [ac for ac in ADMIN_LEVELS if ac in gs_df.columns]
                full_customdata = []
                for i, (uuid_v, lat_v, lon_v, pop_v) in enumerate(zip(
                    gs_df['_uuid'].tolist(), lat_vals, lon_vals, pop_s.tolist()
                )):
                    full_customdata.append([uuid_v, lat_v, lon_v, pop_v] + [gs_df.iloc[i][ac] if ac in gs_df.columns else '' for ac in geo_admin_cols])
                hover_parts = [
                    '<b>ID: %{customdata[0]}</b>',
                    'Cooling DD: %{x:,.0f}°C·days',
                    'Solar PV: %{y:,.0f} kWh/kW/yr',
                    'Population: %{customdata[3]:,.0f}',
                ]
                for i, ac in enumerate(geo_admin_cols):
                    hover_parts.append(f'{ac}: %{{customdata[{4+i}]}}')
                hover_parts.append('<extra></extra>')
                fig_cooling_solar = themed_fig()
                fig_cooling_solar.add_trace(go.Scatter(
                    x=gs_df['Cooling Degree Days Cold Room at 4C'].tolist(),
                    y=gs_df['Solar PV Output (kWh/year per installed kW)'].tolist(),
                    mode='markers',
                    marker=dict(size=sizes, color='#3498db', opacity=0.65,
                                line=dict(width=0.5, color='rgba(255,255,255,0.3)')),
                    customdata=full_customdata,
                    hovertemplate='<br>'.join(hover_parts),
                    name='settlements',
                    selected=dict(marker=dict(color='#3498db', opacity=1)),
                    unselected=dict(marker=dict(color='#3498db', opacity=0.15)),
                ))
                if geo_clicked_point and geo_clicked_point.get('uuid'):
                    hl_row = gs_df[gs_df['_uuid'] == geo_clicked_point['uuid']]
                    if not hl_row.empty:
                        r = hl_row.iloc[0]
                        fig_cooling_solar.add_trace(go.Scatter(
                            x=[r['Cooling Degree Days Cold Room at 4C']],
                            y=[r['Solar PV Output (kWh/year per installed kW)']],
                            mode='markers',
                            marker=dict(size=20, color='rgba(0,255,136,0.15)',
                                        line=dict(width=2, color='#00ff88')),
                            showlegend=False, hoverinfo='skip', name='highlight'
                        ))
                fig_cooling_solar.update_layout(
                    title=dict(text=t('analysis.geography.cooling_solar_title', lang), font=dict(size=10), x=0, xanchor='left'),
                    showlegend=False,
                    xaxis=dict(title='Cooling Degree Days (4°C)', gridcolor='rgba(255,255,255,0.08)'),
                    yaxis=dict(title='Solar PV Output (kWh/kW/yr)', gridcolor='rgba(255,255,255,0.08)'),
                )
            else:
                fig_cooling_solar = empty_fig
        else:
            fig_cooling_solar = empty_fig

        # ── MARKET ACCESSIBILITY ──────────────────────────────────────────────────
        mkt_cols = ['Export Market Accessibility', 'National Market Accessibility', 'Fresh Markets Accessibility']
        mkt_present = [c for c in mkt_cols if c in df.columns]
        market_kpis = [
            kpi_badge(get_column_translation(c, lang, COUNTRY_CODE), fmt(safe_mean(df.get(c)), 3))
            for c in mkt_present
        ]

        mkt_index_cols = ['Export Market Accessibility', 'National Market Accessibility',
                          'Fresh Markets Accessibility', 'Farming Activity']
        if not df.empty and any(c in df.columns for c in mkt_index_cols[:3]):
            idx_cols  = [c for c in mkt_index_cols if c in df.columns]
            idx_vals  = [safe_mean(df[c]) or 0 for c in idx_cols]
            idx_labels = [get_column_translation(c, lang, COUNTRY_CODE) for c in idx_cols]
            idx_colors = ['#e67e22' if 'Export' in c else '#FBB800' if 'National' in c else '#3498db' if 'Fresh' in c else '#2ecc71' for c in idx_cols]
            fig_market_indices = themed_fig()
            fig_market_indices.add_trace(go.Bar(
                y=idx_labels, x=idx_vals, orientation='h',
                marker_color=idx_colors,
                text=[f"{v:.3f}" for v in idx_vals],
                textposition='outside',
                textfont=dict(size=9, color='#ecf0f1'),
                hovertemplate='<b>%{y}</b><br>Avg score: %{x:.3f}<extra></extra>'
            ))
            fig_market_indices.update_layout(
                title=dict(text=t('analysis.market.indices_title', lang), font=dict(size=10), x=0, xanchor='left'),
                margin=dict(l=10, r=40, t=28, b=10),
                xaxis=dict(title='', range=[0, 1.1], gridcolor='rgba(255,255,255,0.08)'),
                yaxis=dict(title='', automargin=True),
            )
        else:
            fig_market_indices = empty_fig

        return (fig_ag, fig_fish, fig_prod, prod_text, settle_text, pop_text, fig_pop_sunburst, fig_elec_sunburst,
                economic_kpis, fig_livelihoods, fig_wealth_food,
                geography_kpis, fig_climate, fig_cooling_solar,
                market_kpis, fig_market_indices)

    # --- BUBBLE CHART CALLBACK ---
    @app.callback(
        Output('bubble-chart', 'figure'),
        Input('bubble-x-col', 'value'),
        Input('bubble-y-col', 'value'),
        Input('bubble-size-col', 'value'),
        Input('bubble-color-col', 'value'),
        Input('analysis-subset-selector', 'value'),
        Input('layer-manager-trigger', 'data'),
        Input('table-layer-dropdown', 'value'),
        Input({'type': 'histogram-slider', 'column': ALL}, 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'id'),
        Input({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        Input('current-selection-store', 'data'),
        Input('bubble-clear-trigger', 'data'),
        State('user-session-id', 'data'),
        State('bubble-clicked-point', 'data'),
    )
    def update_bubble_chart(x_col, y_col, size_col, color_col, subset_mode, trigger,
                            active_layer_id, slider_vals, slider_ids, cat_vals, cat_ids,
                            selected_uuids, bubble_clear, session_id, bubble_clicked):
        lang = app.lang
        empty_fig = go.Figure()
        empty_fig.update_layout(
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            xaxis={'visible': False}, yaxis={'visible': False},
            annotations=[dict(text=t('analysis.bubble.select_columns', lang),
                              xref='paper', yref='paper', x=0.5, y=0.5, showarrow=False,
                              font=dict(size=11, color='#888', family='Mulish'))]
        )

        if not x_col or not y_col or not session_id or not active_layer_id:
            return empty_fig

        df = dm.get_df(session_id, active_layer_id)
        if df is None:
            return empty_fig

        user_meta = dm.get_metadata(session_id)
        layer_meta = next((l for l in user_meta if l['id'] == active_layer_id), None)
        is_primary = layer_meta.get('is_primary', False) if layer_meta else False

        # Apply subset logic (same as analysis charts)
        if subset_mode in ['filtered', 'selected']:
            if is_primary:
                if slider_ids and slider_vals:
                    for value, sid in zip(slider_vals, slider_ids):
                        col = sid['column']
                        if col in df.columns and value:
                            df = df[(df[col] >= value[0]) & (df[col] <= value[1])]
                if cat_ids:
                    for values, cid in zip(cat_vals, cat_ids):
                        col = cid['column']
                        if col in df.columns and values:
                            df = df[df[col].isin(values)]

        if subset_mode == 'selected':
            is_select_all = (selected_uuids and len(selected_uuids) == 1 and selected_uuids[0] == 'ALL')
            if is_select_all:
                pass  # Use all filtered data
            elif selected_uuids:
                df = df[df.index.isin(selected_uuids)]
            else:
                df = df.iloc[0:0]

        if df.empty or x_col not in df.columns or y_col not in df.columns:
            return empty_fig

        # Include uuid/lat/lon in customdata for click-to-map-zoom (assign returns new DF)
        df = df.assign(_uuid=df.index)
        custom_cols = ['_uuid', 'lat', 'lon']

        # Build scatter kwargs
        scatter_kwargs = dict(x=x_col, y=y_col, opacity=0.7, custom_data=custom_cols)
        if size_col and size_col in df.columns:
            scatter_kwargs['size'] = size_col
            scatter_kwargs['size_max'] = 30
        if color_col and color_col in df.columns:
            scatter_kwargs['color'] = color_col
            if pd.api.types.is_numeric_dtype(df[color_col]):
                scatter_kwargs['color_continuous_scale'] = 'YlOrRd'

        # Add admin hover info
        hover_cols = [c for c in ADMIN_LEVELS if c in df.columns]
        if hover_cols:
            scatter_kwargs['hover_data'] = hover_cols

        fig = px.scatter(df, **scatter_kwargs)

        x_label = get_column_translation(x_col, lang, COUNTRY_CODE)
        y_label = get_column_translation(y_col, lang, COUNTRY_CODE)

        layout_kwargs = dict(
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(30,30,30,0.3)',
            font=dict(family='Mulish', color='#ecf0f1', size=10),
            # top margin leaves room for the Plotly modebar so it doesn't overlap the chart
            margin=dict(l=40, r=10, t=30, b=40),
            xaxis=dict(title=x_label, gridcolor='rgba(255,255,255,0.1)', zerolinecolor='rgba(255,255,255,0.2)'),
            yaxis=dict(title=y_label, gridcolor='rgba(255,255,255,0.1)', zerolinecolor='rgba(255,255,255,0.2)'),
            dragmode='zoom',
        )

        # Legend / colorbar styling — horizontal below chart to save horizontal space
        if color_col and color_col in df.columns:
            color_label = get_column_translation(color_col, lang, COUNTRY_CODE)
            if pd.api.types.is_numeric_dtype(df[color_col]):
                layout_kwargs['showlegend'] = False
                layout_kwargs['coloraxis_colorbar'] = dict(
                    orientation='h',
                    thickness=10, len=0.9,
                    x=0.5, xanchor='center',
                    y=-0.22, yanchor='top',
                    title=dict(text=color_label, side='top', font=dict(size=9)),
                    tickfont=dict(size=8),
                )
                # Extra bottom margin for horizontal colorbar
                layout_kwargs['margin']['b'] = 80
            else:
                layout_kwargs['showlegend'] = True
                layout_kwargs['legend'] = dict(
                    orientation='h',
                    title=dict(text=color_label, side='left', font=dict(size=9)),
                    font=dict(size=8), bgcolor='rgba(0,0,0,0)',
                    itemsizing='constant',
                    x=0, y=-0.18, xanchor='left', yanchor='top',
                )
                layout_kwargs['margin']['b'] = 70
        else:
            layout_kwargs['showlegend'] = False

        fig.update_layout(**layout_kwargs)

        # Selected / unselected styling for click highlight
        fig.update_traces(
            selected=dict(marker=dict(opacity=1)),
            unselected=dict(marker=dict(opacity=0.3)),
        )

        if not color_col:
            fig.update_traces(marker=dict(color='#FBB800'))

        # Add highlight trace for bubble-clicked point
        if bubble_clicked and bubble_clicked.get('uuid'):
            hl_uuid = bubble_clicked['uuid']
            matching = df[df.index == hl_uuid]
            if not matching.empty:
                row = matching.iloc[0]
                fig.add_trace(go.Scatter(
                    x=[row[x_col]], y=[row[y_col]],
                    mode='markers',
                    marker=dict(size=18, color='rgba(0, 255, 136, 0.15)',
                                line=dict(width=2, color='#00ff88')),
                    showlegend=False, hoverinfo='skip', name='highlight'
                ))

        return fig

    # --- BUBBLE CHART CLICK → MAP ZOOM + HIGHLIGHT ---
    app.clientside_callback(
        """
        function(clickData) {
            if (!clickData || !clickData.points || !clickData.points.length) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var pt = clickData.points[0];
            if (!pt.customdata) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var uuid = pt.customdata[0];
            var lat = pt.customdata[1];
            var lon = pt.customdata[2];
            if (lat == null || lon == null) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }

            // Immediate visual highlight on bubble chart via Plotly.restyle
            var gd = document.getElementById('bubble-chart');
            var plotGd = gd && gd._fullLayout ? gd : (gd ? gd.querySelector('.js-plotly-plot') || gd : null);
            if (plotGd && plotGd._fullData) {
                var selPts = plotGd._fullData.map(function(trace, i) {
                    if (trace.name === 'highlight') return null;
                    var idx = -1;
                    if (trace.customdata) {
                        for (var j = 0; j < trace.customdata.length; j++) {
                            if (trace.customdata[j][0] === uuid) { idx = j; break; }
                        }
                    }
                    return idx >= 0 ? [idx] : [];
                });
                Plotly.restyle(plotGd, {selectedpoints: selPts});
            }

            return [
                {center: {lat: lat, lon: lon}, zoom: 12},
                {uuid: uuid, lat: lat, lon: lon}
            ];
        }
        """,
        [Output('map-view-store', 'data', allow_duplicate=True),
         Output('bubble-clicked-point', 'data')],
        Input('bubble-chart', 'clickData'),
        prevent_initial_call=True
    )

    # --- CLEAR BUBBLE HIGHLIGHT (× button) ---
    app.clientside_callback(
        """
        function(n_clicks) {
            if (!n_clicks) return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            // Clear restyle on bubble chart
            var gd = document.getElementById('bubble-chart');
            var plotGd = gd && gd._fullLayout ? gd : (gd ? gd.querySelector('.js-plotly-plot') || gd : null);
            if (plotGd && plotGd._fullData) {
                var nullSel = plotGd._fullData.map(function() { return null; });
                Plotly.restyle(plotGd, {selectedpoints: nullSel});
            }
            return [null, (window.__bubble_clear_count || 0) + 1];
        }
        """,
        [Output('bubble-clicked-point', 'data', allow_duplicate=True),
         Output('bubble-clear-trigger', 'data', allow_duplicate=True)],
        Input('bubble-clear-btn', 'n_clicks'),
        prevent_initial_call=True
    )

    # --- Show/hide × button based on bubble-clicked-point ---
    app.clientside_callback(
        """
        function(bubble_clicked) {
            if (bubble_clicked && bubble_clicked.uuid) {
                return {'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                        'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                        'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                        'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer',
                        'display':'block', 'padding':'0'};
            }
            return {'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                    'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                    'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                    'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer',
                    'display':'none', 'padding':'0'};
        }
        """,
        Output('bubble-clear-btn', 'style'),
        Input('bubble-clicked-point', 'data')
    )

    # --- ECON CHART CLICK → MAP ZOOM + GREEN PIN + HIGHLIGHT ---
    app.clientside_callback(
        """
        function(clickData) {
            if (!clickData || !clickData.points || !clickData.points.length) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var pt = clickData.points[0];
            if (!pt.customdata) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var uuid = pt.customdata[0];
            var lat  = pt.customdata[1];
            var lon  = pt.customdata[2];
            if (lat == null || lon == null) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var gd = document.getElementById('chart-wealth-food');
            var plotGd = gd && gd._fullLayout ? gd : (gd ? gd.querySelector('.js-plotly-plot') || gd : null);
            if (plotGd && plotGd._fullData) {
                var selPts = plotGd._fullData.map(function(trace) {
                    if (trace.name === 'highlight') return null;
                    var idx = -1;
                    if (trace.customdata) {
                        for (var j = 0; j < trace.customdata.length; j++) {
                            if (trace.customdata[j][0] === uuid) { idx = j; break; }
                        }
                    }
                    return idx >= 0 ? [idx] : [];
                });
                Plotly.restyle(plotGd, {selectedpoints: selPts});
            }
            return [
                {center: {lat: lat, lon: lon}, zoom: 12},
                {uuid: uuid, lat: lat, lon: lon},
                {uuid: uuid, lat: lat, lon: lon}
            ];
        }
        """,
        [Output('map-view-store', 'data', allow_duplicate=True),
         Output('econ-clicked-point', 'data'),
         Output('bubble-clicked-point', 'data', allow_duplicate=True)],
        Input('chart-wealth-food', 'clickData'),
        prevent_initial_call=True
    )

    # --- CLEAR ECON HIGHLIGHT ---
    app.clientside_callback(
        """
        function(n_clicks) {
            if (!n_clicks) return [window.dash_clientside.no_update, window.dash_clientside.no_update, null];
            var gd = document.getElementById('chart-wealth-food');
            var plotGd = gd && gd._fullLayout ? gd : (gd ? gd.querySelector('.js-plotly-plot') || gd : null);
            if (plotGd && plotGd._fullData) {
                var nullSel = plotGd._fullData.map(function() { return null; });
                Plotly.restyle(plotGd, {selectedpoints: nullSel});
            }
            return [null, (window.__econ_clear_count || 0) + 1, null];
        }
        """,
        [Output('econ-clicked-point', 'data', allow_duplicate=True),
         Output('econ-clear-trigger', 'data', allow_duplicate=True),
         Output('bubble-clicked-point', 'data', allow_duplicate=True)],
        Input('econ-clear-btn', 'n_clicks'),
        prevent_initial_call=True
    )

    app.clientside_callback(
        """
        function(econ_clicked) {
            var base = {'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                        'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                        'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                        'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer', 'padding':'0'};
            if (econ_clicked && econ_clicked.uuid) {
                return Object.assign({}, base, {'display':'block'});
            }
            return Object.assign({}, base, {'display':'none'});
        }
        """,
        Output('econ-clear-btn', 'style'),
        Input('econ-clicked-point', 'data')
    )

    # --- GEO CHART CLICK → MAP ZOOM + GREEN PIN + HIGHLIGHT ---
    app.clientside_callback(
        """
        function(clickData) {
            if (!clickData || !clickData.points || !clickData.points.length) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var pt = clickData.points[0];
            if (!pt.customdata) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var uuid = pt.customdata[0];
            var lat  = pt.customdata[1];
            var lon  = pt.customdata[2];
            if (lat == null || lon == null) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var gd = document.getElementById('chart-cooling-solar');
            var plotGd = gd && gd._fullLayout ? gd : (gd ? gd.querySelector('.js-plotly-plot') || gd : null);
            if (plotGd && plotGd._fullData) {
                var selPts = plotGd._fullData.map(function(trace) {
                    if (trace.name === 'highlight') return null;
                    var idx = -1;
                    if (trace.customdata) {
                        for (var j = 0; j < trace.customdata.length; j++) {
                            if (trace.customdata[j][0] === uuid) { idx = j; break; }
                        }
                    }
                    return idx >= 0 ? [idx] : [];
                });
                Plotly.restyle(plotGd, {selectedpoints: selPts});
            }
            return [
                {center: {lat: lat, lon: lon}, zoom: 12},
                {uuid: uuid, lat: lat, lon: lon},
                {uuid: uuid, lat: lat, lon: lon}
            ];
        }
        """,
        [Output('map-view-store', 'data', allow_duplicate=True),
         Output('geo-clicked-point', 'data'),
         Output('bubble-clicked-point', 'data', allow_duplicate=True)],
        Input('chart-cooling-solar', 'clickData'),
        prevent_initial_call=True
    )

    # --- CLEAR GEO HIGHLIGHT ---
    app.clientside_callback(
        """
        function(n_clicks) {
            if (!n_clicks) return [window.dash_clientside.no_update, window.dash_clientside.no_update, null];
            var gd = document.getElementById('chart-cooling-solar');
            var plotGd = gd && gd._fullLayout ? gd : (gd ? gd.querySelector('.js-plotly-plot') || gd : null);
            if (plotGd && plotGd._fullData) {
                var nullSel = plotGd._fullData.map(function() { return null; });
                Plotly.restyle(plotGd, {selectedpoints: nullSel});
            }
            return [null, (window.__geo_clear_count || 0) + 1, null];
        }
        """,
        [Output('geo-clicked-point', 'data', allow_duplicate=True),
         Output('geo-clear-trigger', 'data', allow_duplicate=True),
         Output('bubble-clicked-point', 'data', allow_duplicate=True)],
        Input('geo-clear-btn', 'n_clicks'),
        prevent_initial_call=True
    )

    app.clientside_callback(
        """
        function(geo_clicked) {
            var base = {'position':'absolute', 'top':'4px', 'right':'4px', 'zIndex':10,
                        'background':'rgba(255,80,80,0.85)', 'color':'#fff', 'border':'none',
                        'borderRadius':'50%', 'width':'20px', 'height':'20px', 'fontSize':'14px',
                        'lineHeight':'18px', 'textAlign':'center', 'cursor':'pointer', 'padding':'0'};
            if (geo_clicked && geo_clicked.uuid) {
                return Object.assign({}, base, {'display':'block'});
            }
            return Object.assign({}, base, {'display':'none'});
        }
        """,
        Output('geo-clear-btn', 'style'),
        Input('geo-clicked-point', 'data')
    )

    # --- TABLE ROW CLICK → MAP ZOOM + GREEN PIN ---
    app.clientside_callback(
        """
        function(active_cell, table_data, store_data, active_layer_id) {
            if (!active_cell || !table_data || !store_data || !active_layer_id) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var row = table_data[active_cell.row];
            if (!row || !row.orig_uuid) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }
            var uuid = String(row.orig_uuid);

            var layerObj = store_data[active_layer_id];
            if (!layerObj) return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            var columns = layerObj.columns;
            var colIdx = {};
            columns.forEach(function(c, i) { colIdx[c] = i; });
            var uuidIdx = colIdx['row_uuid'] !== undefined ? colIdx['row_uuid'] : 0;
            var latIdx = colIdx['lat'];
            var lonIdx = colIdx['lon'];
            if (latIdx === undefined || lonIdx === undefined) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }

            var rawData = layerObj.data;
            var lat = null, lon = null;
            for (var i = 0; i < rawData.length; i++) {
                if (String(rawData[i][uuidIdx]) === uuid) {
                    lat = rawData[i][latIdx];
                    lon = rawData[i][lonIdx];
                    break;
                }
            }
            if (lat == null || lon == null) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }

            return [
                {center: {lat: lat, lon: lon}, zoom: 12},
                {uuid: uuid, lat: lat, lon: lon}
            ];
        }
        """,
        [Output('map-view-store', 'data', allow_duplicate=True),
         Output('bubble-clicked-point', 'data', allow_duplicate=True)],
        Input({'type': 'table-export', 'index': 0}, 'active_cell'),
        State({'type': 'table-export', 'index': 0}, 'data'),
        State('client-layer-data', 'data'),
        State('table-layer-dropdown', 'value'),
        prevent_initial_call=True
    )

    # --- TABLE ROW HIGHLIGHT (full row on click) ---
    app.clientside_callback(
        """
        function(active_cell) {
            var base = [
                {'if': {'state': 'selected'}, 'backgroundColor': 'rgba(88, 60, 165, 0.35)', 'color': '#fff', 'border': '1px solid #583CA5'},
                {'if': {'state': 'active'}, 'backgroundColor': 'rgba(255,200,0,0.25)', 'color': '#fff', 'border': '1px solid rgba(255,200,0,0.6)'}
            ];
            if (!active_cell || active_cell.row == null) return base;
            base.push({'if': {'row_index': active_cell.row}, 'backgroundColor': 'rgba(255,200,0,0.18)', 'color': '#fff'});
            return base;
        }
        """,
        Output({'type': 'table-export', 'index': 0}, 'style_data_conditional'),
        Input({'type': 'table-export', 'index': 0}, 'active_cell'),
        prevent_initial_call=True
    )

    # --- CLICK GREEN PIN ON MAP → CLEAR IT ---
    app.clientside_callback(
        """
        function(clickData, bp) {
            if (!bp || bp.lat == null || !clickData || !clickData.points || !clickData.points.length) {
                return window.dash_clientside.no_update;
            }
            var pt = clickData.points[0];
            var dlat = Math.abs((pt.lat || 0) - bp.lat);
            var dlon = Math.abs((pt.lon || 0) - bp.lon);
            if (dlat < 0.001 && dlon < 0.001) {
                return null;
            }
            return window.dash_clientside.no_update;
        }
        """,
        Output('bubble-clicked-point', 'data', allow_duplicate=True),
        Input('map', 'clickData'),
        State('bubble-clicked-point', 'data'),
        prevent_initial_call=True
    )

    @app.callback(
        Output("download-session", "data"),
        Input("save-session-btn", "n_clicks"),
        State("map-view-store", "data"),
        State("column-dropdown", "value"),
        State("user-session-id", "data"),
        State({'type': 'histogram-slider', 'column': ALL}, 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'id'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        State('current-selection-store', 'data'),
        prevent_initial_call=True
    )
    def save_session(n_clicks, map_view, filter_columns, session_id,
                     slider_vals, slider_ids, cat_vals, cat_ids, selected_points):
        if not n_clicks or not session_id: return dash.no_update
        session_data = dm.get_session_state(session_id)
        session_data['map_view'] = map_view
        session_data['filter_columns'] = filter_columns
        session_data['selected_points'] = selected_points or []
        # Save filter values (slider ranges and categorical selections)
        filter_values = {}
        if slider_ids:
            for id_dict, val in zip(slider_ids, slider_vals):
                filter_values[id_dict['column']] = {'type': 'numeric', 'value': val}
        if cat_ids:
            for id_dict, val in zip(cat_ids, cat_vals):
                filter_values[id_dict['column']] = {'type': 'categorical', 'value': val}
        session_data['filter_values'] = filter_values
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"agcap_session_{timestamp}.agcap"
        return dict(content=json.dumps(session_data), filename=filename)

    @app.callback(
        [Output('layer-manager-trigger', 'data', allow_duplicate=True),
         Output('layer-data-content-trigger', 'data', allow_duplicate=True),
         Output('map-view-store', 'data', allow_duplicate=True),
         Output('column-dropdown', 'value'),
         Output('current-selection-store', 'data', allow_duplicate=True),
         Output('pending-filter-values', 'data')],
        Input('load-session-upload', 'contents'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def load_session(contents, session_id):
        no_up = (dash.no_update,) * 6
        if not contents or not session_id: return no_up
        try:
            content_type, content_string = contents.split(',')
            decoded = base64.b64decode(content_string)
            session_data = json.loads(decoded.decode('utf-8'))

            success, msg = dm.load_session_state(session_id, session_data)
            if not success:
                print(msg)
                return no_up

            map_view = session_data.get('map_view', {'center': None, 'zoom': None})
            filter_columns = session_data.get('filter_columns', [])
            selected_points = session_data.get('selected_points', [])
            filter_values = session_data.get('filter_values', {})
            trigger = np.random.randint(100000)
            return trigger, trigger, map_view, filter_columns, selected_points, filter_values

        except Exception as e:
            print(f"Error parsing session file: {e}")
            return no_up

    @app.callback(
        Output('table-edits-store', 'data'),
        Input({'type': 'table-export', 'index': ALL}, 'data'),
        State('table-edits-store', 'data'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def save_table_changes(rows_list, existing_edits, session_id):
        rows = rows_list[0] if rows_list else None
        if not rows or not session_id: return dash.no_update

        edits = dict(existing_edits) if existing_edits else {}
        changed = False
        user_layers = dm.get_user_data(session_id)['layers']

        for row in rows:
            uuid_idx = row.get('orig_uuid')
            if not uuid_idx: continue

            for lid, df in user_layers.items():
                if uuid_idx in df.index:
                    old_notes = str(df.at[uuid_idx, 'Notes']) if 'Notes' in df.columns else ''
                    new_notes = str(row.get('Notes', ''))
                    old_score = int(df.at[uuid_idx, 'Score']) if 'Score' in df.columns else 0
                    try: new_score = int(row.get('Score', 0)) if row.get('Score') not in [None, ''] else 0
                    except: new_score = 0

                    if new_notes != old_notes or new_score != old_score:
                        df.at[uuid_idx, 'Notes'] = new_notes
                        df.at[uuid_idx, 'Score'] = new_score
                        edits[uuid_idx] = {'Notes': new_notes, 'Score': new_score}
                        changed = True
                    break
        return edits if changed else dash.no_update

    @app.callback(
        Output('layer-manager-trigger', 'data'),
        Output('layer-data-content-trigger', 'data'), # <--- KEEPS MAP SYNC WORKING
        Output('upload-status', 'children'),
        Input('upload-data', 'contents'),
        State('upload-data', 'filename'),
        State('layer-manager-trigger', 'data'),
        State('layer-data-content-trigger', 'data'),  # <--- NEEDED FOR SYNC
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def parse_upload(contents, filename, ui_trigger, data_trigger, session_id):
        lang = app.lang  # Get language from app config
        # 1. Basic Checks
        if contents is None or not session_id:
            return dash.no_update, dash.no_update, ""

        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        # File size gate: reject files > 30 MB
        if len(decoded) > 30 * 1024 * 1024:
            return dash.no_update, dash.no_update, "File exceeds 30 MB limit. Please upload a smaller file."

        try:
            new_df = None
            fname_lower = filename.lower()

            # CSV files
            if fname_lower.endswith('.csv'):
                try:
                    df = pd.read_csv(io.StringIO(decoded.decode('utf-8')), low_memory=False)
                except Exception as e:
                    return dash.no_update, dash.no_update, f"CSV read error: {str(e)}"

                # Check for various geometry column names
                wkt_col = next((c for c in df.columns if c.lower() in ['wkt', 'wkt_geom', 'geometry', 'geom']), None)
                lat_col = next((c for c in df.columns if c.lower() in ['lat', 'latitude', 'y']), None)
                lon_col = next((c for c in df.columns if c.lower() in ['lon', 'longitude', 'x']), None)

                if wkt_col:
                    try:
                        df[wkt_col] = df[wkt_col].fillna('').astype(str)
                        geometry = df[wkt_col].apply(wkt.loads)
                        gdf = gpd.GeoDataFrame(df, geometry=geometry, crs="EPSG:4326")
                        new_df = pd.DataFrame(gdf.drop(columns='geometry'))
                        new_df['lat'] = gdf.geometry.y
                        new_df['lon'] = gdf.geometry.x
                    except Exception as e:
                        return dash.no_update, dash.no_update, f"WKT processing error: {str(e)}"
                elif lat_col and lon_col:
                    new_df = df.copy()
                    new_df['lat'] = pd.to_numeric(new_df[lat_col], errors='coerce')
                    new_df['lon'] = pd.to_numeric(new_df[lon_col], errors='coerce')
                else:
                    return dash.no_update, dash.no_update, "No lat/lon or WKT geometry found in CSV"

            # Geospatial files (GeoJSON, GeoPackage, FlatGeobuf, Shapefiles, etc.)
            elif fname_lower.endswith(('.geojson', '.json', '.gpkg', '.fgb', '.shp', '.zip')):
                suffix = "." + filename.split('.')[-1]
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                    tmp.write(decoded)
                    tmp_path = tmp.name

                try:
                    # Read with geopandas (supports .gpkg, .fgb, .geojson, .shp, .zip)
                    if fname_lower.endswith('.zip'):
                        gdf = gpd.read_file(f"zip://{tmp_path}")
                    else:
                        gdf = gpd.read_file(tmp_path)

                    # Reproject to WGS84 if needed
                    if gdf.crs and gdf.crs.to_string() != "EPSG:4326":
                        gdf = gdf.to_crs(epsg=4326)

                    if 'geometry' not in gdf.columns or gdf.geometry.isna().all():
                        return dash.no_update, dash.no_update, "No geometry column found"

                    # Detect geometry type — must be homogeneous
                    geom_types = set(gdf.geometry.dropna().geom_type.unique())
                    polygon_types = geom_types & {'Polygon', 'MultiPolygon'}
                    point_types = geom_types & {'Point', 'MultiPoint'}
                    is_polygon = bool(polygon_types)
                    is_point = bool(point_types)

                    if is_polygon and is_point:
                        return dash.no_update, dash.no_update, "Mixed geometry types not supported. Please upload a single geometry type."

                    if is_polygon:
                        # Polygon path — delegate entirely to add_polygon_layer
                        dm.add_polygon_layer(session_id, gdf, filename.split('.')[0], filename)
                        return (ui_trigger or 0) + 1, (data_trigger or 0) + 1, f"✓ Loaded polygon layer: {filename}"

                    # Point path — existing logic
                    new_df = pd.DataFrame(gdf.drop(columns='geometry'))
                    new_df['lat'] = gdf.geometry.centroid.y.values
                    new_df['lon'] = gdf.geometry.centroid.x.values
                except Exception as e:
                    return dash.no_update, dash.no_update, f"Geospatial file error: {str(e)}"
                finally:
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
            else:
                return dash.no_update, dash.no_update, f"Unsupported file type: {filename}"

            # Robust type conversion (point layers only reach here)
            if new_df is not None:
                for col in new_df.columns:
                    if pd.api.types.is_numeric_dtype(new_df[col]):
                        new_df[col] = pd.to_numeric(new_df[col], errors='coerce').astype('float64')
                    else:
                        # Safe string conversion: use .apply(str) to handle mixed types
                        new_df[col] = new_df[col].fillna('').apply(str)

                # Add layer to session
                dm.add_layer(session_id, new_df, filename.split('.')[0], filename)

                return (ui_trigger or 0) + 1, (data_trigger or 0) + 1, f"✓ Loaded: {filename}"

        except Exception as e:
            import traceback
            return dash.no_update, dash.no_update, f"Error: {str(e)}\n{traceback.format_exc()}"

        return dash.no_update, dash.no_update, ""

    # --- CLIENTSIDE: GEOCODER (Async Fetch) ---
    app.clientside_callback(
        """
        async function(n_clicks, n_submit, query) {
            // Check if we have a query
            if (!query) {
                return window.dash_clientside.no_update;
            }

            try {
                // Perform the fetch directly from the browser
                const response = await fetch(`https://nominatim.openstreetmap.org/search?q=${query}&format=json&limit=1`, {
                    headers: {'User-Agent': 'DashAgCAP'}
                });
                const data = await response.json();

                if (data && data.length > 0) {
                    const res = data[0];
                    // Return the new map view store data
                    return {
                        'center': {'lat': parseFloat(res.lat), 'lon': parseFloat(res.lon)}, 
                        'zoom': 11
                    };
                }
            } catch (error) {
                console.error("Geocoding error:", error);
            }
            
            return window.dash_clientside.no_update;
        }
        """,
        Output('map-view-store', 'data', allow_duplicate=True),
        Input('geocoder-btn', 'n_clicks'),
        Input('geocoder-input', 'n_submit'),
        State('geocoder-input', 'value'),
        prevent_initial_call=True
    )

    @app.callback(
        Output('layer-manager-trigger', 'data', allow_duplicate=True),
        Input({'type': 'layer-color-col', 'index': ALL}, 'value'),
        Input({'type': 'layer-colorscale', 'index': ALL}, 'value'),
        Input({'type': 'layer-single-color', 'index': ALL}, 'value'),
        Input({'type': 'layer-fill-color', 'index': ALL}, 'value'),
        Input({'type': 'layer-fill-opacity', 'index': ALL}, 'value'),
        Input({'type': 'layer-outline-color', 'index': ALL}, 'value'),
        Input({'type': 'layer-outline-weight', 'index': ALL}, 'value'),
        Input({'type': 'layer-outline-opacity', 'index': ALL}, 'value'),
        State('layer-manager-trigger', 'data'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def update_layer_ui_on_color_change(col_vals, scale_vals, color_vals,
                                         fill_color_vals, fill_op_vals,
                                         outline_color_vals, outline_weight_vals, outline_op_vals,
                                         current_trigger, session_id):
        ctx = callback_context
        if not ctx.triggered or not session_id: return dash.no_update

        try:
            prop_id = ctx.triggered[0]['prop_id']
            trigger_val = ctx.triggered[0]['value']
            trigger_dict = json.loads(prop_id.split('.')[0])

            l_id = trigger_dict['index']
            t_type = trigger_dict['type']

            if t_type == 'layer-color-col':
                dm.update_setting(session_id, l_id, 'color_column', trigger_val)
                new_mode = 'single' if trigger_val == 'Single Color' else 'column'
                dm.update_setting(session_id, l_id, 'color_mode', new_mode)
            elif t_type == 'layer-colorscale':
                dm.update_setting(session_id, l_id, 'color_scale', trigger_val)
            elif t_type == 'layer-single-color':
                val = trigger_val['hex'] if isinstance(trigger_val, dict) and 'hex' in trigger_val else trigger_val
                dm.update_setting(session_id, l_id, 'single_color_hex', val)
            elif t_type == 'layer-fill-color':
                dm.update_setting(session_id, l_id, 'fill_color', trigger_val)
            elif t_type == 'layer-fill-opacity':
                dm.update_setting(session_id, l_id, 'fill_opacity', trigger_val)
            elif t_type == 'layer-outline-color':
                dm.update_setting(session_id, l_id, 'outline_color', trigger_val)
            elif t_type == 'layer-outline-weight':
                dm.update_setting(session_id, l_id, 'outline_weight', trigger_val)
            elif t_type == 'layer-outline-opacity':
                dm.update_setting(session_id, l_id, 'outline_opacity', trigger_val)

            return (current_trigger or 0) + 1
        except: return dash.no_update
    
    # --- NEW: SYNC VISIBILITY LIST ---
    @app.callback(
        Output('visible-layers-list', 'data'),
        Input('layer-manager-trigger', 'data'),
        State('user-session-id', 'data')
    )
    def update_visible_list(trigger, session_id):
        if not session_id: return []
        user_meta = dm.get_metadata(session_id)
        # Return list of IDs where visible is True
        return [l['id'] for l in user_meta if l['visible']]

    @app.callback(
        Output('layers-list-container', 'children'),
        Input('layer-manager-trigger', 'data'),
        State('user-session-id', 'data')
    )
    def render_layer_list(trigger, session_id):
        if not session_id: return []

        user_meta = dm.get_metadata(session_id)

        # Shared styles
        lbl = {'fontSize':'10px', 'color':'#bdc3c7'}
        lbl_sm = {'fontSize':'9px', 'color':'#bdc3c7'}
        color_input_style = {'width':'100%','height':'26px','padding':'0','border':'none','backgroundColor':'transparent','cursor':'pointer'}

        children = []
        for i, layer in enumerate(user_meta):
            lid = layer['id']
            is_polygon = layer.get('layer_type') == 'polygon'

            # --- Header row (shared by both types) ---
            header = html.Div([
                dcc.Input(id={'type': 'layer-name', 'index': lid}, value=layer['name'],
                          style={'backgroundColor':'transparent','border':'none','color':'white','fontWeight':'bold','flexGrow':1,'width':'80px','fontSize':'12px','fontFamily':'Mulish'}),
                html.Button("▲", id={'type': 'layer-up', 'index': lid},
                            style={'padding':'0 4px','background':'transparent','border':'none','color':'#bdc3c7','cursor':'pointer','fontSize':'10px'}),
                html.Button("▼", id={'type': 'layer-down', 'index': lid},
                            style={'padding':'0 4px','background':'transparent','border':'none','color':'#bdc3c7','cursor':'pointer','fontSize':'10px'}),
                html.Button(html.I(className="fas fa-eye" if layer['visible'] else "fas fa-eye-slash",
                                   style={'color':'#2ecc71' if layer['visible'] else '#95a5a6'}),
                            id={'type': 'layer-vis-btn', 'index': lid},
                            style={'background':'transparent','border':'none','cursor':'pointer','padding':'0 4px','fontSize':'12px'}),
                html.Button(html.I(className="fas fa-trash-alt", style={'color':'#e74c3c'}),
                            id={'type': 'layer-del-btn', 'index': lid},
                            style={'background':'transparent','border':'none','cursor':'pointer','padding':'0 4px','fontSize':'12px'}),
            ], style={'display':'flex','alignItems':'center','marginBottom':'2px'})

            if is_polygon:
                # --- POLYGON CARD ---
                all_attr_cols = [c for c in layer.get('columns', [])
                                 if c not in ('centroid_lat','centroid_lon','ID','row_uuid','orig_uuid')]
                color_col_opts = [{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c}
                                  for c in all_attr_cols] + [{'label': t('layer_settings.single_color', lang), 'value': 'Single Color'}]

                poly_dropdown = dcc.Dropdown(
                    id={'type': 'layer-color-col', 'index': lid},
                    options=color_col_opts,
                    value=layer.get('color_column', 'Single Color'),
                    clearable=False,
                    style={'color':'#333','fontSize':'11px','width':'100%','marginTop':'0px'}
                )

                _scale_display = 'none' if layer.get('color_mode', 'single') == 'single' else None
                poly_secondary = html.Div([
                    dbc.Input(id={'type': 'layer-single-color', 'index': lid}, type='color',
                              value=layer.get('single_color_hex', '#3498db'),
                              style={**color_input_style, 'display': 'none'}),
                    dcc.Dropdown(id={'type': 'layer-colorscale', 'index': lid},
                                 options=[{'label': c, 'value': c} for c in SCALE_MAPPINGS.keys()],
                                 value=layer.get('color_scale', 'Portland'),
                                 clearable=False,
                                 style={'color':'#333','fontSize':'11px','width':'100%',
                                        **(({'display':'none'}) if _scale_display else {})})
                ])

                settings_body = html.Div([
                    html.Div([
                        html.Label('Fill color', style=lbl),
                        dbc.Input(id={'type': 'layer-fill-color', 'index': lid}, type='color',
                                  value=layer.get('fill_color', '#3498db'), style=color_input_style),
                    ], style={'marginBottom':'2px'}),
                    html.Div([html.Label('Fill opacity', style=lbl),
                              dcc.Slider(id={'type': 'layer-fill-opacity', 'index': lid},
                                         min=0, max=1, step=0.05, value=layer.get('fill_opacity', 0.4), marks=None)],
                             style={'marginBottom':'0px','height':'24px'}),
                    html.Div([
                        html.Label('Outline color', style=lbl),
                        dbc.Input(id={'type': 'layer-outline-color', 'index': lid}, type='color',
                                  value=layer.get('outline_color', '#333333'), style=color_input_style),
                    ], style={'marginBottom':'2px'}),
                    html.Div([html.Label('Outline weight', style=lbl),
                              dcc.Slider(id={'type': 'layer-outline-weight', 'index': lid},
                                         min=0, max=8, step=0.5, value=layer.get('outline_weight', 2), marks=None)],
                             style={'marginBottom':'0px','height':'24px'}),
                    html.Div([html.Label('Outline opacity', style=lbl),
                              dcc.Slider(id={'type': 'layer-outline-opacity', 'index': lid},
                                         min=0, max=1, step=0.05, value=layer.get('outline_opacity', 1.0), marks=None)],
                             style={'marginBottom':'2px','height':'24px'}),
                    # Point layer dummies (hidden) so ALL pattern-match callbacks don't miss this layer
                    html.Div(dcc.Slider(id={'type': 'layer-size', 'index': lid}, min=1, max=20, step=1,
                               value=layer.get('size', 8), marks=None), style={'display':'none'}),
                    html.Div(dcc.Slider(id={'type': 'layer-opacity', 'index': lid}, min=0, max=1, step=0.1,
                               value=0.7, marks=None), style={'display':'none'}),
                    # Color-by-attribute
                    html.Div([html.Label('Color by attribute', style=lbl), poly_dropdown, poly_secondary]),
                ], style={'backgroundColor':'#34495e','padding':'2px 5px','borderRadius':'4px'})

            else:
                # --- POINT CARD (existing logic unchanged) ---
                dropdown = dcc.Dropdown(
                    id={'type': 'layer-color-col', 'index': lid},
                    options=[{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c}
                             for c in layer.get('numeric_columns', [])] + [{'label': t('layer_settings.single_color', lang), 'value': 'Single Color'}],
                    value=layer.get('color_column'),
                    clearable=False,
                    style={'color':'#333','fontSize':'11px','width':'100%','marginTop':'0px'}
                )
                if layer['color_mode'] == 'single':
                    secondary = html.Div([
                        html.Label(t('layer_settings.color', lang), style={**lbl, 'marginRight':'8px'}),
                        dbc.Input(id={'type': 'layer-single-color', 'index': lid}, type='color',
                                  value=layer.get('single_color_hex', '#FBB800'), style=color_input_style)
                    ], style={'marginTop':'5px','display':'flex','alignItems':'center'})
                else:
                    secondary = html.Div([
                        html.Label(t('layer_settings.scale', lang), style=lbl_sm),
                        dcc.Dropdown(id={'type': 'layer-colorscale', 'index': lid},
                                     options=[{'label': c, 'value': c} for c in SCALE_MAPPINGS.keys()],
                                     value=layer.get('color_scale', 'Portland'),
                                     clearable=False, style={'color':'#333','fontSize':'11px','width':'100%'})
                    ])
                # Polygon-specific dummies (hidden) so ALL pattern-match callbacks work
                poly_dummies = html.Div([
                    dbc.Input(id={'type': 'layer-fill-color', 'index': lid}, type='color',
                              value='#3498db', style={'display':'none'}),
                    html.Div(dcc.Slider(id={'type': 'layer-fill-opacity', 'index': lid}, min=0, max=1, step=0.05,
                               value=0.4, marks=None), style={'display':'none'}),
                    dbc.Input(id={'type': 'layer-outline-color', 'index': lid}, type='color',
                              value='#333333', style={'display':'none'}),
                    html.Div(dcc.Slider(id={'type': 'layer-outline-weight', 'index': lid}, min=0, max=8, step=0.5,
                               value=2, marks=None), style={'display':'none'}),
                    html.Div(dcc.Slider(id={'type': 'layer-outline-opacity', 'index': lid}, min=0, max=1, step=0.05,
                               value=1.0, marks=None), style={'display':'none'}),
                ])
                settings_body = html.Div([
                    html.Div([html.Label(t('layer_settings.size', lang), style=lbl),
                              dcc.Slider(id={'type': 'layer-size', 'index': lid}, min=1, max=20, step=1,
                                         value=layer.get('size', 4), marks=None)],
                             style={'marginBottom':'0px','height':'24px'}),
                    html.Div([html.Label(t('layer_settings.opacity', lang), style=lbl),
                              dcc.Slider(id={'type': 'layer-opacity', 'index': lid}, min=0, max=1, step=0.1,
                                         value=layer.get('opacity', 0.8), marks=None)],
                             style={'marginBottom':'2px','height':'24px'}),
                    html.Div([html.Label(t('layer_settings.color', lang), style=lbl), dropdown, secondary]),
                    poly_dummies,
                ], style={'backgroundColor':'#34495e','padding':'2px 5px','borderRadius':'4px'})

            card = html.Div([header, settings_body],
                            style={'padding':'5px','backgroundColor':'#555555','borderRadius':'4px',
                                   'marginBottom':'5px',
                                   'borderLeft': f"4px solid {'#2ecc71' if layer['visible'] else '#95a5a6'}"})
            children.append(card)
        return children
    
    # @app.callback(
    #     Output({'type': 'color-picker-container', 'index': MATCH}, 'style'),
    #     Input({'type': 'color-toggle-btn', 'index': MATCH}, 'n_clicks'),
    #     prevent_initial_call=True
    # )
    # def toggle_color_picker(n_clicks):
    #     # If even clicks (0, 2, 4...), hide. If odd (1, 3...), show.
    #     if n_clicks is None or n_clicks % 2 == 0:
    #         return {'display': 'none'}
    #     else:
    #         return {'display': 'flex', 'justifyContent': 'center', 'marginTop': '5px'}

    @app.callback(
        Output('layer-manager-trigger', 'data', allow_duplicate=True),
        Input({'type': 'layer-up', 'index': ALL}, 'n_clicks'),
        Input({'type': 'layer-down', 'index': ALL}, 'n_clicks'),
        Input({'type': 'layer-vis-btn', 'index': ALL}, 'n_clicks'),
        Input({'type': 'layer-del-btn', 'index': ALL}, 'n_clicks'),
        State({'type': 'layer-size', 'index': ALL}, 'value'),
        State({'type': 'layer-size', 'index': ALL}, 'id'),
        State({'type': 'layer-opacity', 'index': ALL}, 'value'),
        State({'type': 'layer-opacity', 'index': ALL}, 'id'),
        State({'type': 'layer-color-col', 'index': ALL}, 'value'),
        State({'type': 'layer-color-col', 'index': ALL}, 'id'),
        State({'type': 'layer-colorscale', 'index': ALL}, 'value'),
        State({'type': 'layer-colorscale', 'index': ALL}, 'id'),
        State({'type': 'layer-single-color', 'index': ALL}, 'value'),
        State({'type': 'layer-single-color', 'index': ALL}, 'id'),
        State({'type': 'layer-fill-color', 'index': ALL}, 'value'),
        State({'type': 'layer-fill-color', 'index': ALL}, 'id'),
        State({'type': 'layer-fill-opacity', 'index': ALL}, 'value'),
        State({'type': 'layer-fill-opacity', 'index': ALL}, 'id'),
        State({'type': 'layer-outline-color', 'index': ALL}, 'value'),
        State({'type': 'layer-outline-color', 'index': ALL}, 'id'),
        State({'type': 'layer-outline-weight', 'index': ALL}, 'value'),
        State({'type': 'layer-outline-weight', 'index': ALL}, 'id'),
        State({'type': 'layer-outline-opacity', 'index': ALL}, 'value'),
        State({'type': 'layer-outline-opacity', 'index': ALL}, 'id'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def manage_layers(up, down, vis, delete,
                      sizes, size_ids,
                      ops, op_ids,
                      cols, col_ids,
                      scales, scale_ids,
                      colors, color_ids,
                      fill_colors, fill_color_ids,
                      fill_ops, fill_op_ids,
                      outline_colors, outline_color_ids,
                      outline_weights, outline_weight_ids,
                      outline_ops, outline_op_ids,
                      session_id):

        ctx = callback_context
        if not ctx.triggered or not session_id: return dash.no_update

        # --- 1. SNAPSHOT: Save current UI state to Backend before moving ---
        try:
            def sync_settings(values, ids, key):
                if not values or not ids: return
                for val, id_dict in zip(values, ids):
                    final_val = val
                    if key == 'single_color_hex' and isinstance(val, dict) and 'hex' in val:
                        final_val = val['hex']
                    dm.update_setting(session_id, id_dict['index'], key, final_val)

            sync_settings(sizes, size_ids, 'size')
            sync_settings(ops, op_ids, 'opacity')
            sync_settings(cols, col_ids, 'color_column')
            sync_settings(scales, scale_ids, 'color_scale')
            sync_settings(colors, color_ids, 'single_color_hex')
            sync_settings(fill_colors, fill_color_ids, 'fill_color')
            sync_settings(fill_ops, fill_op_ids, 'fill_opacity')
            sync_settings(outline_colors, outline_color_ids, 'outline_color')
            sync_settings(outline_weights, outline_weight_ids, 'outline_weight')
            sync_settings(outline_ops, outline_op_ids, 'outline_opacity')
        except Exception as e:
            print(f"Error syncing settings: {e}")

        # --- 2. PERFORM ACTION ---
        try:
            payload = json.loads(ctx.triggered[0]['prop_id'].split('.')[0])
            layer_id = payload['index']
            action = payload['type']

            if action == 'layer-up': dm.move_layer(session_id, layer_id, 'up')
            elif action == 'layer-down': dm.move_layer(session_id, layer_id, 'down')
            elif action == 'layer-vis-btn': dm.toggle_visibility(session_id, layer_id)
            elif action == 'layer-del-btn': dm.delete_layer(session_id, layer_id)

            return np.random.randint(100000)
        except: return dash.no_update

    @app.callback(
        Output('table-layer-dropdown', 'options'),
        Output('table-layer-dropdown', 'value'),
        Input('layer-manager-trigger', 'data'),
        State('table-layer-dropdown', 'value'),
        State('user-session-id', 'data') # <--- ADDED STATE
    )
    def update_table_dropdown_options(trigger, current_val, session_id):
        if not session_id: return [], None
        user_meta = dm.get_metadata(session_id)
        if not user_meta: return [], None
        
        options = [{'label': l['name'], 'value': l['id']} for l in user_meta]
        existing_ids = [l['id'] for l in user_meta]
        if current_val in existing_ids: return options, current_val
        primary = next((l for l in user_meta if l.get('is_primary')), None)
        default_val = primary['id'] if primary else user_meta[0]['id']
        return options, default_val

    @app.callback(
        [Output('legend-container', 'children'), Output('legend-container', 'style'),
         Output('filter-histogram-data', 'data')],
        Input('layer-manager-trigger', 'data'),
        Input({'type': 'layer-size', 'index': ALL}, 'value'),
        Input({'type': 'layer-opacity', 'index': ALL}, 'value'),
        Input({'type': 'layer-color-col', 'index': ALL}, 'value'),
        Input({'type': 'layer-colorscale', 'index': ALL}, 'value'),
        Input({'type': 'layer-single-color', 'index': ALL}, 'value'),
        Input({'type': 'histogram-slider', 'column': ALL}, 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'id'),
        Input({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        Input('table-layer-dropdown', 'value'),
        State('user-session-id', 'data'),
        State('column-dropdown', 'value'),
        State('filter-log-scale-store', 'data')
    )
    def update_legend_only(trigger, sizes, opacities, color_cols, color_scales, single_colors, slider_vals, slider_ids, cat_vals, cat_ids, active_table_layer, session_id, selected_columns, clip_cols):
        # NOTE: This callback now ONLY handles the Legend. The Map is updated Client-Side.

        if not session_id: return [], {'display': 'none'}, {}

        # --- REPLICATE FILTERING LOGIC FOR LEGEND STATS ---
        # We need to filter data here just to show the correct Min/Max in the legend

        legend_items = [html.H5("LEGEND", style={'fontSize': '12px', 'marginBottom': '8px', 'borderBottom': '1px solid #777', 'paddingBottom': '4px'})]
        legend_style = {'display': 'block', 'position': 'absolute', 'top': '60px', 'left': '15px', 'zIndex': 5, 'width': '220px', 'maxHeight': '80vh', 'overflowY': 'auto'}

        user_data = dm.get_user_data(session_id)
        visible_layers = [l for l in reversed(user_data['metadata']) if l['visible']]

        if not visible_layers: return [], {'display': 'none'}, {}

        # Build slider/cat lookup for cross-filter histogram computation
        slider_map = {sid['column']: val for sid, val in zip(slider_ids or [], slider_vals or [])}
        cat_map = {cid['column']: val for cid, val in zip(cat_ids or [], cat_vals or [])}
        clip_cols = clip_cols or {}

        # Compute cross-filtered histogram data for each selected filter column
        hist_data = {}
        if selected_columns and not df_main.empty:
            for col in selected_columns:
                if col in ['lat', 'lon', 'row_uuid', 'geometry', 'index', 'orig_uuid', 'ID']:
                    continue
                # Build cross-filter mask (exclude current col)
                mask = pd.Series(True, index=df_main.index)
                for other_col, val in slider_map.items():
                    if other_col == col or other_col not in df_main.columns or not val:
                        continue
                    mask &= (df_main[other_col] >= val[0]) & (df_main[other_col] <= val[1])
                for other_col, vals in cat_map.items():
                    if other_col == col or other_col not in df_main.columns or not vals:
                        continue
                    mask &= df_main[other_col].isin(vals)
                cross_filtered_col = df_main.loc[mask, col].dropna()

                if pd.api.types.is_numeric_dtype(df_main[col]):
                    if cross_filtered_col.empty:
                        hist_data[col] = {'counts': []}
                        continue
                    clean_full = df_main[col].dropna()
                    mn_full = float(clean_full.min())
                    mx_full = float(clean_full.max())
                    use_clip = clip_cols.get(col, False)
                    p95 = float(np.percentile(clean_full, 95))
                    if p95 <= mn_full: p95 = mx_full
                    sl_max = p95 if use_clip else mx_full
                    counts, _ = np.histogram(cross_filtered_col, bins=50, range=(mn_full, sl_max))
                    hist_data[col] = {'counts': counts.tolist()}
                else:
                    counts_series = cross_filtered_col.value_counts()
                    hist_data[col] = {'cat_counts': {str(k): int(v) for k, v in counts_series.items()}}

        for layer in visible_layers:
            dff = user_data['layers'][layer['id']]

            # Filter Logic (Duplicate of Client Side, but needed for Stats)
            if layer.get('is_primary', False):
                filtered_df = dff.copy()
                if slider_ids and slider_vals:
                    for value, sid in zip(slider_vals, slider_ids):
                        col = sid['column']
                        if col in filtered_df.columns and value:
                            filtered_df = filtered_df[(filtered_df[col] >= value[0]) & (filtered_df[col] <= value[1])]
                if cat_ids and cat_vals:
                    for values, cid in zip(cat_vals, cat_ids):
                        col = cid['column']
                        if col in filtered_df.columns and values:
                            filtered_df = filtered_df[filtered_df[col].isin(values)]
                dff = filtered_df

            # Construct Legend Item
            layer_display_name = t('layer.default_name', lang) if layer['name'] == 'Settlements' else layer['name']

            if layer.get('layer_type') == 'polygon':
                # --- POLYGON LEGEND ---
                col = layer.get('color_column', 'Single Color')
                if col == 'Single Color' or layer.get('color_mode', 'single') == 'single':
                    fill_c = layer.get('fill_color', '#3498db')
                    outline_c = layer.get('outline_color', '#333333')
                    legend_items.append(html.Div([
                        html.Div(style={
                            'width': '18px', 'height': '14px', 'flexShrink': '0',
                            'backgroundColor': fill_c, 'border': f"2px solid {outline_c}",
                            'borderRadius': '2px', 'marginRight': '6px'
                        }),
                        html.Div(layer_display_name, className="legend-label")
                    ], className="legend-item"))
                else:
                    col_display_name = get_column_translation(col, lang, COUNTRY_CODE)
                    scale_name = layer.get('color_scale', 'Portland')
                    if col in dff.columns and not dff.empty:
                        grad_css = SCALE_MAPPINGS.get(scale_name, 'linear-gradient(to right, #ccc, #333)')
                        is_numeric_col = pd.api.types.is_numeric_dtype(dff[col])
                        if is_numeric_col:
                            mn, mx = dff[col].min(), dff[col].max()
                            range_label = html.Div([
                                html.Span(f"{mn:.1f}", style={'fontSize':'9px','marginRight':'3px'}),
                                html.Div(className="legend-gradient-bar", style={'background': grad_css}),
                                html.Span(f"{mx:.1f}", style={'fontSize':'9px','marginLeft':'3px'})
                            ], style={'display':'flex','alignItems':'center'})
                        else:
                            n_cats = dff[col].nunique()
                            range_label = html.Div([
                                html.Span(f"{n_cats} categories", style={'fontSize':'9px','marginRight':'3px'}),
                                html.Div(className="legend-gradient-bar", style={'background': grad_css}),
                            ], style={'display':'flex','alignItems':'center'})
                        legend_items.append(html.Div([
                            html.Div(f"{layer_display_name} ({col_display_name})", style={'fontWeight':'bold','fontSize':'10px','marginBottom':'2px'}),
                            range_label
                        ], style={'marginBottom':'8px'}))
                    else:
                        legend_items.append(html.Div(f"{layer_display_name} ({t('legend.no_data', lang)})", className="legend-item"))
            else:
                # --- POINT LEGEND ---
                base_size = layer['size']
                if layer['color_mode'] == 'single':
                    color_val = layer.get('single_color_hex', '#FBB800')
                    legend_items.append(html.Div([
                        html.Div(className="legend-marker", style={'backgroundColor': color_val, 'width': f"{base_size}px", 'height': f"{base_size}px"}),
                        html.Div(layer_display_name, className="legend-label")
                    ], className="legend-item"))
                else:
                    col = layer['color_column']
                    col_display_name = get_column_translation(col, lang, COUNTRY_CODE)
                    scale_name = layer.get('color_scale', 'Portland')
                    if col in dff.columns and not dff.empty:
                        grad_css = SCALE_MAPPINGS.get(scale_name, 'linear-gradient(to right, #ccc, #333)')
                        mn, mx = dff[col].min(), dff[col].max()
                        legend_items.append(html.Div([
                            html.Div(f"{layer_display_name} ({col_display_name})", style={'fontWeight':'bold', 'fontSize':'10px', 'marginBottom':'2px'}),
                            html.Div([
                                html.Span(f"{mn:.1f}", style={'fontSize':'9px', 'marginRight':'3px'}),
                                html.Div(className="legend-gradient-bar", style={'background': grad_css}),
                                html.Span(f"{mx:.1f}", style={'fontSize':'9px', 'marginLeft':'3px'})
                            ], style={'display':'flex', 'alignItems':'center'})
                        ], style={'marginBottom': '8px'}))
                    else:
                        legend_items.append(html.Div(f"{layer_display_name} ({t('legend.no_data', lang)})", className="legend-item"))

        return legend_items, legend_style, hist_data

    @app.callback(
        [Output('histograms-container', 'children'),
         Output('pending-filter-values', 'data', allow_duplicate=True)],
        Input('column-dropdown', 'value'),
        Input('filter-log-scale-store', 'data'),
        State({'type': 'histogram-slider', 'column': ALL}, 'value'),
        State({'type': 'histogram-slider', 'column': ALL}, 'id'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        State('pending-filter-values', 'data'),
        prevent_initial_call='initial_duplicate'
    )
    def update_filter_panel(selected_columns, clip_cols, slider_vals, slider_ids, cat_vals, cat_ids, pending_filters):
        if not selected_columns: return [], None
        clip_cols = clip_cols or {}
        lang = app.lang
        current_numeric, current_cat = {}, {}
        if slider_ids: current_numeric = {id_dict['column']: val for id_dict, val in zip(slider_ids, slider_vals)}
        if cat_ids: current_cat = {id_dict['column']: val for id_dict, val in zip(cat_ids, cat_vals)}
        if pending_filters:
            for col, fv in pending_filters.items():
                if fv.get('type') == 'numeric' and col not in current_numeric:
                    current_numeric[col] = fv['value']
                elif fv.get('type') == 'categorical' and col not in current_cat:
                    current_cat[col] = fv['value']
        def fmt_num(v):
            if abs(v) >= 1_000_000: return f'{v/1_000_000:.2g}M'
            if abs(v) >= 10_000:    return f'{v/1_000:.3g}k'
            if abs(v) >= 1_000:     return f'{v/1_000:.2g}k'
            if v == int(v):         return str(int(v))
            return f'{v:.3g}'

        def slider_marks(mn, sl_max, n_marks=5):
            step = (sl_max - mn) / (n_marks - 1)
            vals = [mn + i * step for i in range(n_marks)]
            return {v: {'label': fmt_num(v), 'style': {'color': '#aaa', 'fontSize': '9px'}} for v in vals}

        children = []
        for col in selected_columns:
            if col in ['lat', 'lon', 'row_uuid', 'geometry', 'index', 'orig_uuid', 'ID']: continue
            translated_col = get_column_translation(col, lang, COUNTRY_CODE)
            if pd.api.types.is_numeric_dtype(df_main[col]):
                clean = df_main[col].dropna()
                mn, mx = (float(clean.min()), float(clean.max())) if not clean.empty else (0.0, 1.0)
                use_clip = clip_cols.get(col, False)
                p95 = float(np.percentile(clean, 95))
                if p95 <= mn: p95 = mx
                sl_max = p95 if use_clip else mx
                prev_val = current_numeric.get(col, [mn, sl_max])
                val = [prev_val[0], min(prev_val[1], sl_max)]
                n_bins = 50
                counts, _ = np.histogram(clean, bins=n_bins, range=(mn, sl_max))
                max_count = counts.max() if counts.max() > 0 else 1
                bar_divs = [html.Div(style={
                    'flex': '1', 'height': f'{int(counts[i] / max_count * 100)}%',
                    'backgroundColor': '#FBB800', 'alignSelf': 'flex-end',
                    'marginRight': '1px' if i < n_bins - 1 else '0'
                }) for i in range(n_bins)]
                btn_title = t('filters.p95_active', lang) if use_clip else t('filters.p95_inactive', lang)
                btn_style = {
                    'fontSize': '9px', 'padding': '1px 5px', 'border': 'none',
                    'borderRadius': '3px', 'cursor': 'pointer', 'fontFamily': 'Mulish',
                    'backgroundColor': '#583CA5' if use_clip else '#555',
                    'color': '#fff'
                }
                input_style = {
                    'flex': '1', 'minWidth': '0', 'backgroundColor': '#2a2a2a',
                    'color': '#ecf0f1', 'border': '1px solid #555', 'fontSize': '10px',
                    'borderRadius': '3px', 'padding': '2px 4px', 'fontFamily': 'Mulish'
                }
                is_active = (val[0] > mn or val[1] < sl_max)
                reset_style = {
                    'fontSize': '13px', 'padding': '0px 4px', 'border': 'none',
                    'borderRadius': '3px', 'cursor': 'pointer', 'fontFamily': 'Mulish',
                    'backgroundColor': '#583CA5' if is_active else 'transparent',
                    'color': '#FBB800' if is_active else '#777',
                    'lineHeight': '1'
                }
                children.append(html.Div([
                    html.Div([
                        html.Label(translated_col, style={'fontSize':'11px','fontWeight':'bold','color':'#ecf0f1','fontFamily':'Mulish','flex':'1','overflow':'hidden','textOverflow':'ellipsis','whiteSpace':'nowrap'}),
                        html.Button('↺', id={'type':'filter-reset','column':col},
                                    title=t('filters.reset_active', lang) if is_active else t('filters.reset_inactive', lang),
                                    style=reset_style,
                                    **{'data-mn': mn, 'data-slmax': sl_max}),
                        html.Button('p95', id={'type':'log-toggle','column':col}, n_clicks=1 if use_clip else 0,
                                    title=btn_title, style=btn_style),
                    ], style={'display':'flex','alignItems':'center','gap':'4px','marginBottom':'4px'}),
                    html.Div([
                        dcc.Input(id={'type':'min-input','column':col}, type='text', value=fmt_num(val[0]), debounce=True, style=input_style),
                        html.Span('–', style={'color':'#aaa','padding':'0 4px','lineHeight':'22px'}),
                        dcc.Input(id={'type':'max-input','column':col}, type='text', value=fmt_num(val[1]), debounce=True, style=input_style),
                    ], style={'display':'flex','alignItems':'center','marginBottom':'4px'}),
                    html.Div(id={'type':'mini-hist','column':col}, children=bar_divs, style={
                        'display': 'flex', 'alignItems': 'flex-end',
                        'height': '50px', 'padding': '0 25px', 'marginBottom': '2px',
                        'boxSizing': 'border-box'
                    }),
                    dcc.RangeSlider(id={'type':'histogram-slider','column':col}, min=mn, max=sl_max, value=val,
                                    marks=slider_marks(mn, sl_max), tooltip={"always_visible":False})
                ], style={'marginBottom':'10px','backgroundColor':'#2e2e2e','padding':'8px','borderRadius':'6px','border':'1px solid #444'}))
            else:
                val_counts = df_main[col].dropna().value_counts()
                cat_options = [{'label': f'{str(v)} [{val_counts.get(v, 0)}]', 'value': str(v)} for v in sorted(val_counts.index.tolist())]
                children.append(html.Div([
                    html.Label(translated_col, style={'fontSize':'11px','fontWeight':'bold','color':'#ecf0f1','fontFamily':'Mulish','marginBottom':'4px','display':'block'}),
                    dcc.Dropdown(id={'type':'categorical-dropdown','column':col}, options=cat_options, value=current_cat.get(col,[]), multi=True, style={'color':'#333','fontSize':'10px'})
                ], style={'marginBottom':'10px','backgroundColor':'#2e2e2e','padding':'8px','borderRadius':'6px','border':'1px solid #444'}))
        return children, None

    # --- P95 TOGGLE: n_clicks odd = clipped, even = full ---
    app.clientside_callback(
        """
        function(n_clicks_list, ids, current_store) {
            const ctx = dash_clientside.callback_context;
            if (!ctx.triggered.length) return window.dash_clientside.no_update;
            const triggered = ctx.triggered[0].prop_id;
            try {
                const id_part = triggered.split('.')[0];
                const id_obj = JSON.parse(id_part);
                const col = id_obj.column;
                const idx = ids.findIndex(function(id) { return id.column === col; });
                const n = n_clicks_list[idx] || 0;
                const store = Object.assign({}, current_store || {});
                store[col] = (n % 2 === 1);
                return store;
            } catch(e) {
                return window.dash_clientside.no_update;
            }
        }
        """,
        Output('filter-log-scale-store', 'data'),
        Input({'type': 'log-toggle', 'column': ALL}, 'n_clicks'),
        State({'type': 'log-toggle', 'column': ALL}, 'id'),
        State('filter-log-scale-store', 'data'),
        prevent_initial_call=True
    )

    # --- CLIENTSIDE: SYNC HISTOGRAM SLIDER & INPUTS ---
    app.clientside_callback(
        """
        function(slider_val, min_input_val, max_input_val, range_min, range_max) {
            const ctx = dash_clientside.callback_context;
            if (!ctx.triggered.length) return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];

            const trigger_id = ctx.triggered[0].prop_id;

            function fmtNum(v) {
                if (Math.abs(v) >= 1e6) return (v/1e6).toPrecision(3).replace(/\\.?0+$/, '') + 'M';
                if (Math.abs(v) >= 1e4) return (v/1e3).toPrecision(3).replace(/\\.?0+$/, '') + 'k';
                if (Math.abs(v) >= 1e3) return (v/1e3).toPrecision(2).replace(/\\.?0+$/, '') + 'k';
                if (v === Math.floor(v)) return String(Math.floor(v));
                return parseFloat(v.toPrecision(3)).toString();
            }

            function parseNum(s) {
                if (s === null || s === undefined || s === '') return null;
                if (typeof s === 'number') return s;
                s = String(s).trim().replace(/,/g, '');
                var m = s.match(/^(-?[\\d.]+)\\s*([kKmM]?)$/);
                if (!m) return null;
                var n = parseFloat(m[1]);
                if (m[2].toLowerCase() === 'k') n *= 1e3;
                if (m[2].toLowerCase() === 'm') n *= 1e6;
                return isNaN(n) ? null : n;
            }

            // 1. If slider changed, update text inputs
            if (trigger_id.includes('histogram-slider')) {
                return [window.dash_clientside.no_update, fmtNum(slider_val[0]), fmtNum(slider_val[1])];
            }

            // 2. If text inputs changed, update slider
            var mn = parseNum(min_input_val);
            var mx = parseNum(max_input_val);
            if (mn === null) mn = range_min;
            if (mx === null) mx = range_max;

            // Clamp values to the allowed range
            mn = Math.max(range_min, mn);
            mx = Math.min(range_max, mx);

            // Prevent min > max
            if (mn > mx) {
                if (trigger_id.includes('min-input')) {
                    mn = mx;
                } else {
                    mx = mn;
                }
            }

            return [[mn, mx], window.dash_clientside.no_update, window.dash_clientside.no_update];
        }
        """,
        Output({'type': 'histogram-slider', 'column': MATCH}, 'value'),
        Output({'type': 'min-input', 'column': MATCH}, 'value'),
        Output({'type': 'max-input', 'column': MATCH}, 'value'),
        Input({'type': 'histogram-slider', 'column': MATCH}, 'value'),
        Input({'type': 'min-input', 'column': MATCH}, 'value'),
        Input({'type': 'max-input', 'column': MATCH}, 'value'),
        State({'type': 'histogram-slider', 'column': MATCH}, 'min'),
        State({'type': 'histogram-slider', 'column': MATCH}, 'max'),
        prevent_initial_call=True
    )

    # --- CLIENTSIDE: RESET FILTER BUTTON → resets slider + inputs to full range ---
    app.clientside_callback(
        """
        function(n_clicks, id_obj, sl_min, sl_max) {
            if (!n_clicks) return [window.dash_clientside.no_update, window.dash_clientside.no_update, window.dash_clientside.no_update];

            function fmtNum(v) {
                if (Math.abs(v) >= 1e6) return (v/1e6).toPrecision(3).replace(/\\.?0+$/, '') + 'M';
                if (Math.abs(v) >= 1e4) return (v/1e3).toPrecision(3).replace(/\\.?0+$/, '') + 'k';
                if (Math.abs(v) >= 1e3) return (v/1e3).toPrecision(2).replace(/\\.?0+$/, '') + 'k';
                if (v === Math.floor(v)) return String(Math.floor(v));
                return parseFloat(v.toPrecision(3)).toString();
            }

            return [[sl_min, sl_max], fmtNum(sl_min), fmtNum(sl_max)];
        }
        """,
        Output({'type': 'histogram-slider', 'column': MATCH}, 'value', allow_duplicate=True),
        Output({'type': 'min-input', 'column': MATCH}, 'value', allow_duplicate=True),
        Output({'type': 'max-input', 'column': MATCH}, 'value', allow_duplicate=True),
        Input({'type': 'filter-reset', 'column': MATCH}, 'n_clicks'),
        State({'type': 'filter-reset', 'column': MATCH}, 'id'),
        State({'type': 'histogram-slider', 'column': MATCH}, 'min'),
        State({'type': 'histogram-slider', 'column': MATCH}, 'max'),
        prevent_initial_call=True
    )

    # --- CLIENTSIDE: RESET BUTTON VISUAL — highlight when filter is active ---
    app.clientside_callback(
        """
        function(slider_val, sl_min, sl_max) {
            var is_active = (slider_val[0] > sl_min + 1e-9 || slider_val[1] < sl_max - 1e-9);
            return {
                'fontSize': '13px', 'padding': '0px 4px', 'border': 'none',
                'borderRadius': '3px', 'cursor': 'pointer', 'fontFamily': 'Mulish',
                'backgroundColor': is_active ? '#583CA5' : 'transparent',
                'color': is_active ? '#FBB800' : '#777',
                'lineHeight': '1'
            };
        }
        """,
        Output({'type': 'filter-reset', 'column': MATCH}, 'style'),
        Input({'type': 'histogram-slider', 'column': MATCH}, 'value'),
        Input({'type': 'histogram-slider', 'column': MATCH}, 'min'),
        Input({'type': 'histogram-slider', 'column': MATCH}, 'max'),
        prevent_initial_call=True
    )

    # --- CLIENTSIDE: UPDATE HISTOGRAM BARS from cross-filter store ---
    app.clientside_callback(
        """
        function(hist_data, hist_ids) {
            if (!hist_data || !hist_ids || !hist_ids.length) return window.dash_clientside.no_update;
            return hist_ids.map(function(id_obj) {
                var col = id_obj.column;
                var d = hist_data[col];
                if (!d || !d.counts || !d.counts.length) return window.dash_clientside.no_update;
                var counts = d.counts;
                var maxCount = Math.max.apply(null, counts);
                if (maxCount === 0) maxCount = 1;
                var n = counts.length;
                return counts.map(function(c, i) {
                    return {
                        namespace: 'dash_html_components', type: 'Div',
                        props: { style: {
                            flex: '1', height: Math.round(c / maxCount * 100) + '%',
                            backgroundColor: '#FBB800', alignSelf: 'flex-end',
                            marginRight: i < n - 1 ? '1px' : '0'
                        }}
                    };
                });
            });
        }
        """,
        Output({'type': 'mini-hist', 'column': ALL}, 'children'),
        Input('filter-histogram-data', 'data'),
        State({'type': 'mini-hist', 'column': ALL}, 'id'),
        prevent_initial_call=True
    )

    # --- CLIENTSIDE: UPDATE CATEGORICAL OPTIONS with cross-filter counts ---
    app.clientside_callback(
        """
        function(hist_data, ids) {
            if (!hist_data || !ids || !ids.length) return window.dash_clientside.no_update;
            return ids.map(function(id_obj) {
                var col = id_obj.column;
                var d = hist_data[col];
                if (!d || !d.cat_counts) return window.dash_clientside.no_update;
                var counts = d.cat_counts;
                return Object.keys(counts).sort().map(function(v) {
                    return { label: v + ' [' + counts[v] + ']', value: v };
                });
            });
        }
        """,
        Output({'type': 'categorical-dropdown', 'column': ALL}, 'options'),
        Input('filter-histogram-data', 'data'),
        State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        prevent_initial_call=True
    )

    # --- CLIENTSIDE TABLE ENGINE (Custom Paging - no server round-trip) ---
    app.clientside_callback(
        """
        function(page_current, sort_by, selected_uuids, active_layer_id, slider_vals, cat_vals,
                 slider_ids, cat_ids, store_data, col_translations, edits_store, col_tooltips) {

            var placeholderShow = {'display':'flex','flexDirection':'column','alignItems':'center','justifyContent':'center','height':'100%','color':'#7f8c8d','textAlign':'center'};
            var placeholderHide = {'display':'none'};
            var tableShow = {'height':'100%','minWidth':'100%','overflowX':'auto','display':'block'};
            var tableHide = {'height':'100%','minWidth':'100%','overflowX':'auto','display':'none'};
            var PAGE_SIZE = 50;
            var emptyResult = [[], [{'name':' ','id':'_init'}], 1, placeholderShow, tableHide, [], {}];

            if (!store_data || !active_layer_id || !store_data[active_layer_id]) return emptyResult;

            var layerObj = store_data[active_layer_id];
            var columns = layerObj.columns;
            var rawData = layerObj.data;
            var colIdx = {};
            columns.forEach(function(c, i) { colIdx[c] = i; });
            var uuidIdx = colIdx['row_uuid'] !== undefined ? colIdx['row_uuid'] : 0;
            var totalRows = rawData.length;

            // --- Overlay edits from table-edits-store onto rawData BEFORE filtering/sorting ---
            var edits = edits_store || {};
            if (Object.keys(edits).length > 0) {
                var _notesI = colIdx['Notes'], _scoreI = colIdx['Score'];
                if (_notesI !== undefined || _scoreI !== undefined) {
                    rawData = rawData.map(function(row) {
                        var uuid = row[uuidIdx];
                        if (uuid && edits[uuid]) {
                            var nr = row.slice();
                            if (_notesI !== undefined && 'Notes' in edits[uuid]) nr[_notesI] = edits[uuid]['Notes'];
                            if (_scoreI !== undefined && 'Score' in edits[uuid]) nr[_scoreI] = edits[uuid]['Score'];
                            return nr;
                        }
                        return row;
                    });
                }
            }

            // --- Apply sidebar filters ---
            var filtered = [];
            for (var i = 0; i < rawData.length; i++) {
                var row = rawData[i];
                var pass = true;
                if (slider_ids && slider_vals) {
                    for (var s = 0; s < slider_ids.length; s++) {
                        var sc = slider_ids[s].column;
                        var sv = slider_vals[s];
                        if (sc && sv && colIdx[sc] !== undefined) {
                            var val = row[colIdx[sc]];
                            if (val !== null && val !== undefined && !isNaN(val)) {
                                if (val < sv[0] || val > sv[1]) { pass = false; break; }
                            }
                        }
                    }
                }
                if (pass && cat_ids && cat_vals) {
                    for (var ci = 0; ci < cat_ids.length; ci++) {
                        var cc = cat_ids[ci].column;
                        var cv = cat_vals[ci];
                        if (cc && cv && cv.length > 0 && colIdx[cc] !== undefined) {
                            if (cv.indexOf(row[colIdx[cc]]) === -1) { pass = false; break; }
                        }
                    }
                }
                if (pass) filtered.push(row);
            }
            var filteredCount = filtered.length;

            // --- Selection ---
            var isSelectAll = (selected_uuids && selected_uuids.length === 1 && selected_uuids[0] === 'ALL');
            var selectedSet = isSelectAll ? null : new Set(selected_uuids || []);
            var selectedRows;

            if (isSelectAll) {
                selectedRows = filtered;
            } else if (selectedSet && selectedSet.size > 0) {
                selectedRows = filtered.filter(function(r) { return selectedSet.has(r[uuidIdx]); });
            } else {
                var stats0 = [
                    {namespace:'dash_html_components',type:'Span',props:{children:[{namespace:'dash_html_components',type:'B',props:{children:String(totalRows)}},' total'],style:{fontSize:'10px',color:'#bdc3c7',backgroundColor:'rgba(255,255,255,0.1)',padding:'2px 8px',borderRadius:'10px'}}},
                    {namespace:'dash_html_components',type:'Span',props:{children:[{namespace:'dash_html_components',type:'B',props:{children:String(filteredCount)}},' filtered'],style:{fontSize:'10px',color:'white',backgroundColor:'#583CA5',padding:'2px 8px',borderRadius:'10px'}}},
                    {namespace:'dash_html_components',type:'Span',props:{children:[{namespace:'dash_html_components',type:'B',props:{children:'0'}},' selected'],style:{fontSize:'10px',color:'#333',backgroundColor:'#FBB800',padding:'2px 8px',borderRadius:'10px',fontWeight:'bold'}}}
                ];
                return [[], [{'name':' ','id':'_init'}], 1, placeholderShow, tableHide, stats0, {}];
            }
            var selectedCount = selectedRows.length;

            // --- Sorting ---
            if (sort_by && sort_by.length > 0) {
                var sortColId = sort_by[0].column_id;
                var direction = sort_by[0].direction;
                var origSortCol = sortColId;
                if (col_translations) {
                    for (var key in col_translations) {
                        if (col_translations[key] === sortColId) { origSortCol = key; break; }
                    }
                }
                var sIdx = colIdx[origSortCol];
                if (sIdx !== undefined) {
                    selectedRows = selectedRows.slice().sort(function(a, b) {
                        var va = a[sIdx], vb = b[sIdx];
                        if (va == null) return 1;
                        if (vb == null) return -1;
                        if (va < vb) return direction === 'asc' ? -1 : 1;
                        if (va > vb) return direction === 'asc' ? 1 : -1;
                        return 0;
                    });
                }
            }

            // --- Pagination ---
            var pageCount = Math.max(1, Math.ceil(selectedRows.length / PAGE_SIZE));
            var pg = Math.min(page_current || 0, pageCount - 1);
            var start = pg * PAGE_SIZE;
            var pageSlice = selectedRows.slice(start, start + PAGE_SIZE);

            // --- Build columns (exclude internal cols, move ID to front) ---
            var excludeCols = new Set(['row_uuid','geometry','index','orig_uuid']);
            var viewCols = columns.filter(function(c) { return !excludeCols.has(c) && c !== 'Notes' && c !== 'Score'; });
            var candidates = ['id','fid','name','settlement','station'];
            for (var k = 0; k < candidates.length; k++) {
                var found = viewCols.findIndex(function(vc) { return vc.toLowerCase() === candidates[k]; });
                if (found > 0) { viewCols.unshift(viewCols.splice(found, 1)[0]); break; }
            }
            var tableCols = viewCols.map(function(origCol) {
                var translated = (col_translations && col_translations[origCol]) ? col_translations[origCol] : origCol;
                return {name: translated, id: translated, editable: false};
            });
            var notesIdx = colIdx['Notes'];
            var scoreIdx = colIdx['Score'];
            if (notesIdx !== undefined) tableCols.push({name: 'Notes', id: 'Notes', editable: true});
            if (scoreIdx !== undefined) tableCols.push({name: 'Score', id: 'Score', editable: true});

            // --- Build tooltip_header (EN original col name -> description in active language) ---
            var tooltipHeader = {};
            if (col_tooltips) {
                viewCols.forEach(function(origCol) {
                    var desc = col_tooltips[origCol];
                    if (desc) {
                        var translated = (col_translations && col_translations[origCol]) ? col_translations[origCol] : origCol;
                        tooltipHeader[translated] = {value: desc, use_with: 'header'};
                    }
                });
            }

            // --- Build records (only current page) ---
            var records = pageSlice.map(function(row) {
                var rec = {};
                var uuid = row[uuidIdx];
                viewCols.forEach(function(origCol) {
                    var translated = (col_translations && col_translations[origCol]) ? col_translations[origCol] : origCol;
                    var v = row[colIdx[origCol]];
                    rec[translated] = (v !== null && v !== undefined) ? v : '';
                });
                if (notesIdx !== undefined) rec['Notes'] = row[notesIdx] || '';
                if (scoreIdx !== undefined) rec['Score'] = (row[scoreIdx] != null) ? row[scoreIdx] : 0;
                rec['orig_uuid'] = uuid;
                return rec;
            });

            // --- Stats badges ---
            var stats = [
                {namespace:'dash_html_components',type:'Span',props:{children:[{namespace:'dash_html_components',type:'B',props:{children:String(totalRows)}},' total'],style:{fontSize:'10px',color:'#bdc3c7',backgroundColor:'rgba(255,255,255,0.1)',padding:'2px 8px',borderRadius:'10px'}}},
                {namespace:'dash_html_components',type:'Span',props:{children:[{namespace:'dash_html_components',type:'B',props:{children:String(filteredCount)}},' filtered'],style:{fontSize:'10px',color:'white',backgroundColor:'#583CA5',padding:'2px 8px',borderRadius:'10px'}}},
                {namespace:'dash_html_components',type:'Span',props:{children:[{namespace:'dash_html_components',type:'B',props:{children:String(selectedCount)}},' selected'],style:{fontSize:'10px',color:'#333',backgroundColor:'#FBB800',padding:'2px 8px',borderRadius:'10px',fontWeight:'bold'}}}
            ];

            return [records, tableCols, pageCount, placeholderHide, tableShow, stats, tooltipHeader];
        }
        """,
        [Output({'type': 'table-export', 'index': 0}, 'data'),
         Output({'type': 'table-export', 'index': 0}, 'columns'),
         Output({'type': 'table-export', 'index': 0}, 'page_count'),
         Output('table-placeholder', 'style'),
         Output({'type': 'table-export', 'index': 0}, 'style_table'),
         Output('header-stats-container', 'children'),
         Output({'type': 'table-export', 'index': 0}, 'tooltip_header')],
        [Input({'type': 'table-export', 'index': 0}, 'page_current'),
         Input({'type': 'table-export', 'index': 0}, 'sort_by'),
         Input('current-selection-store', 'data'),
         Input('table-layer-dropdown', 'value'),
         Input({'type': 'histogram-slider', 'column': ALL}, 'value'),
         Input({'type': 'categorical-dropdown', 'column': ALL}, 'value')],
        [State({'type': 'histogram-slider', 'column': ALL}, 'id'),
         State({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
         State('client-layer-data', 'data'),
         State('column-translations-store', 'data'),
         State('table-edits-store', 'data'),
         State('column-tooltips-store', 'data')]
    )

    # --- NEW: CLIENTSIDE MAP UPDATE ---
    # --- NEW: CLIENTSIDE MAP UPDATE (FIXED Z-INDEX ORDER & PERFORMANCE) ---
    # --- NEW: CLIENTSIDE MAP (WITH VISIBILITY FILTER) ---
    # --- NEW: CLIENTSIDE MAP (FIXED INPUT/STATE ORDER) ---
    # --- NEW: CLIENTSIDE MAP (DETAILED POPUPS & NEW COLORS) ---
    # --- NEW: CLIENTSIDE MAP (FIXED DISPLAY ID & NULL HANDLING) ---
    # --- NEW: CLIENTSIDE MAP (FIXED BASEMAP) ---
    # --- NEW: CLIENTSIDE MAP (FIXED TOKEN FOR MAPBOX STYLES) ---
    # --- NEW: CLIENTSIDE MAP (DYNAMIC ZOOM SIZE & OPACITY FIX) ---
    app.clientside_callback(
        """
        function(storeData, slider_vals, cat_vals, active_layer_id,
                 sizes, opacities, color_cols, color_scales, single_colors,
                 slider_ids, cat_ids, size_ids, opacity_ids, color_col_ids, scale_ids, single_color_ids,
                 fill_colors, fill_color_ids, fill_opacities, fill_opacity_ids,
                 outline_colors, outline_color_ids, outline_weights, outline_weight_ids,
                 outline_opacities, outline_opacity_ids,
                 n_clicks_select, n_clicks_deselect, map_selected_data,
                 map_relayout, map_view, visible_layers, basemap_style, mapbox_token, edits_store, bubble_clear,
                 current_selection_list, active_tool, bubble_highlight) {

            // Pass Mapbox token to Plotly config (needed for satellite/dark/light styles)
            if (mapbox_token) Plotly.setPlotConfig({mapboxAccessToken: mapbox_token});

            if (!storeData || Object.keys(storeData).length === 0) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }

            const ctx = dash_clientside.callback_context;
            const trigger = ctx.triggered.length ? ctx.triggered[0].prop_id : "";

            // --- Cache Init ---
            if (!window.agcap_layer_cache) window.agcap_layer_cache = {};
            if (trigger.includes('client-layer-data') || trigger.includes('table-edits-store')) {
                window.agcap_layer_cache = {};
                window.__agcap_geojson = {};
            }

            // --- Early-return on pure pan/zoom: map handles it natively, no rebuild needed ---
            if (trigger.includes('map.relayoutData')) {
                return [window.dash_clientside.no_update, window.dash_clientside.no_update];
            }

            // --- 1. Get Current Zoom for Dynamic Sizing ---
            // Priority: 1. Manual User Zoom (relayout), 2. Programmatic View (store), 3. Default
            let currentZoom = 5;
            if (map_relayout && map_relayout['map.zoom']) {
                currentZoom = map_relayout['map.zoom'];
            } else if (map_view && map_view.zoom) {
                currentZoom = map_view.zoom;
            }

            // --- 2. Selection Intent ---
            let isAllFlag = (current_selection_list && current_selection_list.length === 1 && current_selection_list[0] === 'ALL');
            let targetUUIDs = isAllFlag ? null : new Set(current_selection_list || []);
            let selectAllTriggered = trigger.includes('select-all-btn');
            let deselectTriggered = trigger.includes('deselect-all-btn');
            let mapInteractionTriggered = trigger.includes('map.selectedData');
            let forceViewUpdate = trigger.includes('map-view-store');

            if (mapInteractionTriggered && map_selected_data && map_selected_data.points) {
                isAllFlag = false;
                targetUUIDs = new Set();
                const relevantPoints = map_selected_data.points.filter(p => p.customdata && p.customdata[3] === active_layer_id);
                relevantPoints.forEach(p => targetUUIDs.add(p.customdata[0]));
            } else if (deselectTriggered) {
                isAllFlag = false;
                targetUUIDs = new Set();
            }

            // --- 3. Render Layers ---
            const orderedLayerIDs = size_ids ? size_ids.map(o => o.index).reverse() : Object.keys(storeData);
            const visibleSet = new Set(visible_layers || []);

            const findSetting = (vals, ids, lid, fallback) => {
                if (!ids || !vals) return fallback;
                const idx = ids.findIndex(i => i.index === lid);
                return idx > -1 ? vals[idx] : fallback;
            };

            const dataTraces = [];
            const polygonMapboxLayers = [];
            const finalSelectedUUIDs = [];

            // --- Colorscale helpers for polygon data-driven coloring ---
            const SCALE_STOPS = {
                'Turbo':    [[0,'#30123b'],[0.14,'#4145ab'],[0.28,'#3d87f5'],[0.43,'#1bcfd4'],[0.57,'#4df884'],[0.71,'#a2fc3c'],[0.86,'#fdb827'],[1,'#7a0403']],
                'Viridis':  [[0,'#440154'],[0.25,'#31688e'],[0.5,'#35b779'],[0.75,'#90d743'],[1,'#fde725']],
                'Plasma':   [[0,'#0d0887'],[0.25,'#7e03a8'],[0.5,'#cc4778'],[0.75,'#f89540'],[1,'#f0f921']],
                'Portland': [[0,'#0c3383'],[0.25,'#0a88ba'],[0.5,'#f2d338'],[0.75,'#f28f38'],[1,'#d91e1e']],
                'Jet':      [[0,'#000080'],[0.25,'#00bfff'],[0.5,'#00ff00'],[0.75,'#ffff00'],[1,'#800000']],
                'RdBu':     [[0,'#053061'],[0.25,'#4393c3'],[0.5,'#f7f7f7'],[0.75,'#d6604d'],[1,'#67001f']],
                'Hot':      [[0,'#000000'],[0.33,'#ff0000'],[0.67,'#ffff00'],[1,'#ffffff']],
                'Electric': [[0,'#000000'],[0.15,'#1e0064'],[0.4,'#7b00d4'],[0.65,'#ff0099'],[0.9,'#ff9d00'],[1,'#ffffff']],
                'Rainbow':  [[0,'#96005a'],[0.2,'#0000c8'],[0.4,'#0096ff'],[0.6,'#ffff00'],[0.8,'#ff3a00'],[1,'#c8000a']],
                'Earth':    [[0,'#000082'],[0.25,'#004da8'],[0.5,'#76b900'],[0.75,'#c8a400'],[1,'#ffffff']],
                'Blackbody':[[0,'#000000'],[0.25,'#e00000'],[0.5,'#e0d000'],[0.75,'#ffffff'],[1,'#a0c8ff']],
            };
            function scaleColor(norm, scaleName) {
                const stops = SCALE_STOPS[scaleName] || SCALE_STOPS['Portland'];
                if (norm <= stops[0][0]) return stops[0][1];
                if (norm >= stops[stops.length-1][0]) return stops[stops.length-1][1];
                for (let i = 1; i < stops.length; i++) {
                    if (norm <= stops[i][0]) {
                        const t = (norm - stops[i-1][0]) / (stops[i][0] - stops[i-1][0]);
                        return t < 0.5 ? stops[i-1][1] : stops[i][1];
                    }
                }
                return stops[stops.length-1][1];
            }
            const CAT_PALETTE = ['#583CA5','#e74c3c','#2ecc71','#3498db','#9b59b6','#f1c40f','#e67e22','#1abc9c','#e91e63','#00bcd4','#ff5722','#607d8b'];

            orderedLayerIDs.forEach(lid => {
                if (!visibleSet.has(lid)) return;

                const layerObj = storeData[lid];
                if (!layerObj) return;

                // --- POLYGON LAYER PATH ---
                if (layerObj.layer_type === 'polygon') {
                    if (!layerObj.geojson) return;

                    if (!window.__agcap_geojson) window.__agcap_geojson = {};
                    if (!window.__agcap_geojson[lid]) window.__agcap_geojson[lid] = layerObj.geojson;
                    const geojson = window.__agcap_geojson[lid];

                    const fillColor = findSetting(fill_colors, fill_color_ids, lid, '#3498db');
                    const fillOpacity = Number(findSetting(fill_opacities, fill_opacity_ids, lid, 0.4));
                    const outlineColor = findSetting(outline_colors, outline_color_ids, lid, '#333333');
                    const outlineWeight = Number(findSetting(outline_weights, outline_weight_ids, lid, 2));
                    const outlineOpacity = Number(findSetting(outline_opacities, outline_opacity_ids, lid, 1.0));
                    const colorCol = findSetting(color_cols, color_col_ids, lid, 'Single Color');
                    const colorScale = findSetting(color_scales, scale_ids, lid, 'Portland');
                    const useColormap = (colorCol && colorCol !== 'Single Color');

                    if (!useColormap) {
                        polygonMapboxLayers.push({
                            sourcetype: 'geojson', source: geojson,
                            type: 'fill', color: fillColor, opacity: fillOpacity, below: 'traces'
                        });
                    } else {
                        const columns = layerObj.columns || [];
                        const colIdx2 = {};
                        columns.forEach((c, i) => { colIdx2[c] = i; });
                        const cIdx2 = colIdx2[colorCol];
                        const uuidIdx2 = colIdx2['row_uuid'] !== undefined ? colIdx2['row_uuid'] : 0;

                        if (cIdx2 !== undefined) {
                            const rows = layerObj.data || [];
                            const uuidToVal = {};
                            rows.forEach(r => { uuidToVal[String(r[uuidIdx2])] = r[cIdx2]; });

                            const vals = rows.map(r => r[cIdx2]).filter(v => v !== null && v !== '');
                            const isNumeric = vals.length > 0 && vals.every(v => !isNaN(Number(v)));

                            if (isNumeric) {
                                const nums = vals.map(Number);
                                const mn = nums.reduce((a, b) => a < b ? a : b, nums[0]);
                                const mx = nums.reduce((a, b) => a > b ? a : b, nums[0]);
                                const range = mx - mn || 1;
                                const N_STEPS = 256;
                                const bucketMap = {};
                                geojson.features.forEach(f => {
                                    const uuid = String(f.properties && (f.properties.row_uuid || f.properties.index) || '');
                                    const v = Number(uuidToVal[uuid]);
                                    if (isNaN(v)) return;
                                    const norm = Math.max(0, Math.min(1, (v - mn) / range));
                                    const bucket = Math.floor(norm * (N_STEPS - 1));
                                    if (!bucketMap[bucket]) bucketMap[bucket] = [];
                                    bucketMap[bucket].push(f);
                                });
                                Object.keys(bucketMap).forEach(bKey => {
                                    const norm = Number(bKey) / (N_STEPS - 1);
                                    const color = scaleColor(norm, colorScale);
                                    polygonMapboxLayers.push({
                                        sourcetype: 'geojson',
                                        source: { type: 'FeatureCollection', features: bucketMap[bKey] },
                                        type: 'fill', color: color, opacity: fillOpacity, below: 'traces'
                                    });
                                });
                            } else {
                                const uniqueVals = [...new Set(vals.map(String))];
                                const nCats = uniqueVals.length;
                                const catIndex = {};
                                uniqueVals.forEach((v, i) => { catIndex[v] = i; });
                                const catBuckets = {};
                                geojson.features.forEach(f => {
                                    const uuid = String(f.properties && (f.properties.row_uuid || f.properties.index) || '');
                                    const rawVal = String(uuidToVal[uuid] !== undefined ? uuidToVal[uuid] : '');
                                    if (catIndex[rawVal] === undefined) return;
                                    if (!catBuckets[rawVal]) catBuckets[rawVal] = [];
                                    catBuckets[rawVal].push(f);
                                });
                                Object.keys(catBuckets).forEach(catVal => {
                                    const norm = nCats > 1 ? catIndex[catVal] / (nCats - 1) : 0.5;
                                    const color = scaleColor(norm, colorScale);
                                    polygonMapboxLayers.push({
                                        sourcetype: 'geojson',
                                        source: { type: 'FeatureCollection', features: catBuckets[catVal] },
                                        type: 'fill', color: color, opacity: fillOpacity, below: 'traces'
                                    });
                                });
                            }
                        }
                    }

                    // Outline layer (always on top of fill)
                    if (outlineWeight > 0) {
                        polygonMapboxLayers.push({
                            sourcetype: 'geojson', source: geojson,
                            type: 'line', color: outlineColor,
                            line: { width: outlineWeight }, opacity: outlineOpacity
                        });
                    }

                    // Transparent centroid trace for hover/click interactivity
                    const columns = layerObj.columns || [];
                    const colIdx2 = {};
                    columns.forEach((c, i) => { colIdx2[c] = i; });
                    const cLatIdx = colIdx2['centroid_lat'];
                    const cLonIdx = colIdx2['centroid_lon'];
                    const uuidIdx2 = colIdx2['row_uuid'] !== undefined ? colIdx2['row_uuid'] : 0;
                    if (cLatIdx !== undefined && cLonIdx !== undefined) {
                        const cLats = [], cLons = [], cUuids = [], cHovers = [], cCustom = [];
                        const colorCol2 = findSetting(color_cols, color_col_ids, lid, 'Single Color');
                        const cIdx2 = colIdx2[colorCol2];
                        const rows = layerObj.data || [];
                        const excludeCols = new Set(['centroid_lat','centroid_lon','row_uuid','geometry','index','orig_uuid']);
                        rows.forEach(row => {
                            cLats.push(row[cLatIdx]);
                            cLons.push(row[cLonIdx]);
                            const uuid = row[uuidIdx2];
                            cUuids.push(uuid);
                            cCustom.push([uuid, row[cIdx2 !== undefined ? cIdx2 : 0], lid, lid]);
                            let ht = `<b>${layerObj.name}</b><br>`;
                            let cnt = 0;
                            columns.forEach((c, i) => {
                                if (cnt >= 8 || excludeCols.has(c)) return;
                                ht += `<b>${c}</b>: ${row[i]}<br>`;
                                cnt++;
                            });
                            cHovers.push(ht);
                        });
                        dataTraces.push({
                            type: 'scattermap',
                            lat: cLats, lon: cLons,
                            mode: 'markers',
                            name: lid,
                            marker: { size: 8, opacity: 0, color: 'rgba(0,0,0,0)' },
                            text: cHovers,
                            customdata: cCustom,
                            hovertemplate: '%{text}<extra></extra>',
                        });
                    }
                    return; // done with polygon layer
                }

                // --- POINT LAYER PATH (existing logic) ---

                // 3a. Settings
                const rawSize = findSetting(sizes, size_ids, lid, 4);
                const opacity = Number(findSetting(opacities, opacity_ids, lid, 0.8)); // Force Number
                const colorCol = findSetting(color_cols, color_col_ids, lid, 'Single Color');
                const colorScale = findSetting(color_scales, scale_ids, lid, 'Turbo'); 
                let sColor = findSetting(single_colors, single_color_ids, lid, '#FBB800');
                if (sColor && typeof sColor === 'object' && sColor.hex) sColor = sColor.hex;

                const isActive = (lid === active_layer_id);
                
                // --- DYNAMIC SIZE LOGIC ---
                // If zoom > 8, increase size by 2.5x (Client-side implementation of old logic)
                let renderedSize = rawSize;
                if (currentZoom > 8) {
                    renderedSize = rawSize * 2.5;
                }

                // Cache Signature (Includes renderedSize to invalidate on zoom change)
                const settingsSig = `${renderedSize}-${opacity}-${colorCol}-${colorScale}-${sColor}`;

                if (!isActive && !forceViewUpdate && window.agcap_layer_cache[lid] && window.agcap_layer_cache[lid].sig === settingsSig) {
                    dataTraces.push(window.agcap_layer_cache[lid].trace);
                    return; 
                }
                
                // 3b. Data Processing
                const columns = layerObj.columns;
                let rawData = layerObj.data;
                const layerName = layerObj.name || "Layer";
                const colIdx = {};
                columns.forEach((c, i) => { colIdx[c] = i; });

                // 3b-bis. Overlay edits from table-edits-store onto rawData
                if (edits_store && Object.keys(edits_store).length > 0 && colIdx['row_uuid'] !== undefined) {
                    const ruIdx = colIdx['row_uuid'];
                    const notesIdx = colIdx['Notes'];
                    const scoreIdx = colIdx['Score'];
                    if (notesIdx !== undefined || scoreIdx !== undefined) {
                        rawData = rawData.map(row => {
                            const uuid = row[ruIdx];
                            if (uuid && edits_store[uuid]) {
                                const newRow = row.slice();
                                if (notesIdx !== undefined && 'Notes' in edits_store[uuid]) newRow[notesIdx] = edits_store[uuid]['Notes'];
                                if (scoreIdx !== undefined && 'Score' in edits_store[uuid]) newRow[scoreIdx] = edits_store[uuid]['Score'];
                                return newRow;
                            }
                            return row;
                        });
                    }
                }

                const sliderIndices = (isActive && slider_ids) ? slider_ids.map(s => colIdx[s.column]) : [];
                const catIndices = (isActive && cat_ids) ? cat_ids.map(c => colIdx[c.column]) : [];

                let displayIDIdx = 0;
                const candidates = ['id', 'fid', 'name', 'settlement', 'station'];
                for (let cand of candidates) {
                    const found = columns.findIndex(c => c.toLowerCase() === cand);
                    if (found > -1) { displayIDIdx = found; break; }
                }

                const cIdx = colIdx[colorCol];
                const useColormap = (colorCol && colorCol !== 'Single Color' && cIdx !== undefined);

                const latIdx = colIdx['lat'];
                const lonIdx = colIdx['lon'];
                const idIdx = colIdx['ID'] !== undefined ? colIdx['ID'] : 0; 
                const uuidIdx = colIdx['row_uuid'] !== undefined ? colIdx['row_uuid'] : (colIdx['index'] !== undefined ? colIdx['index'] : idIdx);

                const lats = [];
                const lons = [];
                const uuids = []; 
                const hoverTexts = []; 
                const selectedIndices = [];
                let markerColorArr = [];
                const displayIDs = [];

                for (let i = 0; i < rawData.length; i++) {
                    const row = rawData[i];
                    let keep = true;

                    if (isActive) {
                        if (slider_ids && slider_vals) {
                            for (let j = 0; j < slider_ids.length; j++) {
                                const val = row[sliderIndices[j]];
                                const range = slider_vals[j];
                                if (val !== undefined && range) {
                                    if (val < range[0] || val > range[1]) { keep = false; break; }
                                }
                            }
                        }
                        if (keep && cat_ids && cat_vals) {
                            for (let j = 0; j < cat_ids.length; j++) {
                                const val = row[catIndices[j]];
                                const accepted = cat_vals[j];
                                if (accepted && accepted.length > 0) {
                                    if (!accepted.includes(String(val))) { keep = false; break; }
                                }
                            }
                        }
                    } 

                    if (keep) {
                        const u = row[uuidIdx];
                        lats.push(row[latIdx]);
                        lons.push(row[lonIdx]);
                        uuids.push(u);
                        
                        const dID = row[displayIDIdx];
                        displayIDs.push(dID);

                        if (useColormap) markerColorArr.push(row[cIdx]); 

                        let hText = "";
                        const safeVal = (useColormap && row[cIdx] != null) ? row[cIdx] : "N/A";

                        if (layerName === "Settlements") {
                             const colorDisplay = useColormap ? `<b>${colorCol}</b>: ${safeVal}` : `<b>${layerName}</b>`;
                             hText = `${colorDisplay}<br>ID: ${dID}`;
                        } 
                        else {
                            hText = (useColormap) ? `<b>${colorCol}</b>: ${safeVal}<br>` : "";
                            hText += `<b>ID</b>: ${dID}<br>`;
                            let count = 0;
                            for(let k=0; k < columns.length; k++) {
                                if (count >= 10) break;
                                const colName = columns[k];
                                if (['row_uuid', 'geometry', 'lat', 'lon', 'index', 'orig_uuid', colorCol].includes(colName)) continue;
                                if (k === displayIDIdx) continue;
                                hText += `<b>${colName}</b>: ${row[k]}<br>`;
                                count++;
                            }
                        }
                        hoverTexts.push(hText);

                        if (isActive && !deselectTriggered && (selectAllTriggered || isAllFlag || (targetUUIDs && targetUUIDs.has(u)))) {
                            selectedIndices.push(lats.length - 1);
                            if (!selectAllTriggered && !isAllFlag) finalSelectedUUIDs.push(u);
                        }
                    }
                }

                let finalColor = useColormap ? markerColorArr : sColor;

                const customdata = uuids.map((u, i) => {
                    const cVal = useColormap ? markerColorArr[i] : 0;
                    return [u, cVal, displayIDs[i], lid]; 
                });

                const trace = {
                    type: "scattermap",
                    lat: lats,
                    lon: lons,
                    mode: "markers",
                    name: lid,
                    marker: {
                        size: renderedSize, // <--- Using dynamic size
                        opacity: opacity,
                        color: finalColor,
                        colorscale: colorScale
                    },
                    text: hoverTexts, 
                    customdata: customdata,
                    hovertemplate: "%{text}<extra></extra>",
                    selectedpoints: (isActive && selectedIndices.length > 0) ? selectedIndices : null,
                    selected: { marker: { opacity: 1, size: renderedSize + 8, color: '#583CA5' } }, 
                    unselected: { marker: { opacity: opacity * 0.8 } }
                };
                
                if (!isActive) window.agcap_layer_cache[lid] = { trace: trace, sig: settingsSig };
                dataTraces.push(trace);
            });

            // Emit "ALL" flag for select-all, empty array for deselect
            if (deselectTriggered) {
                finalSelectedUUIDs.length = 0;
            } else if (selectAllTriggered || (isAllFlag && !mapInteractionTriggered)) {
                finalSelectedUUIDs.length = 0;
                finalSelectedUUIDs.push('ALL');
            }

            // --- 4. Layout ---
            let center = (map_view && map_view.center) ? map_view.center : {lat: -18.66, lon: 35.52};
            let zoom = (map_view && map_view.zoom) ? map_view.zoom : 5;
            const uirevision_val = forceViewUpdate ? Date.now() : 'constant';

            const currentStyle = basemap_style || "open-street-map";

            const layout = {
                map: {
                    style: currentStyle,
                    center: center,
                    zoom: zoom,
                    layers: polygonMapboxLayers
                },
                margin: {r: 0, t: 0, l: 0, b: 0},
                showlegend: false,
                paper_bgcolor: '#44546A',
                plot_bgcolor: '#44546A',
                clickmode: 'event+select',
                uirevision: uirevision_val,
                dragmode: active_tool || 'pan'
            };

            // --- Bubble chart highlight marker on map ---
            if (bubble_highlight && bubble_highlight.lat != null && bubble_highlight.lon != null) {
                dataTraces.push({
                    type: "scattermap",
                    lat: [bubble_highlight.lat],
                    lon: [bubble_highlight.lon],
                    mode: "markers",
                    name: "bubble-glow",
                    marker: { size: 35, color: 'rgba(0, 255, 136, 0.25)', opacity: 1 },
                    hoverinfo: 'skip',
                    showlegend: false
                });
                dataTraces.push({
                    type: "scattermap",
                    lat: [bubble_highlight.lat],
                    lon: [bubble_highlight.lon],
                    mode: "markers",
                    name: "bubble-pin",
                    marker: { size: 14, color: '#00ff88', opacity: 1 },
                    hoverinfo: 'skip',
                    showlegend: false
                });
            }

            return [{data: dataTraces, layout: layout}, finalSelectedUUIDs];
        }
        """,
        [Output('map', 'figure'), Output('current-selection-store', 'data')],
        Input('client-layer-data', 'data'),
        Input({'type': 'histogram-slider', 'column': ALL}, 'value'),
        Input({'type': 'categorical-dropdown', 'column': ALL}, 'value'),
        Input('table-layer-dropdown', 'value'), 
        Input({'type': 'layer-size', 'index': ALL}, 'value'),
        Input({'type': 'layer-opacity', 'index': ALL}, 'value'),
        Input({'type': 'layer-color-col', 'index': ALL}, 'value'),
        Input({'type': 'layer-colorscale', 'index': ALL}, 'value'),
        Input({'type': 'layer-single-color', 'index': ALL}, 'value'),
        Input({'type': 'histogram-slider', 'column': ALL}, 'id'),
        Input({'type': 'categorical-dropdown', 'column': ALL}, 'id'),
        Input({'type': 'layer-size', 'index': ALL}, 'id'),
        Input({'type': 'layer-opacity', 'index': ALL}, 'id'),
        Input({'type': 'layer-color-col', 'index': ALL}, 'id'),
        Input({'type': 'layer-colorscale', 'index': ALL}, 'id'),
        Input({'type': 'layer-single-color', 'index': ALL}, 'id'),
        Input({'type': 'layer-fill-color', 'index': ALL}, 'value'),
        Input({'type': 'layer-fill-color', 'index': ALL}, 'id'),
        Input({'type': 'layer-fill-opacity', 'index': ALL}, 'value'),
        Input({'type': 'layer-fill-opacity', 'index': ALL}, 'id'),
        Input({'type': 'layer-outline-color', 'index': ALL}, 'value'),
        Input({'type': 'layer-outline-color', 'index': ALL}, 'id'),
        Input({'type': 'layer-outline-weight', 'index': ALL}, 'value'),
        Input({'type': 'layer-outline-weight', 'index': ALL}, 'id'),
        Input({'type': 'layer-outline-opacity', 'index': ALL}, 'value'),
        Input({'type': 'layer-outline-opacity', 'index': ALL}, 'id'),
        Input('select-all-btn', 'n_clicks'),
        Input('deselect-all-btn', 'n_clicks'),
        Input('map', 'selectedData'),
        Input('map', 'relayoutData'),
        Input('map-view-store', 'data'),
        Input('visible-layers-list', 'data'),
        Input('basemap-dropdown', 'value'),
        Input('mapbox-token-input', 'value'),
        Input('table-edits-store', 'data'),
        Input('bubble-clear-trigger', 'data'),
        State('current-selection-store', 'data'),
        State('active-tool-store', 'data'),
        State('bubble-clicked-point', 'data')
    )

    # ==================== COMPOSITE INDEX BUILDER CALLBACKS ====================

    # 1. Populate Layer Dropdown
    @app.callback(
        Output('composite-layer-dropdown', 'options'),
        Output('composite-layer-dropdown', 'value'),
        Input('user-session-id', 'data')
    )
    def update_composite_layer_dropdown(session_id):
        if not session_id:
            return [], None
        metadata = dm.get_metadata(session_id)
        if not metadata:
            return [], None
        options = [{'label': m['name'], 'value': m['id']} for m in metadata]
        default_value = next((m['id'] for m in metadata if m.get('is_primary')), None)
        return options, default_value

    # 2. Populate Column Dropdown
    @app.callback(
        Output('composite-columns-dropdown', 'options'),
        Input('composite-layer-dropdown', 'value'),
        State('user-session-id', 'data')
    )
    def update_composite_columns_dropdown(layer_id, session_id):
        if not session_id or not layer_id:
            return []
        metadata = dm.get_metadata(session_id)
        layer_meta = next((m for m in metadata if m['id'] == layer_id), None)
        if not layer_meta:
            return []
        numeric_cols = layer_meta.get('numeric_columns', [])
        exclude_cols = ['ID', 'row_uuid', 'orig_uuid', 'Notes', 'Score', 'lat', 'lon']
        numeric_cols = [c for c in numeric_cols if c not in exclude_cols]
        lang = app.lang
        options = [{'label': get_column_translation(c, lang, COUNTRY_CODE), 'value': c}
                   for c in numeric_cols]
        return options

    # 3. Render Weight Controls
    @app.callback(
        Output('composite-weight-controls-container', 'children'),
        Input('composite-columns-dropdown', 'value'),
        State('user-session-id', 'data'),
        State('composite-layer-dropdown', 'value')
    )
    def render_weight_controls(selected_columns, session_id, layer_id):
        if not selected_columns or len(selected_columns) < 2:
            return html.Div(t('composite.error_min2', app.lang),
                          style={'fontSize':'10px', 'color':'#e74c3c', 'fontStyle':'italic', 'padding':'10px'})

        lang = app.lang
        n = len(selected_columns)
        equal_weight = round(100 / n, 2)
        df = dm.get_df(session_id, layer_id) if session_id and layer_id else None

        controls = []
        for i, col in enumerate(selected_columns):
            translated_name = get_column_translation(col, lang, COUNTRY_CODE)
            histogram = html.Div()
            if df is not None and col in df.columns:
                col_data = pd.to_numeric(df[col], errors='coerce').dropna()
                if len(col_data) > 0:
                    hist_fig = go.Figure(go.Histogram(x=col_data, nbinsx=40, marker_color='#FBB800'))
                    hist_fig.update_layout(
                        height=50, margin=dict(l=0, r=0, t=0, b=0),
                        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                        xaxis=dict(visible=False), yaxis=dict(visible=False), bargap=0.05
                    )
                    histogram = dcc.Graph(figure=hist_fig, config={'displayModeBar': False},
                                          style={'height':'50px', 'marginBottom':'2px'})

            controls.append(
                html.Div([
                    histogram,
                    html.Div([
                        html.Label(translated_name, style={'fontSize':'10px', 'color':'#ecf0f1', 'fontWeight':'bold', 'flex':'1'}),
                        html.Div([
                            html.Button(
                                [html.I(className="fas fa-minus-circle", style={'marginRight':'2px'})],
                                id={'type': 'weight-negative', 'index': i},
                                n_clicks=0,
                                title=t('composite.negative_effect', app.lang),
                                style={'padding':'2px 5px', 'fontSize':'10px', 'marginRight':'4px',
                                       'backgroundColor':'#555', 'color':'#ecf0f1', 'border':'none',
                                       'borderRadius':'3px', 'cursor':'pointer'}
                            ),
                            dcc.Input(
                                id={'type': 'weight-input', 'index': i},
                                type='number', value=equal_weight, min=0, max=100, step=0.01,
                                style={'width':'65px', 'fontSize':'10px', 'padding':'2px 4px', 'textAlign':'center',
                                       'backgroundColor':'#333', 'color':'#ecf0f1', 'border':'1px solid #555', 'borderRadius':'3px'}
                            ),
                            html.Span("%", style={'fontSize':'10px', 'color':'#bbb', 'marginLeft':'2px'})
                        ], style={'display':'flex', 'alignItems':'center'})
                    ], style={'display':'flex', 'justifyContent':'space-between', 'alignItems':'center', 'marginBottom':'2px'}),
                    dcc.Slider(
                        id={'type': 'weight-slider', 'index': i},
                        min=0, max=100, step=0.01, value=equal_weight,
                        marks=None, tooltip={"placement": "bottom", "always_visible": False}
                    ),
                    dcc.Store(id={'type': 'weight-column-name', 'index': i}, data=col)
                ], style={'marginBottom':'8px', 'padding':'6px', 'backgroundColor':'#2a2a2a', 'borderRadius':'4px', 'border':'1px solid #444'})
            )
        return controls

    # 3b. Clientside: Sync slider <-> text input
    app.clientside_callback(
        """
        function(slider_values, input_values) {
            const ctx = dash_clientside.callback_context;
            if (!ctx.triggered || !ctx.triggered.length) {
                return [slider_values, input_values];
            }
            const triggered = ctx.triggered[0].prop_id;
            const n = slider_values ? slider_values.length : 0;
            if (n === 0) return [[], []];
            if (triggered.includes('weight-slider')) {
                return [window.dash_clientside.no_update, slider_values.map(v => Math.round(v * 100) / 100)];
            } else if (triggered.includes('weight-input')) {
                return [input_values.map(v => parseFloat(v) || 0), window.dash_clientside.no_update];
            }
            return [slider_values, input_values];
        }
        """,
        [Output({'type': 'weight-slider', 'index': ALL}, 'value'),
         Output({'type': 'weight-input', 'index': ALL}, 'value')],
        [Input({'type': 'weight-slider', 'index': ALL}, 'value'),
         Input({'type': 'weight-input', 'index': ALL}, 'value')]
    )

    # 3c. Clientside: Validate sum & update button style
    app.clientside_callback(
        """
        function(slider_values) {
            if (!slider_values || slider_values.length === 0) {
                return ['0%', {'color': '#e74c3c', 'fontSize':'11px', 'fontWeight':'bold'}, {
                    'flex':'2', 'padding':'8px', 'backgroundColor':'#555',
                    'color':'#888', 'border':'none', 'borderRadius':'4px',
                    'cursor':'not-allowed', 'fontWeight':'bold', 'fontSize':'11px',
                    'fontFamily':'Barlow Condensed', 'letterSpacing':'1px'
                }];
            }
            var sum = slider_values.reduce(function(a, b) { return a + b; }, 0);
            var is_valid = (Math.abs(sum - 100) < 0.1);
            var sum_text = sum.toFixed(2) + '%';
            var sum_style = is_valid
                ? {'color': '#2ecc71', 'fontSize':'11px', 'fontWeight':'bold'}
                : {'color': '#e74c3c', 'fontSize':'11px', 'fontWeight':'bold'};
            var btn_style = is_valid ? {
                'flex':'2', 'padding':'8px', 'backgroundColor':'#FBB800',
                'color':'#333', 'border':'none', 'borderRadius':'4px',
                'cursor':'pointer', 'fontWeight':'bold', 'fontSize':'11px',
                'fontFamily':'Barlow Condensed', 'letterSpacing':'1px'
            } : {
                'flex':'2', 'padding':'8px', 'backgroundColor':'#555',
                'color':'#888', 'border':'none', 'borderRadius':'4px',
                'cursor':'not-allowed', 'fontWeight':'bold', 'fontSize':'11px',
                'fontFamily':'Barlow Condensed', 'letterSpacing':'1px'
            };
            return [sum_text, sum_style, btn_style];
        }
        """,
        [Output('composite-sum-display', 'children'),
         Output('composite-sum-display', 'style'),
         Output('composite-calculate-btn', 'style')],
        Input({'type': 'weight-slider', 'index': ALL}, 'value')
    )

    # 3d. Clientside: Negative effect button toggle
    app.clientside_callback(
        """
        function(click_counts) {
            if (!click_counts || click_counts.length === 0) return [];
            return click_counts.map(function(n) {
                var is_negative = (n % 2 === 1);
                return {
                    'padding':'2px 5px', 'fontSize':'10px', 'marginRight':'4px',
                    'backgroundColor': is_negative ? '#e74c3c' : '#555',
                    'color':'#ecf0f1', 'border':'none',
                    'borderRadius':'3px', 'cursor':'pointer'
                };
            });
        }
        """,
        Output({'type': 'weight-negative', 'index': ALL}, 'style'),
        Input({'type': 'weight-negative', 'index': ALL}, 'n_clicks')
    )

    # 3e. Reset Weights
    @app.callback(
        [Output({'type': 'weight-slider', 'index': ALL}, 'value', allow_duplicate=True),
         Output({'type': 'weight-negative', 'index': ALL}, 'n_clicks', allow_duplicate=True)],
        Input('composite-reset-weights-btn', 'n_clicks'),
        State('composite-columns-dropdown', 'value'),
        prevent_initial_call=True
    )
    def reset_weights(n_clicks, selected_columns):
        if not n_clicks or not selected_columns:
            return dash.no_update, dash.no_update
        n = len(selected_columns)
        equal_weight = round(100 / n, 2)
        return [equal_weight] * n, [0] * n

    # 4. Composite Manager Table
    @app.callback(
        Output('composite-manager-container', 'children'),
        Input('user-session-id', 'data'),
        Input('layer-manager-trigger', 'data')
    )
    def render_composite_manager(session_id, trigger):
        if not session_id:
            return html.Div("No composite indices created yet.",
                          style={'fontSize':'10px', 'color':'#bbb', 'fontStyle':'italic', 'padding':'10px'})
        indices = dm.get_composite_indices(session_id) if hasattr(dm, 'get_composite_indices') else {}
        if not indices:
            return html.Div("No composite indices created yet.",
                          style={'fontSize':'10px', 'color':'#bbb', 'fontStyle':'italic', 'padding':'10px'})
        lang = app.lang
        rows = []
        for idx_name, idx_info in indices.items():
            components = idx_info.get('components', [])
            formula_parts = []
            for comp in components:
                sign = "-" if comp.get('negative', False) else "+"
                col_translated = get_column_translation(comp['column'], lang, COUNTRY_CODE)
                formula_parts.append(f"{sign} {col_translated}: {comp.get('original_weight', 0):.1f}%")
            formula_text = "\n".join(formula_parts) if formula_parts else "N/A"
            formula_preview = formula_parts[0] if formula_parts else ""
            safe_id = idx_name.replace(' ', '-').replace("'", "")
            rows.append(
                html.Div([
                    html.Div([
                        html.B(idx_name, style={'fontSize':'11px', 'color':'#FBB800'}),
                        dbc.Tooltip(formula_text, target=f"idx-name-{safe_id}", placement="right")
                    ], id=f"idx-name-{safe_id}", style={'flex':'2', 'padding':'5px'}),
                    html.Div(formula_preview,
                        style={'flex':'3', 'fontSize':'9px', 'color':'#ecf0f1', 'padding':'5px', 'fontFamily':'monospace'}),
                    html.Div([
                        html.Button(html.I(className="fas fa-trash"),
                            id={'type': 'delete-index-btn', 'index': idx_name}, n_clicks=0, title="Delete",
                            style={'padding':'3px 6px', 'fontSize':'10px', 'backgroundColor':'#e74c3c',
                                   'color':'white', 'border':'none', 'borderRadius':'3px', 'cursor':'pointer'})
                    ], style={'flex':'1', 'padding':'5px', 'textAlign':'right'})
                ], style={'display':'flex', 'alignItems':'center', 'backgroundColor':'#333',
                         'marginBottom':'5px', 'borderRadius':'4px', 'border':'1px solid #555'})
            )
        return rows

    # 4b. Delete Composite Index
    @app.callback(
        Output('layer-manager-trigger', 'data', allow_duplicate=True),
        Output('layer-data-content-trigger', 'data', allow_duplicate=True),
        Input({'type': 'delete-index-btn', 'index': ALL}, 'n_clicks'),
        State({'type': 'delete-index-btn', 'index': ALL}, 'id'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def delete_composite_index(n_clicks_list, button_ids, session_id):
        ctx = callback_context
        if not ctx.triggered or not session_id:
            return dash.no_update, dash.no_update
        triggered_id = ctx.triggered[0]['prop_id']
        if 'delete-index-btn' not in triggered_id:
            return dash.no_update, dash.no_update
        for nc, bid in zip(n_clicks_list, button_ids):
            if nc and nc > 0:
                idx_name = bid['index']
                if hasattr(dm, 'delete_composite_index'):
                    dm.delete_composite_index(session_id, idx_name)
                trigger = str(uuid.uuid4())
                return trigger, trigger
        return dash.no_update, dash.no_update

    # 5. Calculate Composite Index
    @app.callback(
        Output('composite-status-message', 'children'),
        Output('composite-status-message', 'style'),
        Output('layer-manager-trigger', 'data', allow_duplicate=True),
        Output('layer-data-content-trigger', 'data', allow_duplicate=True),
        Input('composite-calculate-btn', 'n_clicks'),
        State('composite-name-input', 'value'),
        State('composite-layer-dropdown', 'value'),
        State('composite-columns-dropdown', 'value'),
        State({'type': 'weight-slider', 'index': ALL}, 'value'),
        State({'type': 'weight-negative', 'index': ALL}, 'n_clicks'),
        State({'type': 'weight-column-name', 'index': ALL}, 'data'),
        State('composite-normalization-method', 'value'),
        State('user-session-id', 'data'),
        prevent_initial_call=True
    )
    def run_composite_calculation(n_clicks, column_name, layer_id, selected_columns,
                                  weights_list, negative_clicks, column_names, method, session_id):
        err_style = {'fontSize':'10px', 'padding':'8px', 'borderRadius':'4px',
                     'backgroundColor':'rgba(231,76,60,0.2)', 'color':'#e74c3c', 'display':'block'}
        ok_style = {'fontSize':'10px', 'padding':'8px', 'borderRadius':'4px',
                    'backgroundColor':'rgba(46,204,113,0.2)', 'color':'#2ecc71', 'display':'block'}
        hide = {'display':'none'}

        if not n_clicks or not session_id:
            return "", hide, dash.no_update, dash.no_update
        if not column_name or not column_name.strip():
            return "Please enter a name for the new column.", err_style, dash.no_update, dash.no_update
        if not layer_id:
            return "Please select a target layer.", err_style, dash.no_update, dash.no_update
        if not selected_columns or len(selected_columns) < 2:
            return "Please select at least 2 columns.", err_style, dash.no_update, dash.no_update
        if not weights_list or not column_names:
            return "Weight configuration error.", err_style, dash.no_update, dash.no_update

        total = sum(weights_list)
        if abs(total - 100) > 0.5:
            return f"Weights must sum to 100% (current: {total:.2f}%).", err_style, dash.no_update, dash.no_update

        df = dm.get_df(session_id, layer_id)
        if df is not None and column_name in df.columns:
            return f"Column '{column_name}' already exists. Choose a different name.", err_style, dash.no_update, dash.no_update

        weights = {col_name: weight for col_name, weight in zip(column_names, weights_list)}
        negative_effects = {col_name: (clicks % 2 == 1 if clicks else False)
                           for col_name, clicks in zip(column_names, negative_clicks or [])}

        try:
            success = dm.create_composite_index(
                session_id=session_id, layer_id=layer_id, column_name=column_name,
                selected_columns=selected_columns, weights=weights,
                method=method, negative_effects=negative_effects
            )
        except Exception as e:
            print(f"[Composite] ERROR creating index: {e}")
            return f"Error: {str(e)}", err_style, dash.no_update, dash.no_update

        if success:
            if hasattr(dm, 'get_user_data'):
                user_data = dm.get_user_data(session_id)
                components = []
                for col_name, weight in zip(column_names, weights_list):
                    components.append({
                        'column': col_name, 'weight': weight / 100.0,
                        'negative': negative_effects.get(col_name, False),
                        'original_weight': weight
                    })
                user_data.setdefault('composite_indices', {})[column_name] = {
                    'layer_id': layer_id, 'components': components, 'method': method
                }

            # Auto-set map color to new composite index
            user_meta = dm.get_metadata(session_id)
            layer_meta = next((m for m in user_meta if m['id'] == layer_id), None)
            if layer_meta:
                layer_meta['color_column'] = column_name
                layer_meta['color_mode'] = 'column'

            trigger = str(uuid.uuid4())
            return f"Created '{column_name}'. Map now colored by this index.", ok_style, trigger, trigger
        else:
            return "Failed to create composite index.", err_style, dash.no_update, dash.no_update

    return app

# --- FACTORY FUNCTION (used by wsgi_protected_auth.py for web deployment) ---
def create_app(lang='en', url_base_pathname=None):
    """
    Load data and create the Mozambique app for a specific language.
    Used by the WSGI dispatcher for web deployment.

    Args:
        lang: Language code ('en' or 'pt')
        url_base_pathname: Dash requests_pathname_prefix (e.g. '/moz/en/').
                           Defaults to f'/moz/{lang}/' if not provided.

    Returns:
        Flask server instance
    """
    if url_base_pathname is None:
        url_base_pathname = f'/moz/{lang}/'

    import glob
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    folder = os.path.join(project_root, DATA_FOLDER)
    candidates = sorted(glob.glob(os.path.join(folder, '*.gpkg')), key=os.path.getmtime, reverse=True)
    if not candidates:
        raise RuntimeError(f"No .gpkg file found in {folder}")
    data_path = candidates[0]
    print(f"[create_app] Loading dataset: {os.path.relpath(data_path, project_root)}")

    gdf = gpd.read_file(data_path)
    app = agcap_explorer(gdf, default_column=DEFAULT_COLUMN,
                         figure_title=FIGURE_TITLE, lang=lang,
                         url_base_pathname=url_base_pathname)
    return app.server


if __name__ == '__main__':
    from werkzeug.serving import run_simple
    server = create_app('en', url_base_pathname='/')
    run_simple('localhost', 8050, server, use_reloader=True, use_debugger=True)